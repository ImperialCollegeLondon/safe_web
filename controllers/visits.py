import datetime
import openpyxl
from collections import Counter

## -----------------------------------------------------------------------------
## Visit controllers
## - controllers to view and edit a user's H&S form and for admin to view records
## - controllers to book accomodation at SAFE camp
## - controllers to approve bookings and manage bed reservations at SAFE camp
##   including an admin interface for creating 'special' bookings
## - controllers to approve field research at SAFE
## - TODO figure out the hell to make the add members options part of the same
##        form as the proposal.
## -----------------------------------------------------------------------------


## -----------------------------------------------------------------------------
## HEALTH AND SAFETY
## -- TODO - how hardline do we want to be about completeness of fields
## -----------------------------------------------------------------------------

@auth.requires_login()
def health_and_safety():
    
    """
    provides access to the health and safety information 
    for the logged in user, both edit and create use the same interface
    """

    # Restrict the user 
    db.health_and_safety.user_id.requires = IS_IN_DB(db(db.auth_user.id == auth.user.id),
                                                     db.auth_user.id, '%(first_name)s %(last_name)s')
    
    # lock the widget for user_id
    db.health_and_safety.user_id.writable = False
    db.health_and_safety.date_last_edited.readable = False
    db.health_and_safety.date_last_edited.writable = False
    
    # look for an existing record, otherwise initialise a blank one
    # - this is a bit of a hack. The DB backend accepts the empty strings
    #   but conveniently the SQLFORM validation doesn't, so you can't just
    #   click through with the blank form 
    rows = db(db.health_and_safety.user_id == auth.user.id).select()
    if len(rows) > 0:
        record = rows.first().id
    else:
        record = db.health_and_safety.insert(user_id = auth.user.id, 
                                             passport_number = '',
                                             emergency_contact_name = '',
                                             emergency_contact_address = '',
                                             emergency_contact_phone = '',
                                             emergency_contact_email = '',
                                             insurance_company = '',
                                             insurance_emergency_phone = '',
                                             insurance_policy_number = '')
    
    # get the form with the existing or new record
    form = SQLFORM(db.health_and_safety, record=record, 
                   showid=False, labels={'user_id': 'Name'})
    
    # now intercept and parse the various inputs
    if form.process(onvalidation=validate_health_and_safety).accepted:
        # insert the h&s record id into the user table 
        #- this field is primarily to avoid a lookup to just to 
        #  populate links from visit and reservation details pages
        db(db.auth_user.id == auth.user.id).update(h_and_s_id = form.vars.id)
        session.flash = CENTER(B('Thanks for providing your health and safety information.'), _style='color: green')
        redirect(URL('default', 'index'))
    elif form.errors:
        response.flash = CENTER(B('Errors in form, please check and resubmit'), _style='color: red')
    else:
        pass
    
    return dict(form=form)


def validate_health_and_safety(form):
    
    """
    Pretty minimal - currently just updates the date edited
    """
    form.vars.date_last_edited = datetime.date.today().isoformat()

@auth.requires_membership('admin')
def admin_view_health_and_safety():
    
    """
    provides access to the health and safety information for admin
    """
    
    user_id = request.args(0)
        
    # look for an existing record
    rows = db(db.health_and_safety.user_id == user_id).select()
    if len(rows) > 0:
        # get the form with the existing or new record
        form = SQLFORM(db.health_and_safety, record=user_id, 
                       showid=False, labels={'user_id': 'Name'},
                       readonly=True)
    else:
        form = None
    return dict(form=form)


## -----------------------------------------------------------------------------
## BED RESERVATIONS
## -- controllers to book blocks of beds and administer those bookings
## -- TODO add named beds (registered users only or just names)
## -----------------------------------------------------------------------------

def bed_availability():

    """
    This controller:
        - creates data for a free beds view using fullcalendar javascript
        - combining this and the booking on a single page causes issues
    """

    # get the dates when beds are booked
    # and select as an iterable of rows
    bed_data = db(db.bed_data)
    rows = bed_data.select()
    
    # Pass a list of events to the view for the javascript calendar
    # - need to handle null values from db
    pend = [0 if r.pending is None else r.pending for r in rows]
    conf = [0 if r.approved is None else r.approved for r in rows]
    avail = [n_beds_available - (x+y) for x, y in zip(pend, conf)]
    # handle admin approved overbooking by trunctating
    avail = [0 if x < 0 else x for x in avail]
    date = [r.day.isoformat() for r in rows]
    
    # now zip into sets of events, with three per day
    # one for each of pending confirmed and available
    n_events = len(date)
    event_n_beds =pend + conf + avail
    event_class  = ['pending'] * n_events + ['confirmed'] * n_events + ['available'] * n_events
    event_title  = [ str(x) + ' ' + y for x,y in zip(event_n_beds, event_class)]
    event_order  = [2] * n_events + [3] * n_events + [1] * n_events
    # pass colour information
    event_backgrounds = ['#CC9900'] * n_events + ['#CC0000'] * n_events + ['#228B22'] * n_events
    
    events = zip(event_title, date * 3, event_class, event_order, event_backgrounds)
    
    return dict(events=events)

@auth.requires_login()
def reserve_beds():

    """
    This controller:
        - creates an SQLFORM form to submit a booking
        - handles responses to errors and success
        - adds bookings to google calendar (TODO - move to confirmed v pending)
        - emails booker (TODO - turned off)
    """
    
    form = SQLFORM(db.bed_reservations, 
                   fields=['research_visit_id', 'arrival_date',
                           'departure_date', 'number_of_visitors',
                           'purpose', 'look_see_visit'],
                   labels = {'research_visit_id':'Approved research visit',
                             'look_see_visit':'These bed bookings are just to explore '
                             'the facilities and research opportunties at SAFE and are '
                             'not associated with an ongoing research project.'},
                   submit_button = 'Submit and add names')
    
    if form.process(onvalidation=validate_reserve_beds).accepted:
        
        # # Mechanism for sending email to logged in user
        # mail.send(to=auth.user.email,
        #           subject='hello',
        #           message='hi there')
        
        # update the bed bookings table
        # - get a list of days to book
        days_occupied = day_range(form.vars.arrival_date, form.vars.departure_date)
        # - check those days are in the bed bookings table
        for d in days_occupied:
            row = db.bed_data(day=d)
            if row is None:
                db.bed_data.insert(day=d, pending=form.vars.number_of_visitors)
            else:
                row.update_record(pending=row.pending + form.vars.number_of_visitors)
        
        # add the proposer to the reservation names list table
        db.bed_reservation_member.insert(bed_reservation_id = form.vars.id,
                                         user_id = auth.user.id)
        
        session.flash = CENTER(B('Bed reservation request submitted.'), _style='color: green')
        print URL('visits','bed_reservation_details', args=form.vars.id)
        
        redirect(URL('visits','bed_reservation_details', args=form.vars.id))
        print 'I am here too'
    
    elif form.errors:
        response.flash = CENTER(B('Errors in form, please check and resubmit'), _style='color: red') 
    else:
        pass
    
    return dict(form=form)


def validate_reserve_beds(form):
    
    # add the user id into the response
    form.vars.reserver_id = auth.user_id
    form.vars.reservation_date =  datetime.date.today().isoformat()
    
    # check the arrival date is more than a fortnight away
    deadline = datetime.date.today() + datetime.timedelta(days=14)
    if form.vars.arrival_date < deadline:
        form.errors.arrival_date = '14 days notice required. Arrival date must be later than {}.'.format(deadline.isoformat())
    
    # check the departure date is after the arrival date
    # TODO - think about day visits
    if form.vars.arrival_date >= form.vars.departure_date:
        form.errors.departure_date = 'The departure date must be later than the arrival date'
    
    # can't be not a look see and not have a visit id - oddly the boolean checkbox widget
    # returns None when not checked...
    if form.vars.look_see_visit in [False, None] and form.vars.research_visit_id is None:
        form.errors.research_visit_id = ('Unless you just want to evaluate research '
                                        'opportunities at SAFE (see check box below), '
                                        'you must specify an approved research visit. ' 
                                        'This menu will only give access to current approved '
                                        'visits for projects of which you are a member.')
    
    # Check there are enough beds for the visit
    # - get a range of days
    days_requested = day_range(form.vars.arrival_date, form.vars.departure_date)
    rows = db(db.bed_data.day.belongs(days_requested)).select()
    pend = [0 if r.pending is None else r.pending for r in rows]
    conf = [0 if r.approved is None else r.approved for r in rows]
    avail = [n_beds_available - (x+y) for x, y in zip(pend, conf)]
    can_fit_request = [x >= form.vars.number_of_visitors for x in avail]
    if not all(can_fit_request):
        form.errors.number_of_visitors = ('There are not enough available beds to cover'
                                          ' the requested visit. There is currently a maximum of ' +
                                          str(min(avail)) + ' beds available throughout this '
                                          'period, although more may become available if some '
                                          'pending reservations are rejected.')


def day_range(start, end):
    
    days = []
    curr = start
    while curr < end: # don't assign a bed on the departure date, so not <=
        days.append(curr)
        curr += datetime.timedelta(days=1)
    
    return(days)


@auth.requires_login()
def bed_reservation_details():
    
    """
    Controller to show the details of a bed reservatoin
     - collects a bunch of information and passes it to the view for processing
     - allows members to be added
    """
    
    # get the reservation_id from the call
    reservation_id = request.args(0)
    
    # get the reservation information
    reservation_record = db.bed_reservations(reservation_id)
    
    if reservation_record is None:
        redirect(URL('default','index'))
    
    # get current rows
    these_rows = db.bed_reservation_member.bed_reservation_id == reservation_id
    
    # and hence members
    reservation_members = db(these_rows).select(db.bed_reservation_member.user_id)
    
    # look up the visit and project details if there are any
    if reservation_record.look_see_visit is False:
        visit_record = db.research_visit(reservation_record.research_visit_id)
    else:
        visit_record = None
    
    # only provide active form to project members - can add themselves
    # although there could be the idea of a coordinator
    # - to implement
    authorised_users = [None]
    
    if auth.user.id in authorised_users:
        # Offer choices of new members to add
        # - find users belonging to the project that this visit is for 
        #   and which are not already members of the visit
        already_users = db(these_rows)._select(db.research_visit_member.user_id)
        which_users = ((db.project_members.project_id == visit_record.project_id) &
                       (~ db.project_members.user_id.belongs(already_users)))
        valid_members = db(which_users)._select(db.project_members.user_id)
        query = db(db.auth_user.id.belongs(valid_members))
    
        # - require the visit members
        db.research_visit_member.user_id.requires = IS_IN_DB(query, db.auth_user.id, '%(last_name)s, %(first_name)s')
    
        # lock down the possible value of research_visit_id
        db.research_visit_member.research_visit_id.default = research_visit_id
        db.research_visit_member.research_visit_id.readable = False
        
        # links for H&S status
        hs_done_icon = SPAN('',_class="glyphicon glyphicon-list-alt", 
                             _style="color:green;font-size: 1.3em;", 
                             _title='H&S Completed')
        hs_none_icon = SPAN('',_class="glyphicon glyphicon-list-alt",
                             _style="color:red;font-size: 1.3em;", 
                             _title='H&S missing')
        
        # create the links to the standalone controllers
        links = [dict(header = '', 
                      body = lambda row: hs_done_icon )]
        #              body = lambda row: approved_icon if row.admin_status == 'Approved' else pending_icon)]
        
        form = SQLFORM(db.research_visit_member,
                       fields = ['user_id'],
                       links = links,
                       labels = {'user_id':'Project member to add'})
    
        # process and reload the page
        if form.process().accepted:
            redirect(URL('visits', 'research_visit_details', args=research_visit_id))
    else:
        form = None
    return dict(reservation_record = reservation_record, visit_record = visit_record,
                reservation_members = reservation_members, form=form)


# decorator restricts access to admin users
# - the link is only revealed in the menu for admin users but 
#   that doesn't prevent pasting in the link!
@auth.requires_membership('admin')
def administer_reserve_beds():
    
    """
    This controller handles:
     - presenting admin users with a list of current visits from the db
     - allowing the admin to approve or reject visit requests
     - Currently this locks all but the admin fields. Might also need
       a separate admin visit booking form with fewer restrictions for
       overbooking and special cases?
    """
    
    # don't want the admin to change any of this about a visit
    db.bed_reservations.research_visit_id.writable = False
    db.bed_reservations.reserver_id.writable = False
    db.bed_reservations.arrival_date.writable = False
    db.bed_reservations.departure_date.writable = False
    db.bed_reservations.number_of_visitors.writable = False
    db.bed_reservations.purpose.writable = False
    db.bed_reservations.look_see_visit.writable = False
    
    # get a query of pending requests with user_id
    form = SQLFORM.grid(query=(db.bed_reservations.admin_status == 'Pending'), csv=False,
                        fields=[db.bed_reservations.arrival_date, 
                                db.bed_reservations.departure_date,
                                db.bed_reservations.number_of_visitors],
                         maxtextlength=250,
                         deletable=False,
                         editable=True,
                         create=False,
                         details=False,
                         editargs = {'fields': ['research_visit_id','reserver_id',
                                                'arrival_date','departure_date',
                                                'number_of_visitors','purpose', 'look_see_visit',
                                                'admin_status','admin_notes'],
                                     'showid': False},
                         onvalidation = validate_administer_reserve_beds,
                         onupdate = update_administer_reserve_beds)
    
    return dict(form=form)

@auth.requires_membership('admin')
def validate_administer_reserve_beds(form):
    
    # validation handles any checking (none here) and also any 
    # amendments to the form variable  - adding user and date of admin
    form.vars.admin_id = auth.user_id
    form.vars.admin_decision_date =  datetime.date.today().isoformat()
    
    return(form)

@auth.requires_membership('admin')
def update_administer_reserve_beds(form):

    """
    This handler updates the bed_data table to remove or confirm beds 
    on each day and then emails the person who made the reservation
    """
    
    # recover the record to get all the stuff that isn't writable
    record = form.record
    
    # get the dates when beds are pending booked
    # and select as an iterable of rows
    days_requested = day_range(record.arrival_date, record.departure_date)
    for d in days_requested:
        row = db.bed_data(day=d)
        if(form.vars.admin_status == 'Approved'):
            row.update_record(pending=row.pending - record.number_of_visitors,
                              approved=(0 if row.approved is None else row.approved) + record.number_of_visitors)
        else:
            row.update_record(pending= row.pending - record.number_of_visitors)
    
    # email the reserver
    if(form.vars.admin_status == 'Approved'):
        
        # email the proposer
        mail.send(to=record.reserver_id.email,
              subject='hello',
              message='Welcome to SAFE')
    
    elif(form.vars.admin_status == 'Rejected'):
        
        mail.send(to=record.reserver_id.email,
              subject='hello',
              message='Nope')
    else:
        pass
    
    return dict(form=form)




@auth.requires_membership('admin')
def manage_bed_reservations():
    
    """
    This controller provides a grid view of approved bookings and allows
    the admin to create new bookings and delete existing ones. Having function
    to _edit_ bookings is probably necessary but would have to track changes in
    date and number of beds, so for the moment just cancel and rebook.
    """
    
    # admins can only make Approved bookings
    db.bed_reservations.admin_status.default = 'Approved'
    db.bed_reservations.admin_status.writable = False
    
    form = SQLFORM.grid(query=(db.bed_reservations.admin_status == 'Approved'),
                        fields = [db.bed_reservations.arrival_date,
                                  db.bed_reservations.departure_date,
                                  db.bed_reservations.number_of_visitors,
                                  db.bed_reservations.reserver_id],
                        create = True,
                        editable = False, 
                        deletable = True,
                        formargs = {'fields': ['research_visit_id', 'arrival_date',
                                               'departure_date','number_of_visitors',
                                               'purpose','look_see_visit','reserver_id',
                                               'admin_status', 'admin_notes'],
                                    'showid': False},
                        csv = False,
                        onvalidation = validate_administer_reserve_beds,
                        oncreate = create_manage_bed_reservations,
                        ondelete = delete_manage_bed_reservations)
    
    return dict(form=form)


@auth.requires_membership('admin')
def create_manage_bed_reservations(form):

    """
    This handler updates the bed_data table to add beds 
    on each day and then emails the person who made the reservation
    but unlike the update_reserve, it know it is going to get approved posts
    and it doesn't get the record but the form.vars
    """
        
    # get the dates of the booking
    days_requested = day_range(form.vars.arrival_date, form.vars.departure_date)
    for d in days_requested:
        row = db.bed_data(day=d)
        if row is None:
            db.bed_data.insert(day = d, approved = form.vars.number_of_visitors)
        else:
            row.update_record(approved=(0 if row.approved is None else row.approved) + form.vars.number_of_visitors)
    
    # email the proposer
    reserver_email = db.auth_user(form.vars.reserver_id).email
    mail.send(to=reserver_email,
              subject='Admin bed reservation at SAFE',
              message='Welcome to SAFE')
    
    response.flash = CENTER(B('Admin bed booking made.'), _style='color: green')



@auth.requires_membership('admin')
def delete_manage_bed_reservations(table, id):

    """
    This handler removes a record and gets rid of the bed data and emails the original reserver
    """
        
    record = table(id)
    
    # get the dates of the booking
    days_requested = day_range(record.arrival_date, record.departure_date)
    for d in days_requested:
        row = db.bed_data(day=d)
        row.update_record(approved=row.approved - record.number_of_visitors)
    
    # email the proposer
    mail.send(to=record.reserver_id.email,
              subject='Your bed reservation at SAFE has been cancelled',
              message='Unlucky')
    
    response.flash = CENTER(B('Bed booking cancelled.'), _style='color: red')


def export_bed_reservation_workbook():
    
    """
    This creates an excel workbook showing the current bed reservation data
    and pokes it out as an http download, so a button can call the controller
    and return the file via the browser. Formatting needs some work but Ryan
    was keen on this.
    """
    
    query = ((db.bed_reservations.admin_status == 'Approved') &
             (db.bed_reservations.departure_date > datetime.date.today().isoformat()))
    data = db(query).select(orderby=db.bed_reservations.arrival_date)
    
    # open the workbook object
    wb = openpyxl.Workbook()
    # get the worksheet
    ws = wb.active
    ws.title = 'Bed reservations'
    ws['A1'] = 'Bed Reservations at SAFE as of {}'.format(datetime.date.today().isoformat())
    
    # get the arrival date of the first row (which is the earliest, due to the orderby)
    end_date = start_date = data.first().arrival_date
    
    current_row = start_row = 6
    last_col = start_col = 1
    
    fill1 = openpyxl.styles.PatternFill(patternType='solid', 
                                        fgColor=openpyxl.styles.Color('55BBBBBB'))
    fill2 = openpyxl.styles.PatternFill(patternType='solid', 
                                        fgColor=openpyxl.styles.Color('FFDDDDDD'))
    
    for i, row in enumerate(data):
        this_start_col = (row.arrival_date - start_date).days + start_col 
        this_end_col = (row.departure_date - start_date).days + start_col 
        ws.merge_cells(start_row=current_row, 
                       start_column=this_start_col,
                       end_row=current_row + row.number_of_visitors - 1 ,
                       end_column=this_end_col)
        if last_col < this_end_col:
            last_col = this_end_col
            end_date = row.departure_date
        c = ws.cell(row=current_row, column=(row.arrival_date - start_date).days + start_col)
        c.style.alignment.vertical = "top"
        c.value = row.purpose
        c.fill = fill1 if (i % 2 == 0) else fill2
        current_row  += row.number_of_visitors 
    
    # label up the dates
    dates = day_range(start_date, end_date + datetime.timedelta(days=1))
    weekday = [d.strftime('%a') for d in dates]
    month = [d.strftime('%B') for d in dates]
    day = [d.day for d in dates]
    
    for i, (d, w) in enumerate(zip(day, weekday)):
        c = ws.cell(row=start_row - 1, column=start_col + i)
        c.value = d
        c = ws.cell(row=start_row - 2, column=start_col + i)
        c.value = w
    
    current_col = start_col
    while len(month) > 0:
        m_counter = Counter(month)
        this_month = month[0]
        n_month = m_counter[this_month] 
        ws.merge_cells(start_row=start_row - 3, 
                       start_column=current_col,
                       end_row=start_row - 3,
                       end_column=current_col + n_month - 1)
        c = ws.cell(row=start_row - 3, column=current_col)
        c.value = this_month
        current_col = current_col + n_month
        month = month[n_month:]
    
    for i in range(start_col, last_col + 1):
        ws.column_dimensions[openpyxl.cell.get_column_letter(i)].width = '3'
    
    # and now poke it out to the browser
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    attachment = 'attachment;filename=SAFE_Bed_reservations_{}.xlsx'.format(datetime.date.today().isoformat())
    response.headers['Content-Disposition'] = attachment
    content = openpyxl.writer.excel.save_virtual_workbook(wb)
    raise HTTP(200, str(content),
               **{'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                  'Content-Disposition':attachment + ';'})


## -----------------------------------------------------------------------------
## RESEARCH VISITS
## -- controllers to register research visits and administer those visits
## -- provide a general grid view of visits for users and a detail view that
##    allows project members to add members to visits
## -----------------------------------------------------------------------------


@auth.requires_login()
def research_visits():
    
    """
    This controller shows the grid view for visits and allows
    logged in users to view details

    It uses a custom button to divert from the SQLFORM.grid view to a
    custom view page that allows project members to add visitors
    """

    # For standard users (need a separate admin projects controller)
    # don't show the authorization fields and don't show a few behind 
    # the scenes fields
    db.research_visit.admin_id.readable = False 
    db.research_visit.admin_notes.readable = False 
    db.research_visit.admin_decision_date.readable = False 
    
    # create a links list that:
    # 1) displays a thumbnail of the  project image
    # 2) creates a custom button to pass the row id to a custom view 

    links = [dict(header = '', body = lambda row: A('View',_class='button btn btn-default'
                  ,_href=URL("visits","research_visit_details", args=[row.id])))
            ]

    form = SQLFORM.grid(db.research_visit, csv=False, 
                        fields=[db.research_visit.project_id, db.research_visit.title, 
                                db.research_visit.arrival_date, db.research_visit.departure_date],
                        maxtextlength=250,
                        deletable=False,
                        editable=False,
                        create=False, 
                        details=False,
                        links=links,
                        links_placement='left',
                        formargs={'showid': False})

    return dict(form=form)



@auth.requires_login()
def new_research_visit():
    
    """
    This controller shows a SQLFORM to propose a research visit. The controller then
    passes the response through validation before sending a confirmation email out.
    """
    
    # Restrict the project choices available
    # - find projects that the logged in user is a member of
    valid_ids = db(db.project_members.user_id == auth.user.id)._select(db.project_members.project_id)
    query = db(db.project.id.belongs(valid_ids))
    
    # - modify the project_id requirements within this controller to only show those projects
    db.research_visit.project_id.requires = IS_EMPTY_OR(IS_IN_DB(query, db.project.id, '%(title)s'))
    
    # Help the user by explaining why project ID might be blank
    if query.count() == 0:
        message = CENTER(B('Sorry, you must have been added as a member of a project'
                           ' to submit a research visit proposal'), _style='color: red')
    else:
        message = None
    
    form = SQLFORM(db.research_visit,
                   fields = ['project_id','title','arrival_date',
                             'departure_date','purpose','licence_details'],
                   labels = {'title':'Reference name for visit'},
                   submit_button = 'Submit and add visit members')
    
    # now intercept and parse the various inputs
    if form.process(onvalidation=validate_new_research_visit).accepted:
        
        # Signal success and email the proposer
        mail.send(to=auth.user.email,
           subject='SAFE research visit proposal submitted',
           message='Many thanks for proposing your research visit')
        response.flash = CENTER(B('SAFE research proposal successfully submitted.'), _style='color: green')
        
        # add the proposer to the visit member table
        db.research_visit_member.insert(research_visit_id = form.vars.id,
                                        user_id = auth.user.id)
        
        redirect(URL('visits','research_visit_details', args=form.vars.id))
        
    elif form.errors:
        response.flash = CENTER(B('Errors in form, please check and resubmit'), _style='color: red')
    else:
        pass
    
    return dict(form=form, message=message)


def validate_new_research_visit(form):
    
    form.vars.proposer_id = auth.user_id
    form.vars.proposal_date = datetime.date.today().isoformat()
    
    # check the arrival date is more than a fortnight away
    deadline = datetime.date.today() + datetime.timedelta(days=14)
    if form.vars.arrival_date < deadline:
        form.errors.arrival_date = '14 days notice required. Arrival date must be later than {}.'.format(deadline.isoformat())
    
    # check the departure date is after the arrival date
    # TODO - think about day visits
    if form.vars.arrival_date >= form.vars.departure_date:
        form.errors.departure_date = 'The departure date must be later than the arrival date'
    

@auth.requires_login()
def research_visit_details():
    
    """
    Controller to show the details of research visits
     - collects a bunch of information and passes it to the view for processing
     - allows members to be added (but this needs to be locked down)
    """
    
    # get the visit_id from the call
    research_visit_id = request.args(0)
    
    # get the visit information
    visit_record = db.research_visit(research_visit_id)
    
    if visit_record is None:
        redirect(URL('default','index'))
    
    # get current rows
    these_rows = db.research_visit_member.research_visit_id == research_visit_id
    # and hence members
    visit_members = db(these_rows).select(db.research_visit_member.user_id)
    
    # only provide active form to project members - can add themselves
    # although there could be the idea of a visit coordinator
    project_members = db(db.project_members.project_id == visit_record.project_id).select()
    project_member_ids = [r.user_id for r in project_members]
    
    if auth.user.id in project_member_ids:
        # Offer choices of new members to add
        # - find users belonging to the project that this visit is for 
        #   and which are not already members of the visit
        already_users = db(these_rows)._select(db.research_visit_member.user_id)
        which_users = ((db.project_members.project_id == visit_record.project_id) &
                       (~ db.project_members.user_id.belongs(already_users)))
        valid_members = db(which_users)._select(db.project_members.user_id)
        query = db(db.auth_user.id.belongs(valid_members))
    
        # - require the visit members
        db.research_visit_member.user_id.requires = IS_IN_DB(query, db.auth_user.id, '%(last_name)s, %(first_name)s')
    
        # lock down the possible value of research_visit_id
        db.research_visit_member.research_visit_id.default = research_visit_id
        db.research_visit_member.research_visit_id.readable = False
    
        form = SQLFORM(db.research_visit_member,
                       fields = ['user_id'],
                       labels = {'user_id':'Project member to add'})
    
        # process and reload the page
        if form.process().accepted:
            redirect(URL('visits', 'research_visit_details', args=research_visit_id))
    else:
        form = None
    return dict(visit_record = visit_record, visit_members = visit_members, form=form)

# decorator restricts access to admin users
# - the link is only revealed in the menu for admin users but 
#   that doesn't prevent pasting in the link!
@auth.requires_membership('admin')
def administer_new_research_visits():
    
    """
    This controller handles:
     - presenting admin users with a list of current proposals for research visits 
     - allowing the admin to approve or reject visit requests
     - Currently this locks all but the admin fields. Might also need
       a separate admin visit booking form with fewer restrictions for
       overbooking and special cases?
    """
    
    # don't want the admin to change any of this about a visit
    db.research_visit.project_id.writable = False
    db.research_visit.title.writable = False
    db.research_visit.arrival_date.writable = False
    db.research_visit.departure_date.writable = False
    db.research_visit.writable = False
    db.research_visit.purpose.writable = False
    
    # get a query of pending requests with user_id
    form = SQLFORM.grid(query=(db.research_visit.admin_status == 'Pending'), csv=False,
                        fields=[db.research_visit.project_id, 
                                db.research_visit.title],
                         maxtextlength=250,
                         deletable=False,
                         editable=True,
                         create=False,
                         details=False,
                         editargs = {'fields': ['project_id','title',
                                                'arrival_date','departure_date',
                                                'purpose', 'admin_status','admin_notes'],
                                     'showid': False},
                         onvalidation = validate_administer_new_research_visits,
                         onupdate = update_administer_new_research_visits)
    
    return dict(form=form)

@auth.requires_membership('admin')
def validate_administer_new_research_visits(form):
    
    # validation handles any checking (none here) and also any 
    # amendments to the form variable  - adding user and date of admin
    form.vars.admin_id = auth.user_id
    form.vars.admin_decision_date =  datetime.date.today().isoformat()


@auth.requires_membership('admin')
def update_administer_new_research_visits(form):

    """
    This handler emails the person who made the proposal
    """
    
    # recover the record to get all the stuff that isn't writable
    record = form.record
        
    # email the reserver
    if(form.vars.admin_status == 'Approved'):
        
        # email the proposer
        mail.send(to=record.proposer_id.email,
              subject='SAFE research visit proposal accepted',
              message='Welcome to SAFE')
        session.flash = CENTER(B('Approval email sent to visit proposer'), _style='color: green')
    
    elif(form.vars.admin_status == 'Rejected'):
        
        mail.send(to=record.proposer_id.email,
              subject='SAFE research visit proposal rejected',
              message='Nope')
        session.flash = CENTER(B('Rejection email sent to visit proposer'), _style='color: green')
    else:
        pass
    
    return dict(form=form)
