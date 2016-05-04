import datetime
import openpyxl
from collections import Counter

## -----------------------------------------------------------------------------
## BED RESERVATIONS
## -- controllers to book blocks of beds and administer those bookings
## -- controllers to book accomodation at SAFE camp
## -- controllers to approve bookings and manage bed reservations at SAFE camp
##    including an admin interface for creating 'special' bookings
## -- TODO Named beds - registered users only or just names?
## -----------------------------------------------------------------------------

def safe_availability():

    """
    This controller:
        - creates data for a free beds view using fullcalendar javascript
        - combining this and the booking on a single page causes issues
    """

    # get the dates when beds are booked
    # and select as an iterable of rows
    bed_data = db(db.bed_reservations.site =='safe').select()
    
    # Pass a list of events to the view for the javascript calendar
    # - need to handle null values from db
    pend = [0 if r.pending is None else r.pending for r in rows]
    conf = [0 if r.approved is None else r.approved for r in rows]
    avail = [n_beds_available - (x+y) for x, y in zip(pend, conf)]
    # handle admin approved overbooking by trunctating
    avail = [0 if x < 0 else x for x in avail]
    date = [r.day.isoformat() for r in rows]
    
    # now zip into sets of events, with three per day
    # one for each of pending confirmed and available
    n_events = len(date)
    event_n_beds =pend + conf + avail
    event_class  = ['pending'] * n_events + ['confirmed'] * n_events + ['available'] * n_events
    event_title  = [ str(x) + ' ' + y for x,y in zip(event_n_beds, event_class)]
    event_order  = [2] * n_events + [3] * n_events + [1] * n_events
    # pass colour information
    event_backgrounds = ['#CC9900'] * n_events + ['#CC0000'] * n_events + ['#228B22'] * n_events
    
    events = zip(event_title, date * 3, event_class, event_order, event_backgrounds)
    
    return dict(events=events)

@auth.requires_login()
def reserve_beds():

    """
    This controller:
        - creates an SQLFORM form to submit a booking
        - handles responses to errors and success
        - adds bookings to google calendar (TODO - move to confirmed v pending)
        - emails booker (TODO - turned off)
    """
    
    form = SQLFORM(db.bed_reservations, 
                   fields=['research_visit_id', 'arrival_date',
                           'departure_date', 'number_of_visitors',
                           'purpose', 'look_see_visit'],
                   labels = {'research_visit_id':'Approved research visit',
                             'look_see_visit':'These bed bookings are just to explore '
                             'the facilities and research opportunties at SAFE and are '
                             'not associated with an ongoing research project.'},
                   submit_button = 'Submit and add names')
    
    if form.process(onvalidation=validate_reserve_beds).accepted:
        
        # # Mechanism for sending email to logged in user
        # mail.send(to=auth.user.email,
        #           subject='hello',
        #           message='hi there')
        
        # update the bed bookings table
        # - get a list of days to book
        days_occupied = day_range(form.vars.arrival_date, form.vars.departure_date)
        # - check those days are in the bed bookings table
        for d in days_occupied:
            row = db.bed_data(day=d)
            if row is None:
                db.bed_data.insert(day=d, pending=form.vars.number_of_visitors)
            else:
                row.update_record(pending=row.pending + form.vars.number_of_visitors)
        
        # add the proposer to the reservation names list table
        db.bed_reservation_member.insert(bed_reservation_id = form.vars.id,
                                         user_id = auth.user.id)

        session.flash = CENTER(B('Bed reservation request submitted.'), _style='color: green')
        redirect(URL('bed_reservations','bed_reservation_details', args=form.vars.id))
    
    elif form.errors:
        response.flash = CENTER(B('Errors in form, please check and resubmit'), _style='color: red') 
    else:
        pass
    
    return dict(form=form)


def validate_reserve_beds(form):
    
    # add the user id into the response
    form.vars.reserver_id = auth.user_id
    form.vars.reservation_date =  datetime.date.today().isoformat()
    
    # check the arrival date is more than a fortnight away
    deadline = datetime.date.today() + datetime.timedelta(days=14)
    if form.vars.arrival_date < deadline:
        form.errors.arrival_date = '14 days notice required. Arrival date must be later than {}.'.format(deadline.isoformat())
    
    # check the departure date is after the arrival date
    # TODO - think about day visits
    if form.vars.arrival_date >= form.vars.departure_date:
        form.errors.departure_date = 'The departure date must be later than the arrival date'
    
    # can't be not a look see and not have a visit id - oddly the boolean checkbox widget
    # returns None when not checked...
    if form.vars.look_see_visit in [False, None] and form.vars.research_visit_id is None:
        form.errors.research_visit_id = ('Unless you just want to evaluate research '
                                        'opportunities at SAFE (see check box below), '
                                        'you must specify an approved research visit. ' 
                                        'This menu will only give access to current approved '
                                        'visits for projects of which you are a member.')
    
    # if this is a project visit then the beds have to be within the project dates
    if form.vars.look_see_visit in [False, None] and form.vars.research_visit_id is not None:
        
        visit_record = db.research_visit(form.vars.research_visit_id)
        
        if form.vars.arrival_date < visit_record.arrival_date:
            form.errors.arrival_date = ('The bed reservation cannot start before the approved '
                                        'research visit start date of {}.').format(visit_record.arrival_date.isoformat())
        
        if form.vars.departure_date > visit_record.departure_date:
            form.errors.departure_date = ('The bed reservation cannot end after the approved '
                                          'research visit start date of {}.').format(visit_record.departure_date.isoformat())
    
    # Check there are enough beds for the visit
    # - get a range of days
    days_requested = day_range(form.vars.arrival_date, form.vars.departure_date)
    rows = db(db.bed_data.day.belongs(days_requested)).select()
    pend = [0 if r.pending is None else r.pending for r in rows]
    conf = [0 if r.approved is None else r.approved for r in rows]
    avail = [n_beds_available - (x+y) for x, y in zip(pend, conf)]
    can_fit_request = [x >= form.vars.number_of_visitors for x in avail]
    if not all(can_fit_request):
        form.errors.number_of_visitors = ('There are not enough available beds to cover'
                                          ' the requested visit. There is currently a maximum of ' +
                                          str(min(avail)) + ' beds available throughout this '
                                          'period, although more may become available if some '
                                          'pending reservations are rejected.')


def day_range(start, end):
    
    days = []
    curr = start
    while curr < end: # don't assign a bed on the departure date, so not <=
        days.append(curr)
        curr += datetime.timedelta(days=1)
    
    return(days)



@auth.requires_login()
def bed_reservation_details():
    
    """
    Controller to show the details of a bed reservation
     - can be called without an argument to create a new reservation
       or with one to expose the reservation details and maybe edit
     - collects a bunch of information and passes it to the view for processing
    """
    
    # get the reservation_id from the call
    reservation_id = request.args(0)
    
    # establish we've got an actual real project ID if one is provided
    if reservation_id is not None:
        reservation_record = db.bed_reservations(reservation_id)
        if reservation_record is None:
            session.flash = CENTER(B('Unknown SAFE reservation id in bed_reservations/bed_reservation_details'), _style='color: red')
            redirect(URL('bed_reservations','bed_reservations'))
    else:
        reservation_record = None
    
    # setup whether editing is allowed for various possibilities
    if reservation_record is not None and reservation_record.look_see_visit is False:
        # Existing project related reservations
        # - Is the user is a coordinator of the project associated with this record?
        editor_set = db((db.project_members.project_id == reservation_record.project_id) &
                        (db.project_members.is_coordinator == 'True') &
                        (db.project_members.user_id == auth.user.id)).select()
        readonly = False if  len(editor_set) > 0 else True
        button_text = 'Update SAFE reservation'
    elif reservation_record is not None and reservation_record.look_see_visit is True:
        # Existing look see reservations
        # Is the user a member of the group?
        editor_set = db((db.bed_reservation_member.bed_reservation_id == reservation_record.id) &
                        (db.bed_reservation_member.user_id == auth.user.id)).select()
        readonly = False if  len(editor_set) > 0 else True
        button_text = 'Update SAFE reservation'
    else:
        # New record so allow anyone to propose
        readonly = False
        button_text = 'Submit SAFE reservation'
    
    # check for existing members
    if reservation_record is not None:
        # deleting records needs to reference bed_reservation_member.id and showing
        # H&S records needs auth_user (and these could be None) so need a left full join
        query = db(db.bed_reservation_member.bed_reservation_id == reservation_record.id)
        reservation_members = query.select(db.auth_user.id,
                                           db.auth_user.last_name,
                                           db.auth_user.first_name,
                                           db.auth_user.h_and_s_id,
                                           db.bed_reservation_member.id,
                                           left=db.bed_reservation_member.on(db.auth_user.id == db.bed_reservation_member.user_id))
    else:
        reservation_members = None
    
    # expose the booking form
    form = SQLFORM(db.bed_reservations, 
                   record = None if reservation_record is None else reservation_record.id,
                   readonly = readonly,
                   fields = ['research_visit_id', 'arrival_date',
                           'departure_date', 'number_of_visitors',
                           'purpose', 'look_see_visit'],
                   labels = {'research_visit_id':'Approved research visit',
                             'look_see_visit':'These bed bookings are just to explore '
                             'the facilities and research opportunties at SAFE and are '
                             'not associated with an ongoing research project.'},
                   submit_button = button_text,
                   showid=False)
    
    # handle form submission
    if form.process(onvalidation=validate_reserve_beds).accepted:
        
        # # Mechanism for sending email to logged in user
        # mail.send(to=auth.user.email,
        #           subject='hello',
        #           message='hi there')
        
        # update the bed bookings table
        # - get a list of days to book
        days_occupied = day_range(form.vars.arrival_date, form.vars.departure_date)
        # - check those days are in the bed bookings table
        for d in days_occupied:
            row = db.bed_data(day=d)
            if row is None:
                db.bed_data.insert(day=d, pending=form.vars.number_of_visitors)
            else:
                row.update_record(pending=row.pending + form.vars.number_of_visitors)
        
        # add the proposer to the reservation names list table
        db.bed_reservation_member.insert(bed_reservation_id = form.vars.id,
                                         user_id = auth.user.id)

        session.flash = CENTER(B('Bed reservation request submitted.'), _style='color: green')
        redirect(URL('bed_reservations','bed_reservation_details', args=form.vars.id))
    
    elif form.errors:
        response.flash = CENTER(B('Errors in form, please check and resubmit'), _style='color: red') 
    else:
        pass
    
    # Now handle members:
    if readonly is False and reservation_record is not None:
        # check who can be added
        # - project members for research visits and anyone for look see visits
        # - and not already a member in either case
        if reservation_record.look_see_visit is False:
            # - project members for research visits and not already a member
            valid_ids = db(db.project_members.project_id == reservation_record.project_id)._select(db.project_members.user_id)
            already_selected = db(db.bed_reservation_member.bed_reservation_id == reservation_record.id)._select(db.bed_reservation_member.user_id)
            query = db(db.auth_user.id.belongs(valid_ids) & ~ db.auth_user.id.belongs(already_selected))
        else:
            # - anyone who is not already a member for look see visits
            already_selected = db(db.bed_reservation_member.bed_reservation_id == reservation_record.id)._select(db.bed_reservation_member.user_id)
            query = db( ~ db.auth_user.id.belongs(already_selected))
        
        db.bed_reservation_member.user_id.requires = IS_IN_DB(query, db.auth_user.id, '%(last_name)s, %(first_name)s')
        db.bed_reservation_member.bed_reservation_id.default = reservation_record.id
        add_member = SQLFORM(db.bed_reservation_member,  fields=['user_id'])
    else:
        members = None
        add_member = None
    
    if add_member is not None:
        if add_member.process(onvalidation=validate_bed_reservation_add_member).accepted:
            session.flash = CENTER(B('New member added to reservation.'), _style='color: green')
            redirect(URL('bed_reservations','bed_reservation_details', args=reservation_record.id))
        elif add_member.errors:
            print add_member.errors
            response.flash = CENTER(B('Problem with adding member to reservation.'), _style='color: red')
        else:
            pass
    
    # # need to handle project reservations and look see reservations
    # if reservation_record.look_see_visit is False:
    #     # get the visit record to give to the view
    #     visit_record = db.research_visit(reservation_record.research_visit_id)
    #     # get project members for that visit (who have the right to add themselves or others)
    #     project_members_query = db.project_members.project_id == visit_record.project_id
    #     project_members_rows = db(project_members_query).select(db.project_members.user_id)
    #     authorised_users = [r.user_id for r in project_members_rows]
    # else:
    #     # otherwise get current visit members and give them the right to add
    #     already_reserved = authorised_users = [r.id for r in reservation_members_rows]
    #     # return an empty visit record
    #     visit_record = None
    #
    # if auth.user.id in authorised_users:
    #
    #     # lock down the possible entry fields to project members if this
    #     # is a project reservation, otherwise can select any registered user
    #     if reservation_record.look_see_visit is False:
    #
    #         # who is already in the project but not visiting
    #         already_reserved = [r.id for r in reservation_members_rows]
    #         not_reserved = list(set(authorised_users) - set(already_reserved))
    #         query = db(db.auth_user.id.belongs(not_reserved))
    #         # - require the visit members to be in that set
    #         db.bed_reservation_member.user_id.requires = IS_IN_DB(query, db.auth_user.id, '%(last_name)s, %(first_name)s')
    #
    #     # lock down the possible value of research_visit_id
    #     db.bed_reservation_member.bed_reservation_id.default = reservation_id
    #     db.bed_reservation_member.bed_reservation_id.readable = False
    #
    #     # set the labels
    #     if reservation_record.look_see_visit is False:
    #         labels = {'user_id':'Project member to add'}
    #     else:
    #         labels = {'user_id':'SAFE registered member to add'}
    #
    #     if len(already_reserved) < reservation_record.number_of_visitors:
    #         form = SQLFORM(db.bed_reservation_member,
    #                        fields = ['user_id'],
    #                        labels = labels)
    #
    #         # process and reload the page
    #         if form.process().accepted:
    #             redirect(URL('bed_reservations', 'bed_reservation_details', args=reservation_id))
    #     else:
    #         form = B(CENTER("The requested number of beds has been filled. No more members can be added."), _style='color:red;')
    # else:
    #     form = None
    
    return dict(reservation_record = reservation_record, reservation_members = reservation_members, 
                add_member=add_member, form=form, readonly=readonly)


def validate_bed_reservation_add_member(form):
    
    # get the reservation record and check that the booking is for enough people
    # TODO - extend this add another space if possible
    
    print form.vars.id
    pass

@auth.requires_login()
def remove_member():

    """
    Removes a row from the bed_reservation_member table and as such needs careful safeguarding
    against use by non-authorised people - must be a logged in user who is a coordinator
    for the project or a member of a visit 
    """

    # get the row id
    row_id = request.args(0)
    
    if row_id is not None:
        member_record = db.bed_reservation_member(row_id)
        visit_record = db.bed_reservations(member_record.bed_reservation_id)
    else:
        session.flash = CENTER(B('Unknown row ID in bed_reservations/remove_member'), _style='color: red')
        redirect(URL('bed_reservations','bed_reservations'))
    
    if member_record is not None:
    
        # # get a set of users who have the right to access this interface for the row
        # project_coords = db((db.project_members.project_id == visit_record.project_id) &
        #                     (db.project_members.is_coordinator == 'True')).select()
        # project_coord_id = [r.user_id for r in project_coords]
    
        # if the user is a member then check it makes sense to delete and do so.
        if  auth.user.id > 0:
            
            # TODO - notify the member that they're being removed?
            session.flash =  CENTER(B('Project member removed'), _style='color: green')
            member_record.delete_record()
            redirect(URL('bed_reservations','bed_reservation_details', args=member_record.bed_reservation_id))
        else:
            session.flash =  CENTER(B('Unauthorised use of research_visits/remove_member'), _style='color: red')
            redirect(URL('bed_reservations','bed_reservation_details', args=member_record.bed_reservation_id))



# decorator restricts access to admin users
# - the link is only revealed in the menu for admin users but 
#   that doesn't prevent pasting in the link!
@auth.requires_membership('admin')
def administer_reserve_beds():
    
    """
    This controller handles:
     - presenting admin users with a list of current visits from the db
     - allowing the admin to approve or reject visit requests
     - Currently this locks all but the admin fields. Might also need
       a separate admin visit booking form with fewer restrictions for
       overbooking and special cases?
    """
    
    # don't want the admin to change any of this about a visit
    db.bed_reservations.research_visit_id.writable = False
    db.bed_reservations.reserver_id.writable = False
    db.bed_reservations.arrival_date.writable = False
    db.bed_reservations.departure_date.writable = False
    db.bed_reservations.number_of_visitors.writable = False
    db.bed_reservations.purpose.writable = False
    db.bed_reservations.look_see_visit.writable = False
    
    # get a query of pending requests with user_id
    form = SQLFORM.grid(query=(db.bed_reservations.admin_status == 'Pending'), csv=False,
                        fields=[db.bed_reservations.arrival_date, 
                                db.bed_reservations.departure_date,
                                db.bed_reservations.number_of_visitors],
                         maxtextlength=250,
                         deletable=False,
                         editable=True,
                         create=False,
                         details=False,
                         editargs = {'fields': ['research_visit_id','reserver_id',
                                                'arrival_date','departure_date',
                                                'number_of_visitors','purpose', 'look_see_visit',
                                                'admin_status','admin_notes'],
                                     'showid': False},
                         onvalidation = validate_administer_reserve_beds,
                         onupdate = update_administer_reserve_beds)
    
    return dict(form=form)

@auth.requires_membership('admin')
def validate_administer_reserve_beds(form):
    
    # validation handles any checking (none here) and also any 
    # amendments to the form variable  - adding user and date of admin
    form.vars.admin_id = auth.user_id
    form.vars.admin_decision_date =  datetime.date.today().isoformat()
    
    return(form)

@auth.requires_membership('admin')
def update_administer_reserve_beds(form):

    """
    This handler updates the bed_data table to remove or confirm beds 
    on each day and then emails the person who made the reservation
    """
    
    # recover the record to get all the stuff that isn't writable
    record = form.record
    
    # get the dates when beds are pending booked
    # and select as an iterable of rows
    days_requested = day_range(record.arrival_date, record.departure_date)
    for d in days_requested:
        row = db.bed_data(day=d)
        if(form.vars.admin_status == 'Approved'):
            row.update_record(pending=row.pending - record.number_of_visitors,
                              approved=(0 if row.approved is None else row.approved) + record.number_of_visitors)
        else:
            row.update_record(pending= row.pending - record.number_of_visitors)
    
    # email the reserver
    if(form.vars.admin_status == 'Approved'):
        
        # email the proposer
        mail.send(to=record.reserver_id.email,
              subject='hello',
              message='Welcome to SAFE')
    
    elif(form.vars.admin_status == 'Rejected'):
        
        mail.send(to=record.reserver_id.email,
              subject='hello',
              message='Nope')
    else:
        pass
    
    return dict(form=form)




@auth.requires_membership('admin')
def manage_bed_reservations():
    
    """
    This controller provides a grid view of approved bookings and allows
    the admin to create new bookings and delete existing ones. Having function
    to _edit_ bookings is probably necessary but would have to track changes in
    date and number of beds, so for the moment just cancel and rebook.
    """
    
    # admins can only make Approved bookings
    db.bed_reservations.admin_status.default = 'Approved'
    db.bed_reservations.admin_status.writable = False
    
    form = SQLFORM.grid(query=(db.bed_reservations.admin_status == 'Approved'),
                        fields = [db.bed_reservations.arrival_date,
                                  db.bed_reservations.departure_date,
                                  db.bed_reservations.number_of_visitors,
                                  db.bed_reservations.reserver_id],
                        create = True,
                        editable = False, 
                        deletable = True,
                        formargs = {'fields': ['research_visit_id', 'arrival_date',
                                               'departure_date','number_of_visitors',
                                               'purpose','look_see_visit','reserver_id',
                                               'admin_status', 'admin_notes'],
                                    'showid': False},
                        csv = False,
                        onvalidation = validate_administer_reserve_beds,
                        oncreate = create_manage_bed_reservations,
                        ondelete = delete_manage_bed_reservations)
    
    return dict(form=form)


@auth.requires_membership('admin')
def create_manage_bed_reservations(form):

    """
    This handler updates the bed_data table to add beds 
    on each day and then emails the person who made the reservation
    but unlike the update_reserve, it know it is going to get approved posts
    and it doesn't get the record but the form.vars
    """
        
    # get the dates of the booking
    days_requested = day_range(form.vars.arrival_date, form.vars.departure_date)
    for d in days_requested:
        row = db.bed_data(day=d)
        if row is None:
            db.bed_data.insert(day = d, approved = form.vars.number_of_visitors)
        else:
            row.update_record(approved=(0 if row.approved is None else row.approved) + form.vars.number_of_visitors)
    
    # email the proposer
    reserver_email = db.auth_user(form.vars.reserver_id).email
    mail.send(to=reserver_email,
              subject='Admin bed reservation at SAFE',
              message='Welcome to SAFE')
    
    response.flash = CENTER(B('Admin bed booking made.'), _style='color: green')



@auth.requires_membership('admin')
def delete_manage_bed_reservations(table, id):

    """
    This handler removes a record and gets rid of the bed data and emails the original reserver
    """
        
    record = table(id)
    
    # get the dates of the booking
    days_requested = day_range(record.arrival_date, record.departure_date)
    for d in days_requested:
        row = db.bed_data(day=d)
        row.update_record(approved=row.approved - record.number_of_visitors)
    
    # email the proposer
    mail.send(to=record.reserver_id.email,
              subject='Your bed reservation at SAFE has been cancelled',
              message='Unlucky')
    
    response.flash = CENTER(B('Bed booking cancelled.'), _style='color: red')


def export_bed_reservation_workbook():
    
    """
    This creates an excel workbook showing the current bed reservation data
    and pokes it out as an http download, so a button can call the controller
    and return the file via the browser. Formatting needs some work but Ryan
    was keen on this.
    """
    
    query = ((db.bed_reservations.admin_status == 'Approved') &
             (db.bed_reservations.departure_date > datetime.date.today().isoformat()))
    data = db(query).select(orderby=db.bed_reservations.arrival_date)
    
    if len(data) > 0:
        # open the workbook object
        wb = openpyxl.Workbook()
        # get the worksheet
        ws = wb.active
        ws.title = 'Bed reservations'
        ws['A1'] = 'Bed Reservations at SAFE as of {}'.format(datetime.date.today().isoformat())
    
        # get the arrival date of the first row (which is the earliest, due to the orderby)
        end_date = start_date = data.first().arrival_date
    
        current_row = start_row = 6
        last_col = start_col = 1
    
        fill1 = openpyxl.styles.PatternFill(patternType='solid', 
                                            fgColor=openpyxl.styles.Color('55BBBBBB'))
        fill2 = openpyxl.styles.PatternFill(patternType='solid', 
                                            fgColor=openpyxl.styles.Color('FFDDDDDD'))
    
        for i, row in enumerate(data):
            this_start_col = (row.arrival_date - start_date).days + start_col 
            this_end_col = (row.departure_date - start_date).days + start_col 
            ws.merge_cells(start_row=current_row, 
                           start_column=this_start_col,
                           end_row=current_row + row.number_of_visitors - 1 ,
                           end_column=this_end_col)
            if last_col < this_end_col:
                last_col = this_end_col
                end_date = row.departure_date
            c = ws.cell(row=current_row, column=(row.arrival_date - start_date).days + start_col)
            c.style.alignment.vertical = "top"
            c.value = row.purpose
            c.fill = fill1 if (i % 2 == 0) else fill2
            current_row  += row.number_of_visitors 
    
        # label up the dates
        dates = day_range(start_date, end_date + datetime.timedelta(days=1))
        weekday = [d.strftime('%a') for d in dates]
        month = [d.strftime('%B') for d in dates]
        day = [d.day for d in dates]
    
        for i, (d, w) in enumerate(zip(day, weekday)):
            c = ws.cell(row=start_row - 1, column=start_col + i)
            c.value = d
            c = ws.cell(row=start_row - 2, column=start_col + i)
            c.value = w
    
        current_col = start_col
        while len(month) > 0:
            m_counter = Counter(month)
            this_month = month[0]
            n_month = m_counter[this_month] 
            ws.merge_cells(start_row=start_row - 3, 
                           start_column=current_col,
                           end_row=start_row - 3,
                           end_column=current_col + n_month - 1)
            c = ws.cell(row=start_row - 3, column=current_col)
            c.value = this_month
            current_col = current_col + n_month
            month = month[n_month:]
    
        for i in range(start_col, last_col + 1):
            ws.column_dimensions[openpyxl.cell.get_column_letter(i)].width = '3'
    
        # and now poke it out to the browser
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        attachment = 'attachment;filename=SAFE_Bed_reservations_{}.xlsx'.format(datetime.date.today().isoformat())
        response.headers['Content-Disposition'] = attachment
        content = openpyxl.writer.excel.save_virtual_workbook(wb)
        raise HTTP(200, str(content),
                   **{'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                      'Content-Disposition':attachment + ';'})
    else:
        session.flash = CENTER(B('No approved bed reservations found.'), _style='color: red')
        redirect(URL('bed_reservations','manage_bed_reservations'))


