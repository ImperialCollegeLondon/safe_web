import datetime
from gluon.storage import Storage
import openpyxl
import itertools
from collections import Counter, OrderedDict
import StringIO

## -----------------------------------------------------------------------------
## RESEARCH VISITS
## -- provide a general grid view of visits for users and routes to a detail view 
##   that allows research visits to be proposed and various bookings to be made
## -----------------------------------------------------------------------------


@auth.requires_login()
def research_visits():
    
    """
    This controller shows the grid view for visits and allows
    logged in users to view details

    It uses a custom button to divert from the SQLFORM.grid view to a
    custom view page that allows project members to add visitors
    """

    # For standard users (need a separate admin projects controller)
    # don't show the authorization fields and don't show a few behind 
    # the scenes fields
    db.research_visit.admin_notes.readable = False 
    
    # create a links list that:
    # 1) creates a custom button to pass the row id to a custom view 
    # Commented out code here allows the form to show a pretty icon for status, BUT
    # blocks it from being used in searches. So don't do that.
    
    links = [#dict(header = 'Admin Status', body = lambda row: approval_icons[row.admin_status]),
             dict(header = '', body = lambda row: A('View',_class='button btn btn-default',
                  _href=URL("research_visits","research_visit_details", args=[row.id]),
                  _style='padding: 3px 10px 3px 10px;'))
            ]
    
    # suppress status in  SQLFORM grid whilst making it available for links
    # db.research_visit.admin_status.readable = True 
    
    form = SQLFORM.grid(db.research_visit, csv=False, 
                        fields=[db.research_visit.project_id, db.research_visit.title, 
                                db.research_visit.arrival_date, db.research_visit.departure_date,
                                db.research_visit.admin_status],
                        maxtextlength=250,
                        deletable=False,
                        editable=False,
                        create=False, 
                        details=False,
                        links=links,
                        links_placement='right',
                        formargs={'showid': False})

    return dict(form=form)


@auth.requires_login()
def research_visit_details():
  
    """
    Complex controller to book visits to give a single page to hold:
        * Complete and submit basic RV details (purpose/period)
        * That exposes a set of booking controls...
            - visitors and H&S
            - booking beds at SAFE and Maliau
            - booking RA time
            - booking transfers
        * ... and a set of tables of existing bookings with cancel buttons
    
    It relies on some client side javascript to:
    - check before deleting visitors
    - update SAFE availability
    - provide check all buttons for selecting sets of checkboxes
    - restricted date ranges using datepicker (parameterised on the fly from the controller)
    """
    
    ##
    ## SECTION 1) CHECK USER STATUS AND SET UP 
    ##
    
    # Three possible entry points, for a three step process to allow project date look ups
    # - Completely new RV request(bare URL)
    # - Project specified for new RV request (project_id as a variable, but no record)
    # - Existing record passed as an argument to the URL.
    
    # 1a) SANITISE THE INPUTS
    rv_id = request.args(0)
     
    if rv_id is not None:
        record = db.research_visit(rv_id)
        new_rv_project_requested = '0'
        # # If the visit is given as an ID, does it really exist?
        if record is None:
            session.flash = B(CENTER('Invalid research visit id'), _style='color:red;')
            redirect(URL('research_visits','research_visits'))
    else:
        record = None
        new_rv_project_requested = request.vars['new']
        # if a project has been requested (and it isn't a 
        # look see project) then check it exists
        if new_rv_project_requested is not None and new_rv_project_requested != '0':
            new_project_record = db.project_id(new_rv_project_requested)
            if new_project_record is None:
                session.flash = B(CENTER('Invalid new visit project reference'), _style='color:red;')
                redirect(URL('research_visits','research_visit_details'))
    
    # 1b) get a list of projects that the user is a coordinator of, to check for both project
    #     selection and subsequent project detail editing and booking.
    coord_query = db((db.project_members.user_id == auth.user_id) &
                     (db.project_members.is_coordinator == 'True') &
                     (db.project_members.project_id == db.project_id.id) &
                     (db.project_id.project_details_id == db.project_details.id))
    
    rows = coord_query.select(db.project_details.project_id, db.project_details.title)
    available_project_ids = [r.project_id for r in rows]
    available_project_titles = [r.title for r in rows]
    
    # 1c) setup whether the record is editable
    if record is None:
        # just in the process of launching a proposal
        readonly = False 
    elif record.admin_status == 'Submitted':
        # this proposal is under consideration, regardless of user
        readonly = True
    elif (record.project_id in available_project_ids):
        # this proposal is in the set that the user is a coordinator for
        readonly = False
    elif record.project_id is None and auth.user.id == record.proposer_id:
        # The proposer is editing his own look see visit.
        readonly = False
    else:
        # just a random viewer, so no right access
        readonly = True
    
    ## SECTION 2) CAPTURE THE BASIC VISIT DETAILS
    ## This is a three step process. 
    ## A) Get a project reference, in order to get date limits, 
    ## B) Provide a fuller visit details form 
    ## C) Expose a set of booking controls for the user to build the plan
    
    if rv_id is None and new_rv_project_requested is None:
        
        # Bare URL submitted: provide a list of available projects + look see visit
        # and redirect back to the page, with the new project_id for the next step
        
        project_selector = SELECT(OPTION('Look see visit', _value='0'),
                                  *[OPTION(title, _value=pid) for title, pid in
                                    zip(available_project_titles, available_project_ids)],
                                  _class='form-control', _name='project_selector')
        
        visit= FORM(DIV(DIV(H5('Research visit summary'), _class="panel-heading"),
                        DIV(DIV(LABEL('Choose project:', _class="control-label col-sm-2" ),
                                DIV(project_selector, _class="col-sm-8"),
                                TAG.BUTTON('Select', _style="padding: 5px 15px", _class='col-sm-2',
                                            _type='submit', _name='submit_project_select'),
                                _class='row', _style='margin:10px 10px'),
                            _class='panel_body'),
                        _class="panel panel-primary"))
        
        if visit.validate():
            # reload the URL with the id of the new project as a variable
            redirect(URL('research_visits','research_visit_details', vars={'new': visit.vars.project_selector}))
            
    else: 
        # Either a) URL with a project requested as a variable 'research_visit_details?new='866' 
        #        b) URL giving an existing record 'research_visit_details/4' 
        if rv_id is not None:
            # intercept existing records first
            buttons = [TAG.button('Save edits',_type="submit",
                                   _name='save_proposal', _style='padding: 5px 15px 5px 15px;')]
        elif new_rv_project_requested is not None:
            # then new ones (as the code sets new_rv_request_submitted = 0 for rv_id calls)
            buttons = [TAG.button('Create proposal',_type="submit",
                                   _name='save_proposal', _style='padding: 5px 15px 5px 15px;')]
        
        # Use SQLFORM for DB input
        visit = SQLFORM(db.research_visit, 
                        record = record,
                        readonly = readonly,
                        fields = ['title','arrival_date',
                                  'departure_date','purpose','licence_details'],
                        buttons = buttons,
                        showid = False)
        
        # process the visit form to create hidden fields and to process input
        if visit.process(onvalidation=validate_research_visit_details, formname='visit').accepted:
            
            # if this is a new proposal then need to insert the project_id,
            # which isn't included in the form, but not if this is a look see visit
            if rv_id is None and new_rv_project_requested != '0':
                db.research_visit(visit.vars.id).update_record(project_id = new_rv_project_requested)
            
            # if this is a new draft, email the proposer the link for the page
            if rv_id is None:
                SAFEmailer(to=auth.user.email,
                           subject='SAFE: draft research visit proposal created',
                           template =  'research_visit_created.html',
                           template_dict = {'name': auth.user.first_name,
                                            'url': URL('research_visits', 'research_visit_details',
                                                       args=[visit.vars.id], scheme=True, host=True)})
                
                db.research_visit(visit.vars.id).update_record(admin_status = 'Draft')
                session.flash = CENTER(B('Research visit proposal created'), _style='color: green')
                
            else:
                session.flash = CENTER(B('Research visit proposal updated'), _style='color: green')
            
            redirect(URL('research_visits','research_visit_details', args=visit.vars.id))
        else:
            
            pass
    
        # Now repackage the form into a custom DIV
        # edit form widgets - notably, override the default date widget classes to allow
        # them to use the daterange datepicker
        if not readonly:
            visit.custom.widget.purpose['_rows'] = 4
            visit.custom.widget.arrival_date['_class'] = "form-control input-sm"
            visit.custom.widget.departure_date['_class'] = "form-control input-sm"
        
        # get the project details
        if (record is not None and record.project_id is not None) or (new_rv_project_requested != '0'):
            
            if new_rv_project_requested != '0':
                pid = int(new_rv_project_requested)
            elif record.project_id is not None:
                pid = record.project_id

            proj_row = db((db.project_id.id == pid) & (db.project_id.project_details_id == db.project_details.id))
            project_details = proj_row.select().first()
            proj_title = project_details.project_details.title
            
        else:
            project_details = None
            proj_title = 'Look see visit'
        
        proj_row = DIV(LABEL('Project title:', _class="control-label col-sm-2" ),
                       DIV(proj_title, _class="col-sm-10"),
                       _class='row', _style='margin:10px 10px')
        
        # get a download link for the budget and timetable spreadsheet
        if rv_id is None:
            download_link = DIV()
        else:
            download_link = CAT('Click on this link to download a spreadsheet of the details and estimated costs: ',
                              A('Download spreadsheet', _href=URL('research_visits', 'export_my_research_visit', args=rv_id)))
        
        # fix up the dates to control the datepicker and the bed booking limits
        if auth.has_membership('admin'):
            # admins can book any number of people and do so retrospectively
            visit_start_min = ''
            visit_end_max = ''
            bed_booking_limit = 1e7
        else:
            if new_rv_project_requested == '0' or (record is not None and record.project_id is None):
                # look see visits can book up to the normal bed limit with a fortnights notice
                visit_start_min = '+14d'
                visit_end_max = ''
            else:
                # project bookings have to be at least 14 days notice and within the project dates
                project_start = project_details.project_details.start_date
                project_end = project_details.project_details.end_date
                fortnight = datetime.date.today() + datetime.timedelta(days=14)
                # check the project hasn't finished
                if project_end < fortnight:
                    session.flash = CENTER(B('The completion date for the proposed project has passed.'), _style='color: red')
                    redirect(URL('research_visits', 'research_visit_details'))
                else:
                    visit_start_min = max(project_start, fortnight).isoformat()
                    visit_end_max = project_end.isoformat()
            
        # javascript to run the datepicker
        visit_js =datepicker_script(id = 'visit_datepicker',
                                    autoclose = 'true',
                                    startDate ='"' + visit_start_min + '"',
                                    endDate ='"' + visit_end_max + '"')
        
        # status flag
        if record is not None:
            status = DIV(approval_icons[record.admin_status], XML('&nbsp'),
                         'Status: ', XML('&nbsp'), record.admin_status, 
                         _class='col-sm-3',
                         _style='padding: 5px 15px 5px 15px;background-color:lightgrey;color:black;')
        else:
            status = DIV()
        
        visit = FORM(DIV(DIV(DIV(H5('Research visit summary', _class='col-sm-9'), status, _class='row', _style='margin:0px 0px'),
                            _class="panel-heading"),
                        DIV(visit.custom.begin, proj_row,
                            DIV(LABEL('Visit title:', _class="control-label col-sm-2" ),
                                DIV(visit.custom.widget.title,  _class="col-sm-10"),
                                _class='row', _style='margin:10px 10px'),
                            DIV(LABEL('Dates:', _class="control-label col-sm-2" ),
                                DIV(DIV(visit.custom.widget.arrival_date,
                                        SPAN('to', _class="input-group-addon input-sm"),
                                        visit.custom.widget.departure_date,
                                        _class="input-daterange input-group", _id="visit_datepicker"),
                                    _class='col-sm-10'),
                                _class='row', _style='margin:10px 10px'),
                            DIV(LABEL('Purpose:', _class="control-label col-sm-2" ),
                                DIV(visit.custom.widget.purpose,  _class="col-sm-10"),
                                _class='row', _style='margin:10px 10px'),
                            DIV(DIV(visit.custom.submit,  _class="col-sm-10 col-sm-offset-2"),
                                _class='row', _style='margin:10px 10px'),
                            visit.custom.end,
                           _class='panel_body'),
                        DIV(download_link, _class='panel-footer'),
                        _class="panel panel-primary"),
                        visit_js)
    
    ##
    ## SECTION 3) PROVIDE THE BOOKING CONTROLS 
    ##
    
    
    ## This consists of a giant form - most panels could work independently 
    ## but the first two (accom/transfers), need access to a list of selected visitors
    ## A) A panel to select and edit the list of visitors
    ## B) A panel to book accommodation at SAFE or Maliau
    ## C) A panel to book site transfers
    ## D) A panel to book research assistants
    ## These are followed by tables of existing bookings that allow record deletion
    ## E) SAFE Accomodation
    ## F) Maliau Accomodation
    ## G) Site transfers
    ## H) Research assistant bookings
    ## I) Submit button panel
    
    
    # setup icons and the instructions
    delete_icon = SPAN(_class="glyphicon glyphicon-remove-sign",
                       _style="color:red;")
                       
    replace_icon = SPAN(_class="glyphicon glyphicon-refresh",
                       _style="color:red;")
    
    add_visitor_icon = CAT(SPAN(_class="glyphicon glyphicon-user"), XML('&nbsp;'), 
                           SPAN(_class="glyphicon glyphicon-plus-sign"))
    
    add_project_icon = CAT(SPAN(_class="glyphicon glyphicon-user"), 
                           SPAN(_class="glyphicon glyphicon-user"), 
                           SPAN(_class="glyphicon glyphicon-user"), XML('&nbsp;'), 
                           SPAN(_class="glyphicon glyphicon-plus-sign"))
    
    reserve_bed_icon = CAT(SPAN(_class="glyphicon glyphicon-bed"), XML('&nbsp;'), 
                           SPAN(_class="glyphicon glyphicon-plus-sign"))
    
    release_bed_icon = CAT(SPAN(_class="glyphicon glyphicon-bed"), XML('&nbsp;'),
                           SPAN(_class="glyphicon glyphicon-refresh"))
    
    reserve_transfer_icon = CAT(SPAN(_class="glyphicon glyphicon-road"), XML('&nbsp;'), 
                                SPAN(_class="glyphicon glyphicon-plus-sign"))
        
    reserve_ra_icon  = CAT(SPAN(_class="glyphicon glyphicon-leaf"), XML('&nbsp;'), 
                           SPAN(_class="glyphicon glyphicon-plus-sign"))
    
    if readonly or rv_id is None:
        instructions = DIV()
    else:
        instructions = CAT(H3('Research visit console'),
                            P('The panels below allow you to add members of your research visit and to book '
                              'accomodation, site transfers and research assistant support.'),
                            TAG.DL(TAG.DT('Research visit members'),
                                   TAG.DD(P('You can update the list to include all the members of a project at once (',
                                            add_project_icon, ') or select and add a single user at a time (',
                                            add_visitor_icon, '). You can add an unknown visitor as a placeholder ',
                                           'for a member to be added later.'),
                                          P('When you are ready to update an unknown visitor (or you want to switch ',
                                            'the bookings for a named visitor to someone else), then select the new  ',
                                            'visitor from the dropdown and press the replace button (', replace_icon, 
                                            ') next to the visitor to be replaced. You can also delete a visitor (',
                                             delete_icon, ') but note that this will ', B('automatically cancel all'), 
                                             ' of their accomodation and transfer bookings.'),
                                          P('You will also be able to see which visitors have a health and safety record ',
                                            'and view that record. These records must be complete and sufficiently detailed '
                                            'for a research visit to be approved.'),
                                            _style='padding:0 0 0 20px'),
                                   TAG.DT('Book accomodation'),
                                   TAG.DD(P('Check the boxes next to the research visitors that you want to book ',
                                            'accomodation for and select your booking dates. We can help book ',
                                            'beds at Maliau and you will need to tell us which accomodation you want '
                                            'and if you want any meals provided. Press the reserve beds button (',
                                            reserve_bed_icon, ') to submit the reservation. The web application will ',
                                            'automatically detect overlapping bookings for a visitor and merge them.'),
                                          P('You can delete an individual booking using the booked accomodation table ',
                                            'below. To remove some nights from a booking, you can also select visitors ',
                                            'and a date range and  click the release beds button (', release_bed_icon ,
                                            '). The web application will automatically truncate or split existing reservations',
                                            'to remove bookings for these dates'),
                                          P('The final date in your booking is the ',B('departure date'), ', so you will not '
                                            'have a bed reserved in the evening of the final date.'),
                                            _style='padding:0 0 0 20px'),
                                   TAG.DT('Book site transfers'),
                                   TAG.DD(P('Use the checkboxes to select the visitors who need to travel, and then select ',
                                            'a transfer option and the date you want to travel. Click the transfer ',
                                            'booking button (', reserve_transfer_icon, ') to make the booking.'),
                                            P('Long distance transfers with SAFE vehicles are only available on ', 
                                              B('Wednesdays'), ' and ', B('Sundays'),
                                              '. You are strongly encouraged to liaise with other '
                                              'researchers visiting the site by examining research visit plans and to coordinate '
                                              'long distance transfers with each other to reduce demand on vehicles.'),
                                            P('If you unavoidably need transfers on another day, you will need to email '
                                              'info@safeproject.net to request that these are added to your research visit. '
                                              'Note that transfers on other days of the week are difficult to organize and '
                                              'subject to availability of vehicles, so you may have to be flexible.'),
                                            _style='padding:0 0 0 20px'),
                                   TAG.DT('Book research assistant support'),
                                   TAG.DD(P('Select a date range, the site and time of the day for which you want support ',
                                            'and the type of support needed. Your booking will include work on the finish ',
                                            'date so, for a single day booking, set the finish date to be the same ',
                                            'as the start date. Click the research assistant support ',
                                            'booking button (', reserve_ra_icon, ') to make the booking.'),
                                            _style='padding:0 0 0 20px')))
   
    if rv_id is None:
        console = DIV()
    else:
        # Define common elements shared across tables:
        # -  an extra column in tables when not readonly to insert delete buttons
        if readonly:
            delete_column_head = ""
        else:
            delete_column_head = TH(_width='30px')
        
        # - Each table has it's own local row packing function as the columns vary
        #   This local function (and the globally available uname) provide common code within row packers
        
        def del_btn(rid, btn_name, readonly):
            
            if readonly:
                delete = ""
            else:
                btn_name = btn_name + str(rid)
                delete = TD(TAG.BUTTON(SPAN(_class="glyphicon glyphicon-remove-sign",
                                            _style="color:red;font-size: 1.6em;padding: 0px 10px;"), 
                                            _type='submit', _name=btn_name, _id = btn_name,
                                            _style='background:None;padding:0px'))
            
            return delete
        
        
        # A ) VISITOR PANEL
        # TODO - consider bootstrap modal for prettiness http://plnkr.co/edit/NePR0BQf3VmKtuMmhVR7?p=preview
        
        # get current visitors
        visit_select = db(db.research_visit_member.research_visit_id == rv_id).select()
        
        # Package rows up in a table with row selectors if non read-only
        def pack_visit(r, readonly):
            
            nm = uname(r.user_id, r.id)
            delete = del_btn(r.id, 'delete_visitor_', readonly)
            replace = del_btn(r.id, 'replace_visitor_', readonly)
            

            if not readonly:
                # edit these delete buttons to trigger the java modal warning
                delete.element('BUTTON').attributes['_onclick'] = 'show_alert(this.id)'
                # change the icon in the replace link
                replace.element('span').attributes['_class'] = "glyphicon glyphicon-refresh"
            
            chk = '' if readonly else CAT(INPUT(_type='checkbox', _name='records', _value=r.id), XML('&nbsp;'))
            
            # link to H&S 
            if r.user_id is None or r.user_id.h_and_s_id is None:
                hs = hs_no 
            elif readonly:
                hs = hs_ok
            else: 
                hs = A(hs_ok, _href=URL('health_safety','health_and_safety', args=r.user_id))
            
            row = TR(TD(LABEL(chk, nm)), TD(hs), delete, replace)
            
            return row
        
        table_rows = [pack_visit(r, readonly) for r in visit_select]
        
        # adjust headings for table and insert controls 
        if readonly:
            headings =TR(TH('Visit members'), TH('H&S'))
        else:
            headings =TR(TH(LABEL(INPUT(_type='checkbox', _id='checkAll'), XML('&nbsp;'), 'Select all')),
                         TH('H&S'), delete_column_head, delete_column_head)
            
            # get a selector of valid users
            # - look see visits can add anyone, but projects can only add project members
            # - both can add 'Unknown' users to be updated later. Users are linked to bookings
            #   by the research visit member id field, so can be replaced across the bookings
            # - don't remove people who are already members - needed for replacement
            if record.project_id is None:
                # anyone can join a look see visit
                users = db(db.auth_user.id > 0).select()
            else:
                # select the rows from auth_users for project members
                users = db((db.project_members.project_id == record.project_id) & 
                           (db.project_members.user_id == db.auth_user.id)).select(db.auth_user.ALL)
            
            options = [OPTION(u.last_name + ', ' + u.first_name, _value=u.id) for u in users]
            
            visitor_select = SELECT(OPTION('Unknown', _value=0), 
                                    *options, 
                                    _name='user', 
                                    _class="generic-widget form-control col-sm-3",
                                    _style='height: 30px;width:200px;')
            
            add_visitor = TAG.BUTTON(add_visitor_icon,
                                     _style='padding: 5px 15px',
                                     _type='submit', _name='add_visitor' )
            
            add_project = TAG.BUTTON(add_project_icon,
                                     _style='padding: 5px 15px;background:lightgrey;color:black;',
                                     _type='submit', _name='add_project', 
                                     _title='Add project members')
            
            table_rows.append(TR(TD(visitor_select),TD(add_visitor, _colspan=3)))

        
        # build table
        visitor_table = TABLE(headings, *table_rows, _class='table table-striped') 
        
        # combine into the panel
        if readonly:
            visitors =  DIV(DIV(H5('Research visit members'),  _class="panel-heading"),
                            visitor_table, # vague TODO - make this table handle squashing better (enable .table-responsive?)
                            _class="panel panel-primary", _name='visitors')
        else:
            visitors =  DIV(DIV(DIV(H5('Research Visit Members', _class='col-sm-8'),
                                    DIV(DIV(add_project, _class=' pull-right'), _class='col-sm-4'),
                                    _class='row'),
                                _class="panel-heading"),
                            visitor_table,
                            _class="panel panel-primary")
        
        # B) Accomodation booking form
        
        # create the panel
        accm_pane = DIV(DIV(DIV(H5('Accomodation requests', _class='col-sm-8'),
                                DIV(DIV(TAG.BUTTON(reserve_bed_icon, _type='submit', _name='reserve_beds',
                                                   _style='padding: 5px 15px;background:lightgrey;color:black;'),
                                        XML('&nbsp;')*5,
                                        TAG.BUTTON(release_bed_icon, _type='submit', _name='release_beds',
                                                    _style='padding: 5px 15px;background:lightgrey;color:black;'),
                                        _class='pull-right'),
                                    _class='col-sm-4'),
                                _class='row'),
                            _class="panel-heading"),
                        DIV(DIV(LABEL('Dates:', _class='col-sm-2'),
                                DIV(DIV(INPUT(type="text", _class="form-control input-sm", _name="accom_arrive",
                                              _onchange='date_change()', _id='accom_arrive'),
                                        SPAN('to', _class="input-group-addon"),
                                        INPUT(type="text", _class="form-control input-sm", _name="accom_depart",
                                              _onchange='date_change()', _id='accom_depart'),
                                        _class="input-daterange input-group", _id="accom_datepicker"),
                                    _class='col-sm-10'),
                                _class='row'),
                            DIV(DIV(_class='col-sm-2'),
                                DIV('Note: these dates are the arrival date and departure dates', _class='col-sm-10'),
                                _class='row', _style='margin:2px'),
                            DIV(LABEL('Location:', _class='col-sm-2'),
                                DIV(LABEL(INPUT(_type='radio', _name='location', 
                                                _value='SAFE', value='SAFE', _onclick='locSAFE()'), 
                                         'SAFE', _class='form-control input-sm'),
                                    _class=' col-sm-5'),
                                DIV(LABEL(INPUT(_type='radio', _name='location', 
                                                _value='Maliau', value='SAFE', _onclick='locMaliau()'), 
                                          'Maliau', _class='form-control input-sm'),
                                    _class=' col-sm-5'),
                                _class='row'),
                            DIV(LABEL('Options:', _class='col-sm-2'),
                                DIV(DIV(LABEL(INPUT(_type='radio', _name='maliau_type', _value='Annex', value='Annex'),
                                             'Annex', _class='form-control input-sm'),
                                        _class='form_control'),
                                    DIV(LABEL(INPUT(_type='radio', _name='maliau_type', _value='Hostel', value='Annex'),
                                             'Hostel',_class='form-control input-sm'),
                                        _class='form_control'),
                                    _class=' col-sm-5'),
                                DIV(DIV(LABEL(INPUT(_type='checkbox', _name='maliau_breakfast'), 
                                             'Breakfast', _class='form-control input-sm'),
                                        _class='form_control'),
                                    DIV(LABEL(INPUT(_type='checkbox', _name='maliau_lunch'), 
                                             'Lunch', _class='form-control input-sm'),
                                         _class='form_control'),
                                    DIV(LABEL(INPUT(_type='checkbox', _name='maliau_dinner'), 
                                             'Dinner', _class='form-control input-sm'),
                                         _class='form_control'),
                                    _class=' col-sm-5'),
                                _class='row', _id='maliau_options', _style='display:none;'),
                            _class='panel-body'),
                        # DIV(_class='panel-footer'),
                        _class='panel panel-primary')
        
        # add javascript to power and constrain the daterange picker and hide/reveal maliau options
        accom_js = datepicker_script(id = 'accom_datepicker',
                                     autoclose = 'true',
                                     startDate ='"' + record.arrival_date.isoformat() + '"',
                                     endDate ='"' + record.departure_date.isoformat() + '"')
        
        maliau_js = SCRIPT('function locSAFE() {document.getElementById("maliau_options").style.display = "none";}',
                           'function locMaliau() {document.getElementById("maliau_options").style.display = "block";}',
                            _type='text/javascript')
        
        accom_test = CAT(accm_pane, accom_js, maliau_js)
        
        ## C) Site transfer bookings panel
        transfers_panel =   DIV(DIV(DIV(H5('Site transfer requests (Wed/Sun only)', _class='col-sm-8'),
                                    DIV(DIV(TAG.BUTTON(reserve_transfer_icon, _type='submit', _name='book_transfer',
                                                       _style='padding: 5px 15px;background:lightgrey;color:black;'),
                                            _class='pull-right'),
                                        _class='col-sm-4'),
                                    _class='row'),
                                _class="panel-heading"),
                                DIV(DIV(LABEL('Date:', _class='col-sm-2'),
                                        DIV(INPUT(type="text", _class="form-control input-sm",
                                                  _name="transfer_datepicker", _id="transfer_datepicker"),
                                            _class='col-sm-4'),
                                        LABEL('Transfer:', _class='col-sm-2'),
                                        DIV(SELECT(transfer_set,
                                                   _name='transfer', 
                                                   _class="form-control input-sm"),
                                            _class = "col-sm-4"),
                                        _class='row'),
                                    _class='panel-body'),
                                _class='panel panel-primary')
        
        # add javascript to constrain datepicker and only let admins add days that aren't We/Su
        if auth.has_membership('admin'):
            transfer_days = '""'
        else:
            transfer_days = '"12456"'
        
        transfers_js = datepicker_script(id = 'transfer_datepicker', 
                                         autoclose = "true",
                                         startDate ='"' + record.arrival_date.isoformat() + '"',
                                         endDate ='"' + record.departure_date.isoformat() + '"',
                                         daysOfWeekDisabled = transfer_days)
        
        transfers_panel = CAT(transfers_panel, transfers_js)
        
        ## D) RA booking panel
        ra_panel  = DIV(DIV(DIV(H5('Research Assistant support requests', _class='col-sm-8'),
                                DIV(DIV(TAG.BUTTON(reserve_ra_icon, _type='submit', _name='book_res_assist',
                                                   _style='padding: 5px 15px;background:lightgrey;color:black;'),
                                        _class='pull-right'),
                                    _class='col-sm-4'),
                                _class='row'),
                            _class="panel-heading"),
                        DIV(DIV(LABEL('Dates:', _class='col-sm-3'),
                                DIV(DIV(INPUT(type="text", _class="form-control input-sm", 
                                              _name="ra_start", _id='ra_start'),
                                        SPAN('to', _class="input-group-addon"),
                                        INPUT(type="text", _class="form-control input-sm", 
                                              _name="ra_end", _id='ra_stop'),
                                        _class="input-daterange input-group", _id="ra_datepicker"),
                                    _class='col-sm-9'),
                                _class='row'),
                            DIV(_class='row', _style='margin:2px'),
                            DIV(LABEL('Site and time:', _class='col-sm-3'),
                                DIV(SELECT(res_assist_set,
                                           _name='ra_site_time', 
                                           _class="form-control input-sm"),
                                    _class = "col-sm-9"),
                                _class='row'),
                            DIV(_class='row', _style='margin:2px'),
                            DIV(LABEL('Work type:', _class='col-sm-3'),
                                DIV(LABEL(INPUT(_type='radio', _name='ra_work_type', 
                                                        _value='Standard', value='Standard'), 
                                                 'Standard', _class='form-control input-sm'),
                                    _class='col-sm-3'),
                                DIV(LABEL(INPUT(_type='radio', _name='ra_work_type', 
                                                _value='Rope work', value='Standard'), 
                                         'Rope work', _class='form-control input-sm'),
                                    _class='col-sm-3'),
                                DIV(LABEL(INPUT(_type='radio', _name='ra_work_type', 
                                                _value='Night work', value='Standard'), 
                                         'Night work', _class='form-control input-sm'),
                                    _class='col-sm-3'),
                                _class='row'),
                            _class='panel-body'),
                        # DIV(_class='panel-footer'),
                        _class='panel panel-primary')
        
        ra_js = datepicker_script(id = 'ra_datepicker', 
                                  autoclose = "true",
                                  startDate ='"' + record.arrival_date.isoformat() + '"',
                                  endDate ='"' + record.departure_date.isoformat() + '"')
        
        ra_panel = CAT(ra_panel, ra_js)
        
        ##
        ## NOW BUILD THE TABLES OF THE EXISTING BOOKINGS
        ##
        
        # E ) Booked SAFE accommodation
        
        # grab the rows joined to the RVM table to get user references and pack into a table
        safe_select = db((db.bed_reservations_safe.research_visit_id == rv_id) &
                         (db.bed_reservations_safe.research_visit_member_id == db.research_visit_member.id)).select()
        
        def pack_safe(r, readonly):
            
            nm = uname(r.research_visit_member.user_id, r.research_visit_member.id)
            delete = del_btn(r.bed_reservations_safe.id, 'delete_safe_', readonly)
            row = TR(TD(nm), TD(r.bed_reservations_safe.arrival_date), 
                     TD(r.bed_reservations_safe.departure_date), delete)
            
            return row
        
        if len(safe_select) > 0:
            safe_table = DIV(DIV(H5('Requested accomodation at SAFE'),_class="panel-heading"),
                             TABLE(TR(TH('Visitor'), TH('Arrival date'), TH('Departure date'), delete_column_head),
                                   *[pack_safe(r, readonly) for r in safe_select],
                                   _class='table table-striped'),
                             _class="panel panel-primary")
        else:
            safe_table = DIV()
        
        # F) Booked MALIAU Accommodation
        
        # grab the rows and pack into a table
        maliau_select = db((db.bed_reservations_maliau.research_visit_id == rv_id) &
                           (db.bed_reservations_maliau.research_visit_member_id == db.research_visit_member.id)).select()
        
        def pack_maliau(r, readonly):
            
            nm = uname(r.research_visit_member.user_id, r.research_visit_member.id)
            delete = del_btn(r.bed_reservations_maliau.id, 'delete_maliau_', readonly)
            row = TR(TD(nm), TD(r.bed_reservations_maliau.arrival_date), 
                     TD(r.bed_reservations_maliau.departure_date), 
                     TD(r.bed_reservations_maliau.type),
                     TD(['B' if r.bed_reservations_maliau.breakfast else ''] + 
                        ['L' if r.bed_reservations_maliau.lunch else ''] +
                        ['D' if r.bed_reservations_maliau.dinner else '']),
                     delete)
                     
            return row
        
        if len(maliau_select) > 0:
            maliau_table = DIV(DIV(H5('Requested accomodation at Maliau'),_class="panel-heading"),
                               TABLE(TR(TH('Visitor'), TH('Arrival date'), TH('Departure date'), 
                                        TH('Type'), TH('Food'), delete_column_head),
                                    *[pack_maliau(r, readonly) for r in maliau_select],
                                    _class='table table-striped'),
                               _class="panel panel-primary")
        else:
            maliau_table = DIV()
        
        # G) Booked transfers
        
        # grab the rows and pack into a table
        transfer_select = db((db.transfers.research_visit_id == rv_id) &
                             (db.transfers.research_visit_member_id == db.research_visit_member.id)).select()
        
        def pack_transfer(r, readonly):
            
            nm = uname(r.research_visit_member.user_id, r.research_visit_member.id)
            delete = del_btn(r.transfers.id, 'delete_transfer_', readonly)
            row = TR(TD(nm), TD(r.transfers.transfer), TD(r.transfers.transfer_date), delete)
            
            return row
        
        if len(transfer_select) > 0:
            transfer_table = DIV(DIV(H5('Requested Site Transfers'),_class="panel-heading"),
                                 TABLE(TR(TH('Visitor'), TH('Transfer'), TH('Transfer date'), delete_column_head),
                                       *[pack_transfer(r, readonly) for r in transfer_select],
                                       _class='table table-striped'),
                                 _class="panel panel-primary")
        else:
            transfer_table = DIV()
        
        ## H) Booked research assistants
        res_assist_select = db(db.research_assistant_bookings.research_visit_id == rv_id).select()
        
        def pack_ra(r, readonly):
            
            delete = del_btn(r.id, 'delete_ra_', readonly)
            row = TR(TD(r.site_time), TD(r.start_date), TD(r.finish_date), TD(r.work_type), delete) 
            
            return row
        
        if len(res_assist_select) > 0:
            ra_table = DIV(DIV(H5('Requested Research Assistant support'),_class="panel-heading"),
                           TABLE(TR(TH('Details'), TH('Start date'), TH('Finish date'), 
                                    TH('Work type'), delete_column_head),
                                *[pack_ra(r, readonly) for r in res_assist_select],
                                _class='table table-striped'),
                            _class="panel panel-primary")
        else:
            ra_table = DIV()
        
        ## I) SUBMIT BUTTON PANEL 
        
        # switch the button class from active to inactive based on
        # number of researchers named
        number_of_visitors = db(db.research_visit_member.research_visit_id == rv_id).count()
        nights_at_safe = db(db.bed_reservations_safe.research_visit_id == rv_id).count()
        nights_at_maliau = db(db.bed_reservations_maliau.research_visit_id == rv_id).count()
        
        
        if number_of_visitors and (nights_at_safe + nights_at_maliau):
            submit_proposal_button = TAG.BUTTON('Submit proposal', _type='submit', 
                                                _name='submit_proposal', _class='btn btn-success btn-md active')
            submit_panel_text = P('Double check you have listed all of the researchers coming to the '
                                  'field and provided details of all accommodation, transfers and RA '
                                  ' support needed and then press the button below to submit your proposal.',
                                  _style='padding: 3px 10px 3px 10px;')
        else:
            submit_proposal_button = TAG.BUTTON('Enter details to submit', _type='submit', 
                                                _name='submit_proposal', _class='btn btn-success btn-md disabled',
                                                _disabled='disabled')
            submit_panel_text = P('You have not identified any researchers taking part in the research visit '
                                  'or requested any accommodation. You will not be able to submit your proposal '
                                  'until you provide the details listed at the top of the page.',
                                  _style='padding: 3px 10px 3px 10px;')
        
        submit_panel =  DIV(DIV(DIV(submit_panel_text, _class='row'),
                                DIV(CENTER(submit_proposal_button), _class='row'),
                                _class='panel-body'),
                            _class='panel panel-primary')
        
        
        ##
        ## SECTION 4) EXPOSE THE DETAILS (AND CONTROLS IF NOT READONLY)
        ## AND PROVIDE THE FORM LOGIC FOR HANDLING THE VARIOUS CONTROLS
        
        if readonly:
            console =  CAT(H3('Proposed research visit details'),
                           DIV(visitors),
                           DIV(safe_table),
                           DIV(maliau_table),
                           DIV(transfer_table),
                           DIV(ra_table))
        
        else:
            # combine the panels into a single form, along with a hidden field containing
            # the research visit id to allow the validation to cross check.
            # - insert anchors between panels to bring the URL back to the pane that was edited
            console =  FORM(A(_name='console'),
                            H3('Proposed research visit details'),
                            DIV(DIV(visitors, _class='col-sm-5'),
                                DIV(DIV(accom_test, _class='row'),
                                   DIV(transfers_panel, _class='row'),
                                   DIV(ra_panel, _class='row'),
                                   _class='col-sm-7'),
                                _class='row'),
                            DIV(safe_table),
                            DIV(maliau_table),
                            DIV(transfer_table),
                            DIV(ra_table),
                            DIV(submit_panel),
                            INPUT(_name='id', _id='id', _value=rv_id, _type='hidden'),
                           _id='console')
            
            # console validation
            if console.process(onvalidation = validate_research_visit_details_console, formname='console').accepted:
            
                # intialise a list to gather history for changes
                new_history = []
                
                # add user and project share the same code, so create a local function
                def add_user(uid):
                
                    new_id = db.research_visit_member.insert(research_visit_id = rv_id,
                                                             user_id = uid)
                    # update the history
                    record = db.research_visit_member(new_id)
                    name = uname(record.user_id, record.id)
                    new_history = ' -- New visitor added: {}\\n'.format(name)
                
                    return(new_history)
                
                # --------------------------------
                # The action to take gets identified by the validator and stored 
                # in the form as console.action
                # --------------------------------
                
                if console.action == 'delete':
                    # --------------------------------
                    # The user has pressed one of the delete buttons and 
                    # there is a delete component to the form giving the table 
                    # and row id to be deleted 
                    # --------------------------------
                    
                    table_dict = {'transfer': 'transfers',
                                  'visitor': 'research_visit_member',
                                  'safe': 'bed_reservations_safe',
                                  'maliau': 'bed_reservations_maliau',
                                  'ra': 'research_assistant_bookings'}
                    
                    # get the information formatted
                    del_type = console.delete[0]
                    del_tab = table_dict[del_type]
                    del_id = int(console.delete[1])
                    
                    # if not an RA booking, look up the name of the RVM
                    if del_type == 'visitor':
                        r = db(db.research_visit_member.id == del_id).select().first()
                        name = uname(r.user_id, r.id)
                    elif del_type != 'ra':
                        r = db((db.research_visit_member.id == db[del_tab].research_visit_member_id) &
                               (db[del_tab].id == del_id)).select().first()
                        name = uname(r.research_visit_member.user_id, r.research_visit_member.id)
                        
                    # delete the record
                    del_record = db[del_tab][del_id]
                    del_record.delete_record()
                    
                    # get the history 
                    if del_type == 'ra':
                        new_history.append(' -- RA cancelled {} from {} to {}\\n'.format(
                                            del_record.site_time, del_record.start_date, del_record.finish_date))
                    elif del_type == 'transfer':
                        new_history.append(' -- Transfer cancelled for {} from {} on {}\\n'.format(
                                            name, del_record.transfer, del_record.transfer_date))
                    elif del_type == 'visitor':
                        new_history.append(' -- Visitor removed: {}\\n'.format(name))
                    elif del_type == 'accom_safe':
                        new_history.append(' -- SAFE bed cancelled for {} between {} - {} \\n'.format(
                                            name, del_record.arrival_date, del_record.departure_date))
                    elif del_type == 'accom_safe':
                        new_history.append(' -- Maliau bed cancelled for {} between {} - {} \\n'.format(
                                            name, del_record.arrival_date, del_record.departure_date))
                    
                # --------------------------------
                # THREE (NON-DELETE) ACTIONS THAT AMEND THE VISITORS
                # --------------------------------
                
                elif console.action == 'add_visitor':
                    
                    new_history.append(add_user(console.vars.user))
                
                elif console.action == 'add_project':
                    
                    for uid in console.vars.user:
                        new_history.append(add_user(uid))
                
                elif console.action == 'replace_visitor':
                    
                    # the current and replacement user ids are passed 
                    # from the validator in console.vars.user as a 2-list:
                    # [row id in research_visit_member to update, user_id to insert]
                    
                    rvm_record = db.research_visit_member(console.vars.user[0])
                    old_name = uname(rvm_record.user_id, rvm_record.id)
                    
                    replace_uid = console.vars.user[1]
                    replace_uid = None if replace_uid == 0 else replace_uid
                    rvm_record.update_record(user_id = replace_uid)
                    rvm_record = db.research_visit_member(console.vars.user[0])
                    new_name = uname(rvm_record.user_id, rvm_record.id)
                    
                    # TODO - could potentially merge reservation records here
                    
                    # update the history
                    new_history.append(' -- Visitor replaced: {} >> {}\\n'.format(old_name, new_name))
                
                # --------------------------------
                # TWO ACTIONS THAT BOOK ACCOMODATION
                # --------------------------------
            
                elif console.action == 'reserve_beds':
                
                    # loop over research_visit_member_ids, creating a dict of fields to insert
                    # via the __book_beds function
                    for rid in console.vars.records:
                    
                        rvm_record = db.research_visit_member(rid)
                        name = uname(rvm_record.user_id, rvm_record.id)

                        flds = dict(arrival_date = console.vars.accom_arrive,
                                    departure_date = console.vars.accom_depart,
                                    research_visit_id = rv_id,
                                    research_visit_member_id = rid)
                        
                        if console.vars.location == 'SAFE':
                            
                            __book_beds(site_table='bed_reservations_safe', flds =flds)
                            
                            new_history.append(' -- SAFE bed booked for {} from {} to {}\\n'.format(
                                                     name, console.vars.accom_arrive, console.vars.accom_depart))
                        
                        else:
                            flds.update(dict(type = console.vars.maliau_type,
                                             breakfast = console.vars.maliau_breakfast,
                                             lunch = console.vars.maliau_lunch,
                                             dinner = console.vars.maliau_dinner))
                            
                            __book_beds(site_table='bed_reservations_maliau', flds =flds)
                        
                            new_history.append(' -- Maliau bed booked for {} from {} to {} \\n'.format(
                                                     name, console.vars.accom_arrive, console.vars.accom_depart))
                
                elif console.action == 'release_beds':
                
                    # loop over research_visit_member_ids, creating a dict of fields to process
                    # via the __book_beds function
                    
                    for rid in console.vars.records:
                        
                        rvm_record = db.research_visit_member(rid)
                        name = uname(rvm_record.user_id, rvm_record.id)
                        
                        flds = dict(arrival_date = console.vars.accom_arrive,
                                    departure_date = console.vars.accom_depart,
                                    research_visit_member_id = rid)
                        
                        if console.vars.location == 'SAFE':
                            __release_beds(site_table='bed_reservations_safe', flds =flds)
                            new_history.append(' -- SAFE beds released for {} between {} to {}\\n'.format(
                                               name, console.vars.accom_arrive, console.vars.accom_depart))
                        else:
                            __release_beds(site_table='bed_reservations_maliau', flds =flds)
                            new_history.append(' -- Maliau beds released for {}, {} between {} to {} \\n'.format(
                                               name, console.vars.accom_arrive, console.vars.accom_depart))
                
                # --------------------------------
                # ACTION TO BOOK TRANSFERS
                # --------------------------------
                
                elif console.action == 'book_transfer':
                    
                    # loop over selected rows from user panel and book transfers
                    for rid in console.vars.records:
                        
                        rvm_record = db.research_visit_member(rid)
                        
                        db.transfers.insert(transfer = console.vars.transfer,
                                            research_visit_id = rv_id,
                                            research_visit_member_id = rid,
                                            transfer_date = console.vars.transfer_datepicker)
                        
                        name = uname(rvm_record.user_id, rvm_record.id)
                        new_history.append(' -- Transfer booked for {} from {} on {}\\n'.format(
                                                 name, console.vars.transfer, console.vars.transfer_datepicker))
                
                # --------------------------------
                # ACTION TO BOOK RAs
                # --------------------------------
                elif console.action == 'book_res_assist':
                
                    ropework = ((console.vars.ra_rope is not None) and (console.vars.ra_rope == 'on'))
                    nightwork = ((console.vars.ra_night is not None) and (console.vars.ra_night == 'on'))
                
                    db.research_assistant_bookings.insert(research_visit_id = rv_id,
                                        start_date = console.vars.ra_start,
                                        finish_date = console.vars.ra_end,
                                        site_time = console.vars.ra_site_time,
                                        work_type = console.vars.ra_work_type)
                                    
                    new_history.append(' -- RA booked: {} from {} to {}\\n'.format(
                                             console.vars.ra_site_time, console.vars.ra_start,
                                             console.vars.ra_end))
                # --------------------------------
                # SUBMIT 
                # --------------------------------
                elif console.action == 'submit_proposal':
                    
                    # i) update the history, change the status and redirect
                    hist_str = '[{}] {} {}\\n -- Proposal submitted\\n'
                    new_history = hist_str.format(datetime.datetime.utcnow().strftime('%Y-%m-%dT%H:%MZ'),
                                                               auth.user.first_name,
                                                               auth.user.last_name) + record.admin_history
            
                    record.update_record(admin_status = 'Submitted',
                                         admin_history = new_history)
            
                    # ii) email the proposer
                    SAFEmailer(to=auth.user.email,
                               subject='SAFE: research visit proposal submitted',
                               template =  'research_visit_submitted.html',
                               template_dict = {'name': auth.user.first_name,
                                                'url': URL('research_visits', 'research_visit_details',
                                                            args=[visit.vars.id], scheme=True, host=True)})
                
                    session.flash = CENTER(B('Research visit proposal submitted.'), _style='color: green')
                    
                else:
                    # datechange causes a processing event that doesn't do anything
                    pass
                
                # update the RV record to catch history changes
                if len(new_history) > 0:
                    history_update = '[{}] {} {}\\n'.format(datetime.datetime.utcnow().strftime('%Y-%m-%dT%H:%MZ'),
                                                            auth.user.first_name,
                                                            auth.user.last_name)
                    history_update += ''.join(new_history)
                    history_update += record.admin_history
                    record.update_record(admin_history = history_update)
                
                # reload the page to update changes
                redirect(URL('research_visit_details', args=rv_id, anchor='console'))
                
            elif console.errors:
                
                session.flash = console.errors
        
    # return the visit history
    if rv_id is not None and record.admin_history is not None:
        history = XML(record.admin_history.replace('\\n', '<br />'), sanitize=True, permitted_tags=['br/'])
        history = CAT(A(_name='history'),
                      DIV(DIV(H5('Research visit history'), _class="panel-heading"),
                          DIV(history, _class='panel_body', _style='margin:10px 10px;height:100px;overflow-y:scroll'),
                          DIV(_class='panel-footer'),
                          _class="panel panel-primary"))
    else:
        history = DIV()
        
    # If the visit record has been created and an admin is viewing, expose the 
    # decision panel
    if rv_id is not None and auth.has_membership('admin') and record.admin_status == 'Submitted':
        
        admin = admin_decision_form(selector_options=['Resubmit', 'Approved'])
        
        if admin.process(formname='admin').accepted:
            
            # update record with decision
            admin_str = '[{}] {} {}\\n ** Decision: {}\\n ** Comments: {}\\n'
            new_history = admin_str.format(datetime.datetime.utcnow().strftime('%Y-%m-%dT%H:%MZ'),
                                           auth.user.first_name,
                                           auth.user.last_name,
                                           admin.vars.decision,
                                           admin.vars.comment) + record.admin_history
            
            record.update_record(admin_status = admin.vars.decision,
                                  admin_history = new_history)
            
            # Email decision
            proposer = record.proposer_id
            template_dict = {'name': proposer.first_name, 
                             'url': URL('research_visits', 'research_visit_details', args=[rv_id], scheme=True, host=True),
                             'admin': auth.user.first_name + ' ' + auth.user.last_name}
            
            # pick an decision
            if admin.vars.decision == 'Approved':
                
                SAFEmailer(to=proposer.email,
                           subject='SAFE: research visit proposal approved',
                           template =  'research_visit_approved.html',
                           template_dict = template_dict)
                
            elif admin.vars.decision == 'Resubmit':
                
                SAFEmailer(to=proposer.email,
                           subject='SAFE: research visit proposal requires resubmission',
                           template =  'research_visit_resubmit.html',
                           template_dict = template_dict)
            
            # reload the page to update changes
            redirect(URL('research_visits','research_visit_details', args=rv_id, anchor='history'))
        
    else:
        admin = DIV()
    
    # Stick up a warning for when admins book stuff!
    if auth.has_membership('admin'):
        admin_warning = DIV(B('Warning!'), ' You are an administrator for the SAFE website. This allows you to',
                            ' edit all research visit proposals and to ignore the usual 14 days notice',
                            ' and SAFE bed limits. Be careful not to accidentally abuse this privilege when booking '
                            ' your own research visits!', _class="alert alert-info", _style='background:darkred', _role="alert")
    else:
        admin_warning = DIV()
        
    
    return dict(visit_record = record, visit=visit, instructions = instructions,
                console=console, history=history, admin=admin, admin_warning=admin_warning)



@auth.requires_membership('admin')
def create_late_research_visit():
  
    """
    Process to create an RV for a late booking user.
    This allows an admin to create an RV that then belongs to that user
    """
    
    # Two possible entry points:
    # - Completely new RV request(bare URL)
    # - Project specified for new RV request (project_id as a variable, but no record)
    
    new_rv_project_requested = request.vars['new']
    
    if new_rv_project_requested is not None and new_rv_project_requested != '0':
        new_project_record = db.project_id(new_rv_project_requested)
        if new_project_record is None:
            session.flash = B(CENTER('Invalid new visit project reference'), _style='color:red;')
            redirect(URL('research_visits','create_late_research_visit'))
    else:
        new_project_record = None
    
    # provide either a project drop down or a RV details form
    if new_rv_project_requested is None:
        
        # Bare URL submitted: provide a list of available projects + look see visit
        # and redirect back to the page, with the new project_id for the next step
        proj_query = db((db.project_id.id > 0) & 
                         (db.project_id.project_details_id == db.project_details.id))
        rows = proj_query.select(db.project_details.project_id, db.project_details.title)
        available_project_ids = [r.project_id for r in rows]
        available_project_titles = [r.title for r in rows]
        
        project_selector = SELECT(OPTION('Look see visit', _value='0'),
                                  *[OPTION(title, _value=pid) for title, pid in
                                    zip(available_project_titles, available_project_ids)],
                                  _class='form-control', _name='project_selector')
        
        visit= FORM(DIV(DIV(H5('Research visit summary'), _class="panel-heading"),
                        DIV(DIV(LABEL('Choose project:', _class="control-label col-sm-2" ),
                                DIV(project_selector, _class="col-sm-8"),
                                TAG.BUTTON('Select', _style="padding: 5px 15px", _class='col-sm-2',
                                            _type='submit', _name='submit_project_select'),
                                _class='row', _style='margin:10px 10px'),
                            _class='panel_body'),
                        _class="panel panel-primary"))
        
        if visit.validate():
            # reload the URL with the id of the new project as a variable
            redirect(URL('research_visits','create_late_research_visit', vars={'new': visit.vars.project_selector}))
            
    else: 
        
        # Now have a URL with a project requested as a variable 'research_visit_details?new='866' 
        buttons = [TAG.button('Create proposal',_type="submit",
                               _name='save_proposal', _style='padding: 5px 15px 5px 15px;')]
        
        # get the project details
        if new_rv_project_requested != '0':
            
            # get the title
            pid = int(new_rv_project_requested)
            proj_row = db((db.project_id.id == pid) & (db.project_id.project_details_id == db.project_details.id))
            project_details = proj_row.select().first()
            proj_title = project_details.project_details.title
            
            # restrict the proposer id to possible coordinators
            coords = db((db.auth_user.id == db.project_members.user_id) &
                        (db.project_members.project_id == pid) &
                        (db.project_members.is_coordinator == 'True'))
            db.research_visit.proposer_id.requires  = IS_IN_DB(coords, db.auth_user.id, '%(last_name)s, %(first_name)s', zero=None)
        else:
            project_details = None
            proj_title = 'Look see visit'
        
        # Use SQLFORM for DB input
        visit = SQLFORM(db.research_visit, 
                        fields = ['title','arrival_date',
                                  'departure_date','purpose',
                                  'licence_details', 'proposer_id'],
                        buttons = buttons,
                        showid = False)
        
        # process the visit form to create hidden fields and to process input
        if visit.process().accepted:
            
            # insert the project_id, but not if this is a look see visit
            if new_rv_project_requested != '0':
                db.research_visit(visit.vars.id).update_record(project_id = new_rv_project_requested)
            
            # email the identified coordinator the link for the page
            proposer = db.auth_user[visit.vars.proposer_id]
            SAFEmailer(to=proposer.email,
                        subject='SAFE: draft short notice research visit proposal created on your behalf',
                        template =  'research_visit_created.html',
                        template_dict = {'name': auth.user.first_name,
                                         'url': URL('research_visits', 'research_visit_details',
                                                    args=[visit.vars.id], scheme=True, host=True)})
                
            db.research_visit(visit.vars.id).update_record(admin_status = 'Draft')
            session.flash = CENTER(B('Research visit proposal created'), _style='color: green')
            
            redirect(URL('research_visits','research_visit_details', args=visit.vars.id))
        else:
            
            pass
    
        # Now repackage the form into a custom DIV
        # edit form widgets - notably, override the default date widget classes to allow
        # them to use the daterange datepicker
        visit.custom.widget.purpose['_rows'] = 4
        visit.custom.widget.arrival_date['_class'] = "form-control input-sm"
        visit.custom.widget.departure_date['_class'] = "form-control input-sm"
        
        proj_row = DIV(LABEL('Project title:', _class="control-label col-sm-2" ),
                       DIV(proj_title, _class="col-sm-10"),
                       _class='row', _style='margin:10px 10px')
        
        # javascript to run the datepicker without date constraints
        visit_js =datepicker_script(id = 'visit_datepicker',
                                    autoclose = 'true',
                                    startDate ='""',
                                    endDate ='""')
        
        visit = FORM(DIV(DIV(DIV(H5('Research visit summary', _class='col-sm-9'), _class='row', _style='margin:0px 0px'),
                            _class="panel-heading"),
                        DIV(visit.custom.begin, proj_row,
                            DIV(LABEL('Visit title:', _class="control-label col-sm-2" ),
                                DIV(visit.custom.widget.title,  _class="col-sm-10"),
                                _class='row', _style='margin:10px 10px'),
                            DIV(LABEL('Coordinator:', _class="control-label col-sm-2" ),
                                DIV(visit.custom.widget.proposer_id,  _class="col-sm-10"),
                                _class='row', _style='margin:10px 10px'),
                            DIV(LABEL('Dates:', _class="control-label col-sm-2" ),
                                DIV(DIV(visit.custom.widget.arrival_date,
                                        SPAN('to', _class="input-group-addon input-sm"),
                                        visit.custom.widget.departure_date,
                                        _class="input-daterange input-group", _id="visit_datepicker"),
                                    _class='col-sm-10'),
                                _class='row', _style='margin:10px 10px'),
                            DIV(LABEL('Purpose:', _class="control-label col-sm-2" ),
                                DIV(visit.custom.widget.purpose,  _class="col-sm-10"),
                                _class='row', _style='margin:10px 10px'),
                            DIV(DIV(visit.custom.submit,  _class="col-sm-10 col-sm-offset-2"),
                                _class='row', _style='margin:10px 10px'),
                            visit.custom.end,
                           _class='panel_body'),
                        _class="panel panel-primary"),
                        visit_js)
    
    return dict(visit=visit)

def uname(uid, rowid):
    
    """
    The RV_member table contains a link to a user that can be NULL
    - this helper function takes a user_id field (which references auth_user)
      and a RVM row id to give a formatted name for both named and unknown users
    """
    
    if uid is None:
        nm = 'Unknown #' + str(rowid)
    else:
        nm = uid.last_name + ", " + uid.first_name
        
    return nm

def validate_research_visit_details(form):
    
    """
    This controller checks the form that creates the initial RV entry
    """
    
    # populate the proposer_id if this is a new entry (form has no record)
    if form.record is None:
        form.vars.proposer_id = auth.user.id
        form.vars.proposal_date = datetime.datetime.utcnow().isoformat()
        form.vars.admin_history =  '[{}] {} {}\\n -- {}\\n'.format(datetime.datetime.utcnow().strftime('%Y-%m-%dT%H:%MZ'),
                    auth.user.first_name,
                    auth.user.last_name, 
                    'Research visit proposal created.')
    
    # check the arrival date is more than a fortnight away
    deadline = datetime.date.today() + datetime.timedelta(days=14)
    if form.vars.arrival_date < deadline and not auth.has_membership('admin'):
        form.errors.arrival_date = '14 days notice required. Arrival date must be later than {}.'.format(deadline.isoformat())
    
    # check the departure date is after the arrival date
    # TODO - think about day visits
    if form.vars.arrival_date >= form.vars.departure_date:
        form.errors.departure_date = 'The departure date must be later than the arrival date'
    


def validate_research_visit_details_console(form):
    
    """
    This function checks the wide range of actions available
    from the panels for a created RV - all the panels form one
    big form that share the visitor selection checkboxes so
    there are multiple submit buttons that use the information
    in the console form in different ways
    """
    
    # the request captures the datechange hidden field and 
    # the name of any submit button pressed
    request_keys = request.vars.keys()
    
    # retrieve the record for the related visit
    rv = db.research_visit(form.vars.id)
    
    # sanitize the visitor row selectors, which can be
    # - missing (none selected) 
    # - a single string (one selected)
    # - a list of strings (2+ selected)
    # so convert to a consistent list and make strings into numbers
    if form.vars.records is None:
        form.vars.records = [] 
    elif type(form.vars.records) is str:
        form.vars.records = [int(form.vars.records)] 
    else:
        form.vars.records = [int(x) for x in form.vars.records]
    
    # Two kinds of actions (delete_XXX and replace_visitor) operate
    # via submit buttons encoding row ids. These submit an empty value
    # but have a key name of format: delete_rowtype_rowid
    # e.g. delete_transfer_47 is a request to delete the transfer record
    # with id 47 
    
    form.action = None
    
    for k in request_keys:
        if "delete_" in k:
            
            form.action = 'delete'
            form.delete = k.split('_')[1:]
            
        elif "replace" in k:
            
            form.action = 'replace_visitor'
            form.vars.user = [int(k.split('_')[2]), int(form.vars.user)]
    
    # if no delete/replace action then look for other actions simply by matching strings
    if form.action is None:
        
        # list of submit buttons (and one hidden action)
        submit_ids = set(["add_visitor", "add_project",
                          "reserve_beds", "release_beds", "book_transfer", 
                          "book_res_assist", "submit_proposal", 'datechange'])
    
        action = list(submit_ids.intersection(request_keys))
    
        # check for oddities and pass the action back across to the processing actions
        if(len(action) <> 1):
            form.errors = 'Error with action identification.'
        else:
            action = action[0]
            form.action = action
        
        # catching and checking dates
        # - onchange event for dates triggers a submit with a hidden datechange field
        #   that we can intercept
        if action == 'datechange':
        
            # store the dates as datetime objects
            arrival_datetime = datetime.datetime.strptime(form.vars.arrival_date, '%Y-%m-%d').date()
            departure_datetime = datetime.datetime.strptime(form.vars.departure_date, '%Y-%m-%d').date()
        
            deadline = datetime.date.today() + datetime.timedelta(days=14)

            # check arrival date and departure date separately to allow defaults on each independently
            if arrival_datetime < deadline:
                # check the from date is more than a fortnight away
                form.errors.arrival_date = '14 days notice required for all bookings. Must be later than {}.'.format(deadline.isoformat())
            elif arrival_datetime < rv.arrival_date:
                # check we are inside the visit window.
                form.errors.arrival_date = 'This date is before the research visit arrival date ({}).'.format(rv.arrival_date.isoformat())
            else:
                # all good, so store the arrival date in session
                session.safe.arrival_datetime = arrival_datetime
                session.safe.arrival_date = form.vars.arrival_date
        
            if departure_datetime > rv.departure_date:
                form.errors.departure_date = 'This date is after the research visit departure date ({}).'.format(rv.departure_date.isoformat())
        
            elif arrival_datetime >= departure_datetime:
                # check the departure date is after the arrival date
                form.errors.departure_date = 'The departure date must be later than the arrival date'
            else:
                # all good, so store the departure date in session
                session.safe.departure_datetime = departure_datetime
                session.safe.departure_date = form.vars.departure_date
        
            if session.safe.departure_date is not None and session.safe.arrival_date is not None:
                __check_availability()
            
        elif action == 'add_visitor':
        
            # get the format correct 
            if form.vars.user == '0':
                form.vars.user = None
            else:
                form.vars.user = int(form.vars.user)
        
            # check to see they aren't already a member
            visit_members = db(db.research_visit_member.research_visit_id == form.vars.id).select(db.research_visit_member.user_id)
            m = [r.user_id for r in visit_members]
        
            # strip None from m to allow multiple unknown users
            m = [x for x in m if x is not None]
        
            if form.vars.user in m:
                form.errors.user = "User already a member of this research visit."
        
        elif action == 'add_project':
        
            # check this is a project visit (don't add everyone in the world!)
            if rv.project_id is None:
                form.errors.user = "Cannot add project members for look see visits."
            
            # get the project members and current visit members
            project_members = db(db.project_members.project_id == rv.project_id).select(db.project_members.user_id)
            visit_members = db(db.research_visit_member.research_visit_id == rv.id).select(db.research_visit_member.user_id)
            
            p = [r.user_id for r in project_members]
            r = [r.user_id for r in visit_members]
            new = set(p).difference(r)
            
            if len(new) > 0:
                form.vars.user = new
            else:
                form.errors.user = "No new project members to add."
        
        elif action == 'reserve_beds':
            
            if len(form.vars.records) == 0:
                form.errors.user = 'You must select visitors to book for.'
            
            if (form.vars.accom_arrive == '') or (form.vars.accom_depart == ''):
                form.errors.accom_arrive = 'You must set dates to book accomodation.'
        
        elif action == 'release_beds':
            
            if len(form.vars.records) == 0:
                form.errors.user = 'You must select which visitors to release beds for.'
            
            if (form.vars.accom_arrive == '') or (form.vars.accom_depart == ''):
                form.errors.accom_arrive = 'You must set dates to release accomodation.'
        
        elif action == 'book_transfer':
            
            if len(form.vars.records) == 0:
                form.errors.user = 'You must select visitors to transfer'
            
            if (form.vars.transfer_datepicker == ''):
                form.errors.transfer_datepicker = 'You must set a date to book transfers.'
        
        elif action == 'book_res_assist':
            
            if (form.vars.ra_start == '') or (form.vars.ra_end == ''):
                form.errors.ra_site_time = 'You must set dates to book research assistants.'


## -----------------------------------------------------------------------------
## HELPER FUNCTIONS - prtotected from being called as a webpage using __name()
## -----------------------------------------------------------------------------

def __check_availability():
    
    """
    Updates the availability information stored in session
    - called from check availability and post booking SAFE
    """
    
    if ((session.safe.arrival_date is not None) &
       (session.safe.arrival_date is not None)):
          # check how many reservations overlap the whole period
          existing_res = db((db.bed_reservations_safe.departure_date > session.safe.arrival_date) &
                            (db.bed_reservations_safe.arrival_date < session.safe.departure_date))
      
          # store the availability
          session.safe.avail = n_beds_available - existing_res.count()


def __book_beds(site_table, flds):
    
    """
    Function to look for existing records in a table and book 
    - expecting to be passed a dictionary of fields to insert 
      via **args and that those will contain some key checking
      fields: research_visit_member_id, arrival_date, departure_date
    """
    
    # look for existing bookings at this site that intersect this one
    existing = db((db[site_table].research_visit_member_id == flds['research_visit_member_id']) &
                  (db[site_table].departure_date >= flds['arrival_date']) &
                  (db[site_table].arrival_date <= flds['departure_date']))
    
    if existing.count() > 0:
        
        # otherwise find the existing selections that overlap and get the date ranges
        existing = existing.select()
        
        # need to convert dates from string
        arrival_date = datetime.datetime.strptime(flds['arrival_date'], '%Y-%m-%d').date()
        departure_date = datetime.datetime.strptime(flds['departure_date'], '%Y-%m-%d').date()
        
        arr_dates = [r.arrival_date for r in existing]
        dep_dates = [r.departure_date for r in existing]
        arr_dates.append(arrival_date)
        dep_dates.append(departure_date)
        flds['arrival_date'] = min(arr_dates)
        flds['departure_date'] = max(dep_dates)
        
        # delete the existing overlapping records
        for e in existing:
            e.delete_record()
        
    # add the new spanning record
    db[site_table].insert(**flds)



def __release_beds(site_table, flds):
    
    """
    Function to look for existing records in a table and book 
    - expecting to be passed a dictionary of fields and that those will contain
      some key checking fields: research_visit_member_id, arrival_date, departure_date
    """
    
    # look for existing bookings at this site that intersect this one
    existing = db((db[site_table].research_visit_member_id == flds['research_visit_member_id']) &
                  (db[site_table].departure_date >= flds['arrival_date']) &
                  (db[site_table].arrival_date <= flds['departure_date']))
    
    if existing.count() > 0:
        
        # can only remove if there are bookings that overlap
        existing = existing.select()
        
        # need to convert dates from string
        release_start = datetime.datetime.strptime(flds['arrival_date'], '%Y-%m-%d').date()
        release_end = datetime.datetime.strptime(flds['departure_date'], '%Y-%m-%d').date()
        
        for e in existing:
            
            arr_in_ex = e.arrival_date < release_start < e.departure_date 
            dep_in_ex = e.arrival_date < release_end < e.departure_date 
            spans = ((release_start < e.arrival_date) & 
                     (e.departure_date < release_end))
            
            # look at each one in turn to see whether to delete/truncate
            if (arr_in_ex and not dep_in_ex):
                     # 1) truncating the end of the visit
                     e.update_record(departure_date = release_start)
            elif (not arr_in_ex and dep_in_ex):
                     # 2) truncating the start of the visit
                     e.update_record(arrival_date = release_end)
            elif (spans):
                     # 3) visit completely covered so delete
                     e.delete_record()
            elif (arr_in_ex and dep_in_ex):
                     # 4) visit split by deletion period, so truncate and insert
                     current_end = e.departure_date
                     e.departure_date = release_start
                     db[site_table].insert(**db[site_table]._filter_fields(e)) 
                     e.arrival_date = release_end
                     e.departure_date = current_end
                     db[site_table].insert(**db[site_table]._filter_fields(e)) 
                     e.delete_record()
            else:
                # non-overlapping
                pass

def __book_release_beds(records, site, mode):
    
    
    # book a bed for each checked record
    for rid in records:
        
        # get the user id from this row in the research_visit_member table
        rid = int(rid)
        row = db.research_visit_member(rid)
        
        # look for existing bookings at this site that intersect this one
        existing = db((db.bed_reservations.research_visit_member_id == rid) &
                      (db.bed_reservations.departure_date >= session.safe.arrival_date) &
                      (db.bed_reservations.arrival_date <= session.safe.departure_date) &
                      (db.bed_reservations.site == site))
        
        if mode == 'book':
            
            if existing.count() == 0:
                # if nothing overlapping already exists, create a new entry
                db.bed_reservations.insert(site = site,
                                           research_visit_id = row.research_visit_id,
                                           research_visit_member_id = rid,
                                           arrival_date = session.safe.arrival_date,
                                           departure_date = session.safe.departure_date,
                                           user_id = row.user_id)
            else:
                # otherwise find everthing that overlaps and get the date ranges
                existing = existing.select()
                arr_dates = [r.arrival_date for r in existing]
                arr_dates.append(datetime.datetime.strptime(session.safe.arrival_date, '%Y-%m-%d').date())
                dep_dates = [r.departure_date for r in existing]
                dep_dates.append(datetime.datetime.strptime(session.safe.departure_date, '%Y-%m-%d').date())
            
                # delete the existing overlapping records
                for e in existing:
                    db(db.bed_reservations.id == e.id).delete()
                
                # add the new spanning record
                db.bed_reservations.insert(site = site,
                                           research_visit_id = row.research_visit_id,
                                           research_visit_member_id = rid,
                                           arrival_date = min(arr_dates),
                                           departure_date = max(dep_dates),
                                           user_id = row.user_id)
        elif mode == 'release':
            
            if existing.count() > 0:
                # can only remove if there are bookings that overlap
                # TODO - add flash if there are no actions?
                existing = existing.select()
                for e in existing:
                    
                    arr_in_ex = e.arrival_date < session.safe.arrival_datetime < e.departure_date 
                    dep_in_ex = e.arrival_date < session.safe.departure_datetime < e.departure_date 
                    spans = ((session.safe.arrival_datetime < e.arrival_date) & 
                             (e.departure_date < session.safe.departure_datetime))
                    
                    # look at each one in turn to see whether to delete/truncate
                    if (arr_in_ex and not dep_in_ex):
                             # 1) truncating the end of the visit
                             db(db.bed_reservations.id == e.id).update(departure_date = session.safe.arrival_date)
                    elif (not arr_in_ex and dep_in_ex):
                             # 2) truncating the start of the visit
                             db(db.bed_reservations.id == e.id).update(departure_date = session.safe.departure_date)
                    elif (spans):
                             # 3) visit completely covered so delete
                             db(db.bed_reservations.id == e.id).delete()
                    elif (arr_in_ex and dep_in_ex):
                             # 4) visit split by deletion period, so truncate and insert
                             db(db.bed_reservations.id == e.id).update(departure_date = session.safe.arrival_date)
                             db.bed_reservations.insert(site = site,
                                                        research_visit_id = row.research_visit_id,
                                                        research_visit_member_id = rid,
                                                        arrival_date = session.safe.departure_date,
                                                        departure_date = e.departure_date,
                                                        user_id = row.user_id)
                    else:
                        # non-overlapping
                        pass
                    

def date_range(start, end):
    
    days = []
    curr = start
    while curr <= end: 
        days.append(curr)
        curr += datetime.timedelta(days=1)
    
    return(days)

def export_my_research_visit():
    
    """
    This creates an excel workbook containing the intinerary for a single research
    visit and pokes it out as an http download, so a button can call the controller
    and return the file via the browser. Costs are loaded from the same costs 
    json used to populate the logistics and costs page, so can be updated from 
    a single location
    """
    
    # load costs from the json data
    f = os.path.join(request.folder, 'private','content/en/info/costs.json')
    costs_dict = simplejson.load(open(f))
    
    # set up the coordinates of the data block
    curr_row = start_row = 8
    data_start_col = 4
    
    # Get the specific visit requested
    rv_id = request.args(0)
    
    if rv_id is not None:
        record = db.research_visit(rv_id)
        if record is None:
            session.flash = CENTER(B('Export request for non existant research visit', _style='color:red'))
            return
    else:
        session.flash = CENTER(B('Export request for research visit missing a visit_id', _style='color:red'))
        return
    
    # GET THE TIMESCALE FOR THE RVs (which _should_ encompass all RV activities)
    
    start_all = record.arrival_date
    end_all   = record.departure_date
    
    # use start as an epoch to give column numbers
    dates  = date_range(start_all, end_all)
    dates_column = [(x - start_all).days + data_start_col for x in dates]
    
    # get column labels and get blocks to label months
    monthyear = [d.strftime('%B %Y') for d in dates]
    monthyear_set = set(monthyear)
    monthyear_dates =  [[i for i, x in zip(dates, monthyear) if x == my] for my in monthyear_set]
    monthyear_range = [[min(x), max(x)] for x in monthyear_dates]
    weekday = [d.strftime('%a') for d in dates]
    weekday = [x[0] for x in weekday]
    day = [d.day for d in dates]
    
    # SETUP THE WORKBOOK
    wb = openpyxl.Workbook()
    
    # spreadsheet styles
    left = openpyxl.styles.Alignment(horizontal='left')
    center = openpyxl.styles.Alignment(horizontal='center')
    weekend = openpyxl.styles.PatternFill(fill_type='solid', start_color='CCCCCC')
    head = openpyxl.styles.Font(size=14, bold=True)
    subhead = openpyxl.styles.Font(bold=True)
    warn = openpyxl.styles.Font(bold=True, color='FF0000')
    cell_shade = {'Approved': openpyxl.styles.PatternFill(fill_type='solid', start_color='8CFF88'),
                  'Submitted': openpyxl.styles.PatternFill(fill_type='solid', start_color='FFCB8B'),
                  'Draft': openpyxl.styles.PatternFill(fill_type='solid', start_color='DDDDDD'),
                  'Resubmit': openpyxl.styles.PatternFill(fill_type='solid', start_color='DDDDDD'),
                  'Rejected': openpyxl.styles.PatternFill(fill_type='solid', start_color='FF96C3')}
    
    # name the worksheet and add a heading
    ws = wb.active
    ws.title = 'Research visits'
    ws['A1'] = 'Research visit plans for the SAFE Project as of {}'.format(datetime.date.today().isoformat())
    ws['A1'].font = head
    
    title =  db((db.project_id.id == record.project_id) &
                (db.project_id.project_details_id == db.project_details.id))
    title = title.select(db.project_details.title).first().title
    ws['A2'] = 'Research visit #' + str(record.id) + ": " + title
    ws['A2'].font = subhead
    
    ws['A3'] = 'Costs are estimated and do not include site transport costs at SAFE'
    ws['A3'].font = warn
    
    # Populate the sheet with the date information
    # months in merged cells with shading on alternate months
    for start, end in monthyear_range:
        start_col = (start - start_all).days + data_start_col
        end_col = (end - start_all).days + data_start_col
        
        ws.merge_cells(start_row=5, start_column= start_col,
                       end_row=5, end_column=end_col)
        c = ws.cell(row=5, column= start_col)
        c.value = start.strftime('%B %Y')
        if (start.month % 2) == 1:
            c.fill = weekend
    
    # day of month and week day, colouring weekends and setting column width
    for i, d, w in zip(dates_column, day, weekday):
        c1 = ws.cell(row = 6, column = i)
        c1.value = d
        c1.alignment = center 
        
        c2 = ws.cell(row = 7, column = i)
        c2.value = w
        c2.alignment = center
        
        if w == 'S':
            c1.fill = weekend
            c2.fill = weekend
        
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 3
    
    # add left hand column headers
    ws['A6'] = 'ID'
    ws['B6'] = 'Type'
    ws['C6'] = 'Cost in RM'
    
    # write event function - don't use merged cells here
    # it seems tempting but text cannot overflow from merged
    # cell blocks into the adjacent blank cells, but it can from 
    # a single cell, so write content into first cell of range
    # and then just colour the rest of the date range.
    def write_event(row, start, end, id, type, content, status, cost):
        
        # range
        this_start_col = (start - start_all).days + data_start_col
        this_end_col = (end - start_all).days + data_start_col
        
        # put content at LHS of range
        c = ws.cell(row=row, column=this_start_col)
        c.value = content
        c.alignment = left # makes text overflow RHS of narrow ranges.
        
        for i in range(this_start_col, this_end_col + 1):
            c = ws.cell(row=row, column=i)
            c.fill = cell_shade[status]
        
        c = ws.cell(row=row, column=1)
        c.value = id
        c = ws.cell(row=row, column=2)
        c.value = type
        c = ws.cell(row=row, column=3)
        c.value = cost
    
    cost = '---'
    
    dat = [curr_row, record.arrival_date, record.departure_date, record.id, 'Research Visit',
           "Project " + str(record.project_id) + ": " + record.title, record.admin_status, cost]
    
    write_event(*dat)
    curr_row += 1
    
    # SAFE bed bookings
    safe_query =  ((db.bed_reservations_safe.research_visit_id == record.id) & 
                   (db.bed_reservations_safe.research_visit_member_id == db.research_visit_member.id))
    safe_data = db(safe_query).select(orderby=db.bed_reservations_safe.arrival_date)
    
    for r in safe_data:
    
        # check for unknown users
        name = uname(r.research_visit_member.user_id, r.research_visit_member.id)
    
        # calculate cost - food charge only
        cost = (r.bed_reservations_safe.departure_date - 
                r.bed_reservations_safe.arrival_date).days * costs_dict['safe_costs']['food']['cost']
    
        # put the list of info to be written together
        dat = [curr_row, r.bed_reservations_safe.arrival_date, 
               r.bed_reservations_safe.departure_date,
               r.bed_reservations_safe.research_visit_id, 
               'SAFE booking', name, record.admin_status, cost]
    
        # write it and move down a row
        write_event(*dat)
        curr_row += 1
    
    # MALIAU bed bookings
    maliau_query =  ((db.bed_reservations_maliau.research_visit_id == record.id) & 
                     (db.bed_reservations_maliau.research_visit_member_id == db.research_visit_member.id))
    maliau_data = db(maliau_query).select(orderby=db.bed_reservations_maliau.arrival_date)
    
    for r in maliau_data:
    
        name = uname(r.research_visit_member.user_id, r.research_visit_member.id)
    
        # calculate cost - entry, bed and food costs
        days = (r.bed_reservations_maliau.departure_date -
                r.bed_reservations_maliau.arrival_date).days
        # admin and conservation on entry
        cost = costs_dict['maliau_entry']['admin']['standard'] + costs_dict['maliau_entry']['cons']['standard']
        # annex/hostel are only alternatives at the moment
        if r.bed_reservations_maliau.type == 'Annex':
            cost += days * costs_dict['maliau_accom']['annex']['standard']
        else:
            cost += days * costs_dict['maliau_accom']['hostel']['standard'] 
        # food
        if r.bed_reservations_maliau.breakfast is True:
            cost +=  costs_dict['maliau_food']['breakfast']['standard'] * days
        if r.bed_reservations_maliau.lunch is True:
            cost += costs_dict['maliau_food']['lunch']['standard'] * days
        if r.bed_reservations_maliau.dinner is True:
            cost += costs_dict['maliau_food']['dinner']['standard'] * days
    
        # content 
        food_labels = ['B' if r.bed_reservations_maliau.breakfast else ''] + \
                      ['L' if r.bed_reservations_maliau.lunch else ''] + \
                      ['D' if r.bed_reservations_maliau.dinner else '']
        content = name + ' (' + r.bed_reservations_maliau.type + ','+ ''.join(food_labels) + ')'
        dat = [curr_row, r.bed_reservations_maliau.arrival_date,
               r.bed_reservations_maliau.departure_date, 
               r.bed_reservations_maliau.research_visit_id,
               'Maliau booking', content, record.admin_status, cost]
    
        write_event(*dat)
        curr_row += 1
    
    #TRANSFERS
    transfer_query =  ((db.transfers.research_visit_id == record.id) & 
                     (db.transfers.research_visit_member_id == db.research_visit_member.id))
    transfer_data = db(transfer_query).select(orderby=db.transfers.transfer_date)
    
    for r in transfer_data:
    
        name = uname(r.research_visit_member.user_id, r.research_visit_member.id)
    
        # costs
        if r.transfers.transfer in ['SAFE to Tawau','Tawau to SAFE']:
            cost = costs_dict['transfers']['tawau_safe']['cost']
        elif r.transfers.transfer in ['SAFE to Maliau','Maliau to SAFE']:
            cost = costs_dict['transfers']['safe_maliau']['cost']
        elif r.transfers.transfer in ['Tawau to Maliau','Maliau to Tawau']:
            cost = costs_dict['transfers']['tawau_maliau']['cost']
        elif r.transfers.transfer in ['SAFE to Danum','Danum to SAFE']:
            cost = costs_dict['transfers']['safe_danum']['cost']
        
        dat = [curr_row, r.transfers.transfer_date, r.transfers.transfer_date, 
               r.transfers.research_visit_id, 'Transfer', 
               name + ': ' + r.transfers.transfer, record.admin_status, cost]
    
        write_event(*dat)
        curr_row += 1
    
    # RA requests
    rassist_query =  ((db.research_assistant_bookings.research_visit_id == record.id))
    rassist_data = db(rassist_query).select(orderby=db.research_assistant_bookings.start_date)
    
    for r in rassist_data:
    
        # costs 
        if r.site_time in ['All day at SAFE', 'All day at Maliau']:
            cost = costs_dict['ra_costs']['full']
        else:
            cost = costs_dict['ra_costs']['half']
    
        cost = cost[r.work_type]
    
        dat = [curr_row, r.start_date, r.finish_date, r.research_visit_id, 'RA booking',
               r.site_time, record.admin_status, cost]
    
        write_event(*dat)
        curr_row += 1

    
    # freeze the rows
    c = ws.cell(row=start_row, column=data_start_col)
    ws.freeze_panes = c
    # and now poke the workbook object out to the browser
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    attachment = 'attachment;filename=SAFE_Bed_reservations_{}.xlsx'.format(datetime.date.today().isoformat())
    response.headers['Content-Disposition'] = attachment
    content = openpyxl.writer.excel.save_virtual_workbook(wb)
    raise HTTP(200, str(content),
               **{'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                  'Content-Disposition':attachment + ';'})


class summary_tracker():
    """
    This class simply keeps track of the number of people requesting
    different resources across a time range. The update method allows
    this tracking to be incorporated into the individual row loops when
    Excel or text reports are being created.
    """
    
    def __init__(self, start, end):
        
        # Initialise ordered dictionary to set the output order
        dates = date_range(start, end)
        counts = [("Beds requested at SAFE", {d: 0 for d in dates}),
                  ("Beds requested at Maliau", {d: 0 for d in dates}),
                  ("RAs requested at SAFE", {d: 0 for d in dates}),
                  ("RAs requested at Maliau", {d: 0 for d in dates})]
        self.summary = OrderedDict(counts)
        
        # add all the transfer types
        for t in transfer_set:
            self.summary[t] = {d: 0 for d in dates}
    
    def update(self, k, start, end):
        
        dates = date_range(start, end)
        for d in dates:
            self.summary[k][d] += 1


def export_ongoing_research_visits():
    
    """
    This creates an excel workbook compiling ongoing and future research visit data
    and pokes it out as an http download, so a button can call the controller
    and return the file via the browser. Costs are loaded from the same costs 
    json used to populate the logistics and costs page, so can be updated from 
    a single location
    """
    
    # load costs from the json data
    f = os.path.join(request.folder, 'private','content/en/info/costs.json')
    costs_dict = simplejson.load(open(f))
    
    # set up the coordinates of the data block
    curr_row = start_row = 13
    data_start_col = 4
    
    # GET THE ONGOING RVs 
    today =datetime.date.today()
    
    rv_query = (db.research_visit.departure_date >= today)
    
    # no records?
    if db(rv_query).count() == 0:
        session.flash = CENTER(B('No research visit data found', _style='color:red'))
        redirect(URL('research_visits','research_visits'))
    else:
        # grab the data from those queries starting with the earliest arrivals
        rv_data = db(rv_query).select(orderby=db.research_visit.arrival_date)
    
    # GET A COMMON TIME SCALE FROM THE RVs (which _should_ encompass all RV activities)
    start_all = min([r.arrival_date for r in rv_data])
    end_all   = max([r.departure_date for r in rv_data])
    
    # use start as an epoch to give column numbers
    dates  = date_range(start_all, end_all)
    dates_column = [(x - start_all).days + data_start_col for x in dates]
    
    # get column labels and get blocks to label months
    monthyear = [d.strftime('%B %Y') for d in dates]
    monthyear_set = set(monthyear)
    monthyear_dates =  [[i for i, x in zip(dates, monthyear) if x == my] for my in monthyear_set]
    monthyear_range = [[min(x), max(x)] for x in monthyear_dates]
    weekday = [d.strftime('%a') for d in dates]
    weekday = [x[0] for x in weekday]
    day = [d.day for d in dates]
    
    # SETUP THE WORKBOOK
    wb = openpyxl.Workbook()
    
    # spreadsheet styles
    left = openpyxl.styles.Alignment(horizontal='left')
    center = openpyxl.styles.Alignment(horizontal='center')
    weekend = openpyxl.styles.PatternFill(fill_type='solid', start_color='CCCCCC')
    head = openpyxl.styles.Font(size=14, bold=True)
    subhead = openpyxl.styles.Font(bold=True)
    warn = openpyxl.styles.Font(bold=True, color='FF0000')
    zero_summary = openpyxl.styles.Font(color='BBBBBB')
    non_zero_summary = openpyxl.styles.Font(bold=True)
    
    cell_shade = {'Approved': openpyxl.styles.PatternFill(fill_type='solid', start_color='8CFF88'),
                  'Submitted': openpyxl.styles.PatternFill(fill_type='solid', start_color='FFCB8B'),
                  'Draft': openpyxl.styles.PatternFill(fill_type='solid', start_color='DDDDDD'),
                  'Resubmit': openpyxl.styles.PatternFill(fill_type='solid', start_color='DDDDDD'),
                  'Rejected': openpyxl.styles.PatternFill(fill_type='solid', start_color='FF96C3')}
    
    # name the worksheet and add a heading
    ws = wb.active
    ws.title = 'Research visits'
    ws['A1'] = 'Research visit plans for the SAFE Project as of {}'.format(today.isoformat())
    ws['A1'].font = head
    
    # Subheading 
    ws['A2'] = 'All research visits'
    ws['A2'].font = subhead
    
    ws['A3'] = 'Costs are estimated and do not include site transport costs at SAFE'
    ws['A3'].font = warn
    
    # Populate the sheet with the date information
    # months in merged cells with shading on alternate months
    for start, end in monthyear_range:
        start_col = (start - start_all).days + data_start_col
        end_col = (end - start_all).days + data_start_col
        
        ws.merge_cells(start_row=5, start_column= start_col,
                       end_row=5, end_column=end_col)
        c = ws.cell(row=5, column= start_col)
        c.value = start.strftime('%B %Y')
        if (start.month % 2) == 1:
            c.fill = weekend
    
    # day of month and week day, colouring weekends and setting column width
    for i, d, w in zip(dates_column, day, weekday):
        c1 = ws.cell(row = 6, column = i)
        c1.value = d
        c1.alignment = center 
        
        c2 = ws.cell(row = 7, column = i)
        c2.value = w
        c2.alignment = center
        
        if w == 'S':
            c1.fill = weekend
            c2.fill = weekend
        
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 3
    
    # add left hand column headers
    ws['A6'] = 'ID'
    ws['B6'] = 'Type'
    ws['C6'] = 'Cost in RM'
    
    # write event function - don't use merged cells here
    # it seems tempting but text cannot overflow from merged
    # cell blocks into the adjacent blank cells, but it can from 
    # a single cell, so write content into first cell of range
    # and then just colour the rest of the date range.
    def write_event(row, start, end, id, type, content, status, cost):
        
        # range
        this_start_col = (start - start_all).days + data_start_col
        this_end_col = (end - start_all).days + data_start_col
        
        # put content at LHS of range
        c = ws.cell(row=row, column=this_start_col)
        c.value = content
        c.alignment = left # makes text overflow RHS of narrow ranges.
        
        for i in range(this_start_col, this_end_col + 1):
            c = ws.cell(row=row, column=i)
            c.fill = cell_shade[status]
        
        c = ws.cell(row=row, column=1)
        c.value = id
        c = ws.cell(row=row, column=2)
        c.value = type
        c = ws.cell(row=row, column=3)
        c.value = cost
    
    # initialise a summary tracker using the full date range
    summary = summary_tracker(start_all, end_all)
    
    # loop over the research visits.
    for v in rv_data:
        
        
        cost = '---'
        
        dat = [curr_row, v.arrival_date, v.departure_date, v.id, 'Research Visit',
               "Project " + str(v.project_id) + ": " + v.title, v.admin_status, cost]
        
        write_event(*dat)
        curr_row += 1
        
        # SAFE bed bookings
        safe_query =  ((db.bed_reservations_safe.research_visit_id == v.id) & 
                       (db.bed_reservations_safe.research_visit_member_id == db.research_visit_member.id))
        safe_data = db(safe_query).select(orderby=db.bed_reservations_safe.arrival_date)
        
        for r in safe_data:
        
            # check for unknown users
            name = uname(r.research_visit_member.user_id, r.research_visit_member.id)
        
            # calculate cost - food charge only
            cost = (r.bed_reservations_safe.departure_date - 
                    r.bed_reservations_safe.arrival_date).days * costs_dict['safe_costs']['food']['cost']
        
            # put the list of info to be written together
            dat = [curr_row, r.bed_reservations_safe.arrival_date, 
                   r.bed_reservations_safe.departure_date,
                   r.bed_reservations_safe.research_visit_id, 
                   'SAFE booking', name, v.admin_status, cost]
        
            # write it, update the summary tracker and move down a row
            write_event(*dat)
            summary.update('Beds requested at SAFE', 
                           r.bed_reservations_safe.arrival_date, 
                           r.bed_reservations_safe.departure_date - datetime.timedelta(days=1))
            curr_row += 1
        
        # MALIAU bed bookings
        maliau_query =  ((db.bed_reservations_maliau.research_visit_id == v.id) & 
                         (db.bed_reservations_maliau.research_visit_member_id == db.research_visit_member.id))
        maliau_data = db(maliau_query).select(orderby=db.bed_reservations_maliau.arrival_date)
        
        for r in maliau_data:
        
            name = uname(r.research_visit_member.user_id, r.research_visit_member.id)
        
            # calculate cost - entry, bed and food costs
            days = (r.bed_reservations_maliau.departure_date -
                    r.bed_reservations_maliau.arrival_date).days
            # admin and conservation on entry
            cost = costs_dict['maliau_entry']['admin']['standard'] + costs_dict['maliau_entry']['cons']['standard']
            # annex/hostel are only alternatives at the moment
            if r.bed_reservations_maliau.type == 'Annex':
                cost += days * costs_dict['maliau_accom']['annex']['standard']
            else:
                cost += days * costs_dict['maliau_accom']['hostel']['standard'] 
            # food
            if r.bed_reservations_maliau.breakfast is True:
                cost +=  costs_dict['maliau_food']['breakfast']['standard'] * days
            if r.bed_reservations_maliau.lunch is True:
                cost += costs_dict['maliau_food']['lunch']['standard'] * days
            if r.bed_reservations_maliau.dinner is True:
                cost += costs_dict['maliau_food']['dinner']['standard'] * days
        
            # content 
            food_labels = ['B' if r.bed_reservations_maliau.breakfast else ''] + \
                          ['L' if r.bed_reservations_maliau.lunch else ''] + \
                          ['D' if r.bed_reservations_maliau.dinner else '']
            content = name + ' (' + r.bed_reservations_maliau.type + ','+ ''.join(food_labels) + ')'
            dat = [curr_row, r.bed_reservations_maliau.arrival_date,
                   r.bed_reservations_maliau.departure_date, 
                   r.bed_reservations_maliau.research_visit_id,
                   'Maliau booking', content, v.admin_status, cost]
        
            write_event(*dat)
            summary.update('Beds requested at Maliau', 
                           r.bed_reservations_maliau.arrival_date, 
                           r.bed_reservations_maliau.departure_date - datetime.timedelta(days=1))
            
            curr_row += 1
        
        #TRANSFERS
        transfer_query =  ((db.transfers.research_visit_id == v.id) & 
                         (db.transfers.research_visit_member_id == db.research_visit_member.id))
        transfer_data = db(transfer_query).select(orderby=db.transfers.transfer_date)
        
        for r in transfer_data:
        
            name = uname(r.research_visit_member.user_id, r.research_visit_member.id)
        
            # costs
            if r.transfers.transfer in ['SAFE to Tawau','Tawau to SAFE']:
                cost = costs_dict['transfers']['tawau_safe']['cost']
            elif r.transfers.transfer in ['SAFE to Maliau','Maliau to SAFE']:
                cost = costs_dict['transfers']['safe_maliau']['cost']
            elif r.transfers.transfer in ['Tawau to Maliau','Maliau to Tawau']:
                cost = costs_dict['transfers']['tawau_maliau']['cost']
            elif r.transfers.transfer in ['SAFE to Danum','Danum to SAFE']:
                cost = costs_dict['transfers']['safe_danum']['cost']
            
            dat = [curr_row, r.transfers.transfer_date, r.transfers.transfer_date,
                   r.transfers.research_visit_id, 'Transfer', 
                   name + ': ' + r.transfers.transfer, v.admin_status, cost]
            
            write_event(*dat)
            # update the transfers summary, keying by the transfer type
            summary.update(r.transfers.transfer, 
                           r.transfers.transfer_date,
                           r.transfers.transfer_date)
            curr_row += 1
        
        # RA requests
        rassist_query =  ((db.research_assistant_bookings.research_visit_id == v.id))
        rassist_data = db(rassist_query).select(orderby=db.research_assistant_bookings.start_date)
        
        for r in rassist_data:
        
            # costs 
            if r.site_time in ['All day at SAFE', 'All day at Maliau']:
                cost = costs_dict['ra_costs']['full']
            else:
                cost = costs_dict['ra_costs']['half']
            
            cost = cost[r.work_type]
            
            dat = [curr_row, r.start_date, r.finish_date, r.research_visit_id, 'RA booking',
                   r.site_time, v.admin_status, cost]
            
            write_event(*dat)
            
            if r.site_time in ['All day at SAFE', 'Morning only at SAFE', 'Afternoon only at SAFE']:
                summary.update('RAs requested at SAFE', r.start_date, r.finish_date)
            else:
                summary.update('RAs requested at Maliau', r.start_date, r.finish_date)
            
            curr_row += 1
    
    # Insert the summary information
    summary_row = 8
    
    # set the order 
    for k in summary.keys():
        
        c = ws.cell(row=summary_row, column=2)
        c.value = summary_labels[k]
        
        for d, c in zip(dates, dates_column):
            c = ws.cell(row=summary_row, column=c)
            c.value = summary.summary[k][d]
            if summary.summary[k][d] == 0:
                c.font = zero_summary
            else:
                c.font = non_zero_summary
                
        summary_row += 1
    
    # freeze the rows
    c = ws.cell(row=start_row, column=data_start_col)
    ws.freeze_panes = c
    
    # and now poke the workbook object out to the browser
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    attachment = 'attachment;filename=SAFE_Bed_reservations_{}.xlsx'.format(datetime.date.today().isoformat())
    response.headers['Content-Disposition'] = attachment
    content = openpyxl.writer.excel.save_virtual_workbook(wb)
    raise HTTP(200, str(content),
               **{'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                  'Content-Disposition':attachment + ';'})


def export_ongoing_research_visits_text():
    
    """
    This just creates text compiling ongoing and future research visit data
    and pokes it out as an http download, so a button can call the controller
    and return the file via the browser. Costs are loaded from the same costs 
    json used to populate the logistics and costs page, so can be updated from 
    a single location
    """
    
    # load costs from the json data
    f = os.path.join(request.folder, 'private','content/en/info/costs.json')
    costs_dict = simplejson.load(open(f))
    
    # GET THE ONGOING RVs 
    today =datetime.date.today()
    
    rv_query = (db.research_visit.departure_date >= today)
    
    # no records?
    if db(rv_query).count() == 0:
        session.flash = CENTER(B('No research visit data found', _style='color:red'))
        redirect(URL('research_visits','research_visits'))
    else:
        # grab the data from those queries starting with the earliest arrivals
        rv_data = db(rv_query).select(orderby=db.research_visit.arrival_date)
    
    # get the date range for all the activities and create a summary tracker
    start_all = min([r.arrival_date for r in rv_data])
    end_all   = max([r.departure_date for r in rv_data])
    summary = summary_tracker(start_all, end_all)
    
    # SETUP THE TEXT FILE AND DETAILS FILE. Because the summary is populated
    # while the details are being written out, and because we want the summary
    # at the top, keep two streams and then merge
    
    output = StringIO.StringIO()
    details = StringIO.StringIO()
    output.write('Research visit plans for the SAFE Project as of {}\n\n'.format(today.isoformat()))
    
    # loop over the research visits.
    for v in rv_data:
        
        details.write("\nProject " + str(v.project_id) + ": " + v.title  + ' [Status: ' + v.admin_status + ']\n')

        # SAFE bed bookings
        safe_query =  ((db.bed_reservations_safe.research_visit_id == v.id) & 
                       (db.bed_reservations_safe.research_visit_member_id == db.research_visit_member.id))
        safe_data = db(safe_query).select(orderby=db.bed_reservations_safe.arrival_date)
        
        for r in safe_data:
        
            # check for unknown users
            name = uname(r.research_visit_member.user_id, r.research_visit_member.id)
            details.write('  SAFE Booking: ' + r.bed_reservations_safe.arrival_date.isoformat() + ' -- ' +
                         r.bed_reservations_safe.departure_date.isoformat() + ' for ' + name + '\n')
            summary.update('Beds requested at SAFE', 
                           r.bed_reservations_safe.arrival_date, 
                           r.bed_reservations_safe.departure_date - datetime.timedelta(days=1))
        
        # MALIAU bed bookings
        maliau_query =  ((db.bed_reservations_maliau.research_visit_id == v.id) & 
                         (db.bed_reservations_maliau.research_visit_member_id == db.research_visit_member.id))
        maliau_data = db(maliau_query).select(orderby=db.bed_reservations_maliau.arrival_date)
        
        for r in maliau_data:
        
            name = uname(r.research_visit_member.user_id, r.research_visit_member.id)
            
            # content 
            food_labels = ['B' if r.bed_reservations_maliau.breakfast else ''] + \
                          ['L' if r.bed_reservations_maliau.lunch else ''] + \
                          ['D' if r.bed_reservations_maliau.dinner else '']
            content = name + ' (' + r.bed_reservations_maliau.type + ','+ ''.join(food_labels) + ')'
            
            details.write('  Maliau Booking: ' + r.bed_reservations_maliau.arrival_date.isoformat() + ' -- ' +
                         r.bed_reservations_maliau.departure_date.isoformat() + ' for ' + content + '\n')
            summary.update('Beds requested at Maliau', 
                           r.bed_reservations_maliau.arrival_date, 
                           r.bed_reservations_maliau.departure_date - datetime.timedelta(days=1))
            
            
        #TRANSFERS
        transfer_query =  ((db.transfers.research_visit_id == v.id) & 
                         (db.transfers.research_visit_member_id == db.research_visit_member.id))
        transfer_data = db(transfer_query).select(orderby=db.transfers.transfer_date)
        
        for r in transfer_data:
        
            name = uname(r.research_visit_member.user_id, r.research_visit_member.id)
            details.write('  Transfer: ' + r.transfers.transfer_date.isoformat() + ' from ' + 
                         r.transfers.transfer + ' for ' + name + '\n')
            # update the transfers summary, keying by the transfer type
            summary.update(r.transfers.transfer, 
                           r.transfers.transfer_date,
                           r.transfers.transfer_date)
        
        # RA requests
        rassist_query =  ((db.research_assistant_bookings.research_visit_id == v.id))
        rassist_data = db(rassist_query).select(orderby=db.research_assistant_bookings.start_date)
        
        for r in rassist_data:
            details.write('  RA time: ' + r.start_date.isoformat() + ' -- ' + r.finish_date.isoformat() + 
                         ' (' + r.site_time + ', ' + r.work_type + ')\n')
            
            if r.site_time in ['All day at SAFE', 'Morning only at SAFE', 'Afternoon only at SAFE']:
                summary.update('RAs requested at SAFE', r.start_date, r.finish_date)
            else:
                summary.update('RAs requested at Maliau', r.start_date, r.finish_date)
    
    # now add the summary to the output and then transfer the details at the bottom
    output.write('Daily resource request totals for current projects\n'
                 '==================================================\n\n')
    
    dates = date_range(start_all, end_all)
    for d in dates:
        # collect non zero counts as string and omit days where nothing happens
        counts = ['    ' + str(summary.summary[k][d]) + " " + k 
                  for k in summary.summary.keys() 
                  if summary.summary[k][d] > 0]
        if len(counts) > 0:
            output.write(d.strftime('%a %d %b %Y') + '\n' + '\n'.join(counts) + '\n')
    
    # now add the details at the end
    output.write('\n\nProject by project details\n'
                 '==========================\n\n')
    output.write(details.getvalue())
    details.close()
    
    # and now poke the text object out to the browser
    response.headers['Content-Type'] = 'text/plain'
    attachment = 'attachment;filename=SAFE_Research_Visit_details_{}.txt'.format(datetime.date.today().isoformat())
    response.headers['Content-Disposition'] = attachment
    raise HTTP(200, output.getvalue(),
               **{'Content-Type':'text/plain',
                  'Content-Disposition':attachment + ';'})
    
    output.close()


def safe_bed_availability():

    """
    This controller:
        - creates data for a free beds view using fullcalendar javascript
        - combining this and the booking on a single page causes issues
    """

    # get the dates when beds are booked and the status of the RV that
    # they are associated with
    # and select as an iterable of rows
    bed_data = db(db.bed_reservations_safe.research_visit_id == db.research_visit.id)
    rows = bed_data.select()
    
    # calculate the beds available by date
    approved = []
    pending = []
    
    for r in rows:
        
        dates = date_range(r.bed_reservations_safe.arrival_date,
                           r.bed_reservations_safe.departure_date)
                           
        if r.research_visit.admin_status == 'Approved':
            approved.extend(dates)
        else:
            pending.extend(dates)
    
    # now tabulate and align to get days
    pending =  Counter(pending)
    approved =  Counter(approved)
    
    # get unique days across pending and approved
    dates = set(pending.keys() + approved.keys())
    pend  = [0 if pending[d] is None else pending[d] for d in dates]
    conf  = [0 if approved[d] is None else approved[d] for d in dates]
    avail = [n_beds_available - (x+y) for x, y in zip(pend, conf)]
    
    # handle admin approved overbooking by truncating
    avail = [0 if x < 0 else x for x in avail]
    date = [d.isoformat() for d in dates]
    
    # now zip into sets of events, with three per day
    # one for each of pending confirmed and available
    n_events = len(dates)
    event_n_beds = pend + conf + avail
    event_class  = ['pending'] * n_events + ['confirmed'] * n_events + ['available'] * n_events
    event_title  = [ str(x) + ' ' + y for x,y in zip(event_n_beds, event_class)]
    event_order  = [2] * n_events + [3] * n_events + [1] * n_events
    # pass colour information
    event_backgrounds = ['#CC9900'] * n_events + ['#CC0000'] * n_events + ['#228B22'] * n_events
    
    events = zip(event_title, date * 3, event_class, event_order, event_backgrounds)
    
    return dict(events=events)

## -----------------------------------------------------------------------------
## ADMINISTER NEW VISITS
## -----------------------------------------------------------------------------


# decorator restricts access to admin users
# - the link is only revealed in the menu for admin users but 
#   that doesn't prevent pasting in the link!
@auth.requires_membership('admin')
def administer_research_visits():
    
    """
    This controller handles:
     - presenting admin users with a list of current submitted proposals for research visits 
     - allows the admin to approve or reject visit requests but also gives write access 
       and admins have unlocked constraints on dates and beds, so can add to a users proposal
    """
    
    links = [dict(header = '', body = lambda row: approval_icons[row.admin_status]),
             dict(header = '', body = lambda row: A('Details',_class='button btn btn-default'
                  ,_href=URL("research_visits","research_visit_details", args=[row.id])))
            ]
    
    db.research_visit.admin_status.readable = False
    
    # get a query of pending requests with user_id
    form = SQLFORM.grid(query=(db.research_visit.admin_status == 'Submitted'), csv=False,
                        fields=[db.research_visit.project_id, 
                                db.research_visit.title,
                                db.research_visit.admin_status],
                         maxtextlength=250,
                         deletable=False,
                         editable=False,
                         create=False,
                         details=False,
                         links=links
                         )
    
    return dict(form=form)
