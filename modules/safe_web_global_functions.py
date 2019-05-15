import os
import datetime
import cStringIO
from collections import Counter
import simplejson
from openpyxl import styles, utils, Workbook, writer
import html2text
import gluon.contrib.fpdf as FPDF

"""
This module holds a set of functions that are shared between controllers

The web2py HTML helpers are provided by gluon. This also provides the 'current' object, which
provides the web2py 'request' API (note the single letter difference from the requests package!).
The 'current' object is also extended by models/db.py to include the current 'db' DAL object
and the 'myconf' AppConfig object so that they can accessed by this module
"""

from gluon import *

"""
Formatting functions for SQLFORM grids, taking values from 
form rows and wrapping them, either using their represent value
or by creating a dictionary for a links object
"""


def thumbnail(value, default, size=100):
    """
    A function to to crop an image within a DIV to make a fixed size
    for a stack of images. It is expecting to be called from a lambda
    expression for the represent of an image field:
    
    db.project_details.thumbnail_figure.represent = lambda value, row: _thumb(value)
    
    """
    if value is None or value == '':
        url = URL('static', 'images/default_thumbnails/' + default) 
    else:
        url = URL('default', 'download', args=value)
    
    return DIV(_style=('background: url(' + url + ') 50% 50% no-repeat; '
                       'background-size: cover;'
                       'width: ' + str(size) + 'px; '
                       'height: ' + str(size) + 'px;'))


def link_button(c, f, arg, text="View", icon="glyphicon glyphicon-zoom-in"):

    """
    A function to provide a dictionary for use in the links argument of
    a SQLFORM.grid. The dictionary creates a button styled link to a specified
    controller (c), function (f) and row property as a link argument.
    """

    return dict(header='',
                body=lambda row: A(SPAN('', _class=icon), XML('&nbsp'),
                                   SPAN(text, _class="buttontext button"),
                                   _class="button btn btn-default",
                                   _href=URL(c, f, args=[row[arg]]),
                                   _style='padding: 3px 5px 3px 5px;'))

def get_frm():
    """
    Function to provide the auth_user row for the current Field Research Manager
    """
    
    # global object containing the Field Research Manager details
    frm = current.db((current.db.contacts.contacts_role == 'Field Research Manager') &
                     (current.db.auth_user.id == current.db.contacts.user_id)
                     ).select(current.db.auth_user.ALL).first()
    
    return frm


def datepicker_script(html_id, **settings):

    """
    This function generates a JS script, keyed to the id of a HTML entity
    that loads it as a datepicker object and allows settings on that datepicker object

    It also loads a set of public holidays into the datepickers and highlights them.

    Finally, it embeds a style class for the highlight, which isn't a horrible kludge
    at all and certainly isn't reflective of my general attitude to CSS in setting this up...
    """

    db = current.db
    session = current.session

    # load the db table into session if it isn't there already
    if not session.public_holidays:
        holidays = db(db.public_holidays).select()
        # get arrays of non-zero padded dates in java friendly D/M/YYYY and titles
        dates = [x.date.strftime('%-d/%-m/%Y') for x in holidays]
        titles = [x.title for x in holidays]
        session.public_holidays = {'dates': str(dates), 'titles': str(titles)}

    # now create the script text
    if settings:
        settings_str = ',\n'.join(item[0] + ':' + str(item[1]) for item in settings.iteritems())
    else:
        settings_str = ''

    javascript = """
        $('head').append($('<link  href="{cssurl}" type="text/css" rel="stylesheet" />'));
        var sabah_holiday_dates = {holiday_dates};
        var sabah_holiday_names = {holiday_titles};
        $.getScript('{scripturl}').done(function(){{
            $('#{html_id}').datepicker({{
                format: w2p_ajax_date_format.replace('%Y', 'yyyy').replace('%m', 
                    'mm').replace('%d', 'dd'),
                {settings},
                beforeShowDay: function(date){{
                    var d = date;
                    var curr_date = d.getDate();
                    var curr_month = d.getMonth() + 1; //Months are zero based
                    var curr_year = d.getFullYear();
                    var formattedDate = curr_date + "/" + curr_month + "/" + curr_year
                    var holiday_index = $.inArray(formattedDate, sabah_holiday_dates) 
                    if (holiday_index != -1) {{
                        return {{
                            classes: 'holiday',
                            tooltip: sabah_holiday_names[holiday_index]
                        }};
                    }}
                    return;
                }}
            }})
        }});
        """
    
    javascript = javascript.format(
        cssurl=URL('static', 'plugin_bs_datepicker/datepicker.css'),
        scripturl=URL('static', 'plugin_bs_datepicker/bootstrap-datepicker.js'),
        html_id=html_id,
        settings=settings_str,
        holiday_dates=session.public_holidays['dates'],
        holiday_titles=session.public_holidays['titles']
    )

    # return CAT(STYLE(XML('.holiday {background: #FE9781;}')), SCRIPT(javascript))
    return CAT(STYLE('.holiday {background: #FE9781;}'), SCRIPT(javascript))

def admin_decision_form(selector_options):

    """
    Global function to create an admin decision form, as used in nearly every
    details page.
    """
    selector = SELECT(selector_options, _class="generic-widget form-control", _name='decision')
    comments = TEXTAREA(_type='text', _class="form-control string", _rows=3, _name='comment')
    submit = TAG.BUTTON('Submit', _type="submit", _class="button btn btn-default",
                        _style='padding: 5px 15px 5px 15px;')

    admin = FORM(DIV(DIV(H5('Admin Decision', ), _class="panel-heading"),
                     DIV(DIV(DIV(LABEL('Select decision', _class='row'),
                                 DIV(selector, _class='row'),
                                 DIV(_class='row', _style='padding:5px'),
                                 DIV(submit,	 _class='row'),
                                 _class='col-sm-2'),
                             DIV(LABEL('Comments', _class='row'),
                                 DIV(comments, _class='row'),
                                 _class='col-sm-9 col-sm-offset-1'),
                             _class='row', _style='margin:10px 10px'),
                         _class='panel_body', _style='margin:10px 10px'),
                     DIV(_class="panel-footer"),
                     _class='panel panel-primary'))

    return admin


def safe_mailer(subject, to, template, template_dict, cc=None,
                reply_to='info@safeproject.net', cc_info=True, bcc=None,
                attachment_string_objects=None):

    """
    Takes a template name, fills it in from the template dictionary
    and then sends it.

    Attachments are currently handled for string representations of
    file objects as a dictionary of {'filename': contents}. This is because
    we're typically attaching virtual files made on the fly. The contents
    object must have a read method.
    """

    db = current.db
    response = current.response
    mail = current.mail

    # get the html version, strip it down to text and combine
    html_msg = response.render('email_templates/' + template, template_dict)
    txt_msg = html2text.html2text(html_msg)
    msg = (txt_msg, html_msg)

    # add the info address into all emails (unless told explicitly not to)
    if cc_info:
        if cc is None:
            cc = ['info@safeproject.net']
        else:
            cc.append('info@safeproject.net')

    # handle string attachments by wrapping them in StringIO (which
    # has a read() method)
    if attachment_string_objects is not None:
        attach = [mail.Attachment(payload=cStringIO.StringIO(v), filename=k)
                  for k, v in attachment_string_objects.iteritems()]
    else:
        attach = None

    # send the mail - the mail.send method actually extends the list of
    # addresses in 'to' by reference to include cc and bcc, so pass it as
    # a sliced copy, to avoid this
    msg_status = mail.send(to=to[:], subject=subject, message=msg,
                           cc=cc, bcc=bcc, reply_to=reply_to,
                           attachments=attach)

    # log it in the database
    # first - Sanitise the template dictionary. There are cases
    # where we want to pass gluon objects to email templates. That's
    # fine, but then that object can't be serialized into JSON. So,
    # using the easier to ask forgiveness principle run everything
    # through the xml() method and catch objects that don't have one.

    for k, v in template_dict.iteritems():
        try:
            template_dict[k] = v.xml()
        except AttributeError:
            pass

    db.safe_web_email_log.insert(email_to=to,
                                 subject=subject,
                                 template=template,
                                 template_dict=simplejson.dumps(template_dict),
                                 email_cc=cc,
                                 email_bcc=bcc,
                                 reply_to=reply_to,
                                 status='sent' if msg_status else 'failed',
                                 message_date=datetime.datetime.now())

    return msg_status


"""
The functions below provide files containing research visit summaries.
"""


def uname(uid, rowid):

    """
    The RV_member table contains a link to a user that can be NULL - this helper
    function takes a user_id field (which references auth_user) and a RVM row id
    to give a formatted name for both named and unknown users
    """

    if uid is None:
        nm = 'Unknown #' + str(rowid)
    else:
        nm = uid.last_name + ", " + uid.first_name

    return nm


def date_range(start, end):
    """
    Provides a set of datetime objects for all days between two end points,
    including those endpoints

    :param start:
    :param end:
    :return:
    """
    days = []
    curr = start
    while curr <= end:
        days.append(curr)
        curr += datetime.timedelta(days=1)

    return days


def single_rv_summary_excel(rv_id):

    """
    This creates an excel workbook containing the intinerary for a single research
    visit as a virtual workbook object. Costs are loaded from the same costs
    json used to populate the logistics and costs page, so can be updated from
    a single location
    """

    db = current.db
    request = current.request

    # load costs from the json data
    f = os.path.join(request.folder, 'private', 'content/en/info/costs.json')
    costs_dict = simplejson.load(open(f))

    # set up the coordinates of the data block
    curr_row = start_row = 8
    data_start_col = 4

    # load record
    record = db.research_visit(rv_id)

    # GET THE TIMESCALE FOR THE RVs (which _should_ encompass all RV activities)

    start_all = record.arrival_date
    end_all = record.departure_date

    # use start as an epoch to give column numbers
    dates = date_range(start_all, end_all)
    dates_column = [(x - start_all).days + data_start_col for x in dates]

    # get column labels and get blocks to label months
    monthyear = [d.strftime('%B %Y') for d in dates]
    monthyear_set = set(monthyear)
    monthyear_dates = [[i for i, x in zip(dates, monthyear) if x == my] for my in monthyear_set]
    monthyear_range = [[min(x), max(x)] for x in monthyear_dates]
    weekday = [d.strftime('%a') for d in dates]
    weekday = [x[0] for x in weekday]
    day = [d.day for d in dates]

    # SETUP THE WORKBOOK
    wb = Workbook()

    # spreadsheet styles
    left = styles.Alignment(horizontal='left')
    center = styles.Alignment(horizontal='center')
    weekend = styles.PatternFill(fill_type='solid', start_color='CCCCCC')
    head = styles.Font(size=14, bold=True)
    subhead = styles.Font(bold=True)
    warn = styles.Font(bold=True, color='FF0000')
    cell_shade = {'Approved': styles.PatternFill(fill_type='solid', start_color='8CFF88'),
                  'Submitted': styles.PatternFill(fill_type='solid', start_color='FFCB8B'),
                  'Draft': styles.PatternFill(fill_type='solid', start_color='DDDDDD'),
                  'Resubmit': styles.PatternFill(fill_type='solid', start_color='DDDDDD'),
                  'Rejected': styles.PatternFill(fill_type='solid', start_color='FF96C3')}

    # name the worksheet and add a heading
    ws = wb.active
    ws.title = 'Research visits'
    ws['A1'] = 'Research visit plans for the SAFE Project as of {}'.format(datetime.date.today().isoformat())
    ws['A1'].font = head

    title = db((db.project_id.id == record.project_id) &
               (db.project_id.project_details_id == db.project_details.id))
    title = title.select(db.project_details.title).first().title
    ws['A2'] = 'Research visit #' + str(record.id) + ": " + title
    ws['A2'].font = subhead

    ws['A3'] = ('Costs are estimated and do not include site transport costs at SAFE. Accomodation '
                'costs at SAFE are estimated assuming all visitors are international and will be '
                'cheaper if your visit includes Malaysian staff or students')
    ws['A3'].font = warn

    # Populate the sheet with the date information
    # months in merged cells with shading on alternate months
    for start, end in monthyear_range:
        start_col = (start - start_all).days + data_start_col
        end_col = (end - start_all).days + data_start_col

        ws.merge_cells(start_row=5, start_column=start_col,
                       end_row=5, end_column=end_col)
        c = ws.cell(row=5, column=start_col)
        c.value = start.strftime('%B %Y')
        if (start.month % 2) == 1:
            c.fill = weekend

    # day of month and week day, colouring weekends and setting column width
    for i, d, w in zip(dates_column, day, weekday):
        c1 = ws.cell(row=6, column=i)
        c1.value = d
        c1.alignment = center

        c2 = ws.cell(row=7, column=i)
        c2.value = w
        c2.alignment = center

        if w == 'S':
            c1.fill = weekend
            c2.fill = weekend

        ws.column_dimensions[utils.get_column_letter(i)].width = 3

    # add left hand column headers
    ws['A6'] = 'ID'
    ws['B6'] = 'Type'
    ws['C6'] = 'Cost in RM'

    def write_event(row, start, end, id, type, content, status, cost):
        """
        Write event function. Don't use merged cells here, although it seems tempting, as text
        cannot overflow from merged cell blocks into the adjacent blank cells, but it can from
        a single cell, so write content into first cell of range and then just colour the rest
        of the date range.

        Note that this operates directly on the worksheet object.

        :param row:
        :param start:
        :param end:
        :param id:
        :param type:
        :param content:
        :param status:
        :param cost:
        :return:
        """
        # range
        this_start_col = (start - start_all).days + data_start_col
        this_end_col = (end - start_all).days + data_start_col

        # put content at LHS of range and align left to make text overflow RHS of narrow ranges.
        ev_cell = ws.cell(row=row, column=this_start_col)
        ev_cell.value = content
        ev_cell.alignment = left

        for ev_col in range(this_start_col, this_end_col + 1):
            ev_cell = ws.cell(row=row, column=ev_col)
            ev_cell.fill = cell_shade[status]

        ev_cell = ws.cell(row=row, column=1)
        ev_cell.value = id
        ev_cell = ws.cell(row=row, column=2)
        ev_cell.value = type
        ev_cell = ws.cell(row=row, column=3)
        ev_cell.value = cost

    cost = '---'

    dat = [curr_row, record.arrival_date, record.departure_date, record.id, 'Research Visit',
           "Project " + str(record.project_id) + ": " + record.title, record.admin_status, cost]

    write_event(*dat)
    curr_row += 1

    # SAFE bed bookings
    safe_query = ((db.bed_reservations_safe.research_visit_id == record.id) &
                  (db.bed_reservations_safe.research_visit_member_id ==
                   db.research_visit_member.id))

    safe_data = db(safe_query).select(orderby=db.bed_reservations_safe.arrival_date)

    for r in safe_data:

        # check for unknown users
        name = uname(r.research_visit_member.user_id, r.research_visit_member.id)

        # calculate cost - food charge only
        cost = (r.bed_reservations_safe.departure_date -
                r.bed_reservations_safe.arrival_date).days * costs_dict['safe_costs']['food']['cost']

        # put the list of info to be written together
        dat = [curr_row, r.bed_reservations_safe.arrival_date,
               r.bed_reservations_safe.departure_date,
               r.bed_reservations_safe.research_visit_id,
               'SAFE booking', name, record.admin_status, cost]

        # write it and move down a row
        write_event(*dat)
        curr_row += 1

    # MALIAU bed bookings
    maliau_query = ((db.bed_reservations_maliau.research_visit_id == record.id) &
                    (db.bed_reservations_maliau.research_visit_member_id ==
                     db.research_visit_member.id))
    maliau_data = db(maliau_query).select(orderby=db.bed_reservations_maliau.arrival_date)

    for r in maliau_data:

        name = uname(r.research_visit_member.user_id, r.research_visit_member.id)

        # calculate cost - entry, bed and food costs
        days = (r.bed_reservations_maliau.departure_date -
                r.bed_reservations_maliau.arrival_date).days
        # admin and conservation on entry
        cost = costs_dict['maliau_entry']['admin']['standard'] + costs_dict['maliau_entry']['cons']['standard']
        # annex/hostel are only alternatives at the moment
        if r.bed_reservations_maliau.type == 'Annex':
            cost += days * costs_dict['maliau_accom']['annex']['standard']
        else:
            cost += days * costs_dict['maliau_accom']['hostel']['standard']
        # food
        if r.bed_reservations_maliau.breakfast is True:
            cost += costs_dict['maliau_food']['breakfast']['standard'] * days
        if r.bed_reservations_maliau.lunch is True:
            cost += costs_dict['maliau_food']['lunch']['standard'] * days
        if r.bed_reservations_maliau.dinner is True:
            cost += costs_dict['maliau_food']['dinner']['standard'] * days

        # content
        food_labels = ['B' if r.bed_reservations_maliau.breakfast else ''] + \
                      ['L' if r.bed_reservations_maliau.lunch else ''] + \
                      ['D' if r.bed_reservations_maliau.dinner else '']
        content = name + ' (' + r.bed_reservations_maliau.type + ',' + ''.join(food_labels) + ')'
        dat = [curr_row, r.bed_reservations_maliau.arrival_date,
               r.bed_reservations_maliau.departure_date,
               r.bed_reservations_maliau.research_visit_id,
               'Maliau booking', content, record.admin_status, cost]

        write_event(*dat)
        curr_row += 1

    # TRANSFERS
    transfer_query = ((db.transfers.research_visit_id == record.id) &
                      (db.transfers.research_visit_member_id == db.research_visit_member.id))
    transfer_data = db(transfer_query).select(orderby=db.transfers.transfer_date)

    for r in transfer_data:

        name = uname(r.research_visit_member.user_id, r.research_visit_member.id)

        # costs
        if r.transfers.transfer in ['SAFE to Tawau', 'Tawau to SAFE']:
            cost = costs_dict['transfers']['tawau_safe']['cost']
        elif r.transfers.transfer in ['SAFE to Maliau', 'Maliau to SAFE']:
            cost = costs_dict['transfers']['safe_maliau']['cost']
        elif r.transfers.transfer in ['Tawau to Maliau', 'Maliau to Tawau']:
            cost = costs_dict['transfers']['tawau_maliau']['cost']
        elif r.transfers.transfer in ['SAFE to Danum', 'Danum to SAFE']:
            cost = costs_dict['transfers']['safe_danum']['cost']

        dat = [curr_row, r.transfers.transfer_date, r.transfers.transfer_date,
               r.transfers.research_visit_id, 'Transfer',
               name + ': ' + r.transfers.transfer, record.admin_status, cost]

        write_event(*dat)
        curr_row += 1

    # RA requests
    rassist_query = (db.research_assistant_bookings.research_visit_id == record.id)
    rassist_data = db(rassist_query).select(orderby=db.research_assistant_bookings.start_date)

    for r in rassist_data:

        # costs
        if r.site_time in ['All day at SAFE', 'All day at Maliau']:
            cost = costs_dict['ra_costs']['full']
        else:
            cost = costs_dict['ra_costs']['half']

        cost = cost[r.work_type]

        dat = [curr_row, r.start_date, r.finish_date, r.research_visit_id, 'RA booking',
               r.site_time, record.admin_status, cost]

        write_event(*dat)
        curr_row += 1

    # freeze the rows
    c = ws.cell(row=start_row, column=data_start_col)
    ws.freeze_panes = c

    # and now return the workbook object as a string
    try:
        return str(writer.excel.save_virtual_workbook(wb))
    finally:
        del wb


class SummaryTracker:
    """
    This class simply keeps track of the number of people requesting
    different resources across a time range. The update method allows
    this tracking to be incorporated into the individual row loops when
    Excel or text reports are being created.
    """

    def __init__(self, start, end):

        # Initialise the various tracked components
        dates = date_range(start, end)
        self.summary = {}
        
        for d in dates:
            self.summary[d] = {'include': False,
                               'include_safe': False,
                               'safe': {'beds':  0,
                                        'diets': [],
                                        'ras': 0},
                               'include_maliau': False,                                
                               'maliau': {'beds':  0,
                                          'diets': [],
                                          'ras': 0}}

            # add all the transfer types
            self.summary[d]['include_transfers'] = False
            self.summary[d]['transfers'] = {}
            for t in current.transfer_set:
                self.summary[d]['transfers'][t] = 0

    def update(self, block, item, start, end, val=None):

        """
        Update the detail for a given key, checking that the
        date provided is included in the tracker date span. This
        can happen when research visits are updated to use new
        arrival and departure dates, leaving old bookings.
        """
        new_dates = date_range(start, end)
        
        for d in new_dates:
            if d in self.summary.keys():
                self.summary[d]['include'] = True
                self.summary[d]['include_' + block] = True
                if item == 'diets':
                    self.summary[d][block]['diets'].append(val)
                else:
                    self.summary[d][block][item] += 1


def all_rv_summary_excel():

    """
    This creates an excel workbook compiling ongoing and future research visit data
    and returns it as virtual file object. Costs are loaded from the same costs
    json used to populate the logistics and costs page, so can be updated from
    a single location
    """

    db = current.db
    request = current.request

    # load costs from the json data
    f = os.path.join(request.folder, 'private', 'content/en/info/costs.json')
    costs_dict = simplejson.load(open(f))

    # set up the coordinates of the data block
    curr_row = start_row = 20
    data_start_col = 4
    food_start_row = 3
    
    # GET THE ONGOING RVs
    today = datetime.date.today()
    rv_query = (db.research_visit.departure_date >= today)

    # no records?
    if db(rv_query).count() == 0:
        return
    else:
        # grab the data from those queries starting with the earliest arrivals
        rv_data = db(rv_query).select(orderby=db.research_visit.arrival_date)

    # GET A COMMON TIME SCALE FROM THE RVs (which _should_ encompass all RV activities)
    start_all = min([r.arrival_date for r in rv_data])
    end_all = max([r.departure_date for r in rv_data])

    # use start as an epoch to give column numbers
    dates = date_range(start_all, end_all)
    dates_column = [(x - start_all).days + data_start_col for x in dates]

    # get column labels and get blocks to label months
    monthyear = [d.strftime('%B %Y') for d in dates]
    monthyear_set = set(monthyear)
    monthyear_dates = [[i for i, x in zip(dates, monthyear) if x == my] for my in monthyear_set]
    monthyear_range = [[min(x), max(x)] for x in monthyear_dates]
    weekday = [d.strftime('%a') for d in dates]
    weekday = [x[0] for x in weekday]
    day = [d.day for d in dates]

    # SETUP THE WORKBOOK
    wb = Workbook()

    # spreadsheet styles
    left = styles.Alignment(horizontal='left')
    center = styles.Alignment(horizontal='center')
    weekend = styles.PatternFill(fill_type='solid', start_color='CCCCCC')
    head = styles.Font(size=14, bold=True)
    subhead = styles.Font(bold=True)
    warn = styles.Font(bold=True, color='FF0000')
    zero_summary = styles.Font(color='BBBBBB')
    non_zero_summary = styles.Font(bold=True)

    cell_shade = {'Approved': styles.PatternFill(fill_type='solid', start_color='8CFF88'),
                  'Submitted': styles.PatternFill(fill_type='solid', start_color='FFCB8B'),
                  'Draft': styles.PatternFill(fill_type='solid', start_color='DDDDDD'),
                  'Resubmit': styles.PatternFill(fill_type='solid', start_color='DDDDDD'),
                  'Rejected': styles.PatternFill(fill_type='solid', start_color='FF96C3')}

    # name the worksheet and add a heading
    ws = wb.active
    ws.title = 'Research visits'
    ws['A1'] = 'Research visit plans for the SAFE Project as of {}'.format(today.isoformat())
    ws['A1'].font = head

    # Subheading
    ws['A2'] = 'All research visits'
    ws['A2'].font = subhead

    ws['A3'] = ('Costs are estimated and do not include site transport costs at SAFE. Accomodation '
                'costs at SAFE are estimated assuming all visitors are international and will be '
                'cheaper if your visit includes Malaysian staff or students')
    ws['A3'].font = warn

    # Add a food summary worksheet - this info doesn't fit neatly into the horizontal layout
    food = wb.create_sheet('Food summary')
    food['A1'] = 'Summary of dietary restrictions for booked beds'
    food['A1'].font = head
    
    food_heads = ['Month', 'Date', 'Day', 
                  'SAFE beds', 'SAFE dietary restrictions',
                  'Maliau beds', 'Maliau dietary restrictions']
                  
    for idx, food_col in enumerate(food_heads):
        f = food.cell(row=food_start_row - 1, column=idx + 1)
        f.value = food_col
        f.font = subhead
    
    # Populate the sheet with the date information
    # months in merged cells with shading on alternate months
    for start, end in monthyear_range:
        start_col = (start - start_all).days + data_start_col
        end_col = (end - start_all).days + data_start_col

        ws.merge_cells(start_row=5, start_column=start_col,
                       end_row=5, end_column=end_col)
        c = ws.cell(row=5, column=start_col)
        c.value = start.strftime('%B %Y')


        food.merge_cells(start_row=start_col - data_start_col + food_start_row, start_column=1,
                         end_row=end_col - data_start_col + food_start_row, end_column=1)
        f = food.cell(row=start_col - data_start_col + food_start_row, column=1)
        f.value = start.strftime('%B %Y')
        f.alignment = styles.Alignment(vertical='top')

        if (start.month % 2) == 1:
            c.fill = weekend
            f.fill = weekend
        
    # day of month and week day, colouring weekends and setting column width
    for i, d, w in zip(dates_column, day, weekday):
        c1 = ws.cell(row=6, column=i)
        c1.value = d
        c1.alignment = center
        
        f1 = food.cell(row=i + food_start_row - data_start_col, column=2)
        f1.value = d

        c2 = ws.cell(row=7, column=i)
        c2.value = w
        c2.alignment = center
        
        f2 = food.cell(row=i + food_start_row - data_start_col, column=3)
        f2.value = w

        if w == 'S':
            c1.fill = weekend
            c2.fill = weekend
            f1.fill = weekend
            f2.fill = weekend

        ws.column_dimensions[utils.get_column_letter(i)].width = 3

    # add left hand column headers
    ws['A6'] = 'ID'
    ws['B6'] = 'Type'
    ws['C6'] = 'Cost in RM'

    # write event function - don't use merged cells here
    # it seems tempting but text cannot overflow from merged
    # cell blocks into the adjacent blank cells, but it can from
    # a single cell, so write content into first cell of range
    # and then just colour the rest of the date range.
    def write_event(row, start, end, id, type, content, status, cost):

        # range
        this_start_col = (start - start_all).days + data_start_col
        this_end_col = (end - start_all).days + data_start_col

        # put content at LHS of range
        c = ws.cell(row=row, column=this_start_col)
        c.value = content
        c.alignment = left # makes text overflow RHS of narrow ranges.

        for i in range(this_start_col, this_end_col + 1):
            c = ws.cell(row=row, column=i)
            c.fill = cell_shade[status]

        c = ws.cell(row=row, column=1)
        c.value = id
        c = ws.cell(row=row, column=2)
        c.value = type
        c = ws.cell(row=row, column=3)
        c.value = cost

    # initialise a summary tracker using the full date range
    summary = SummaryTracker(start_all, end_all)

    # loop over the research visits.
    for v in rv_data:

        cost = '---'

        dat = [curr_row, v.arrival_date, v.departure_date, v.id, 'Research Visit',
               "Project " + str(v.project_id) + ": " + v.title, v.admin_status, cost]

        write_event(*dat)
        curr_row += 1

        # SAFE bed bookings
        safe_query = ((db.bed_reservations_safe.research_visit_id == v.id) &
                      (db.bed_reservations_safe.research_visit_member_id == db.research_visit_member.id) &
                      (db.research_visit_member.user_id == db.health_and_safety.user_id))
        
        safe_data = db(safe_query).select(orderby=db.bed_reservations_safe.arrival_date)

        for r in safe_data:

            # check for unknown users
            name = uname(r.research_visit_member.user_id, r.research_visit_member.id)

            # calculate cost - food charge only
            cost = (r.bed_reservations_safe.departure_date -
                    r.bed_reservations_safe.arrival_date).days * costs_dict['safe_costs']['food']['cost']

            # put the list of info to be written together
            content = '{} (Dietary restrictions: {})\n'.format(
                                name,
                                r.health_and_safety.dietary_requirements)
            
            dat = [curr_row, r.bed_reservations_safe.arrival_date,
                   r.bed_reservations_safe.departure_date,
                   r.bed_reservations_safe.research_visit_id,
                   'SAFE booking', content, v.admin_status, cost]

            # write it, update the summary tracker and move down a row
            write_event(*dat)
            
            summary.update('safe','beds',
                           r.bed_reservations_safe.arrival_date,
                           r.bed_reservations_safe.departure_date - datetime.timedelta(days=1))
            
            summary.update('safe', 'diets',
                           r.bed_reservations_safe.arrival_date,
                           r.bed_reservations_safe.departure_date - datetime.timedelta(days=1),
                           r.health_and_safety.dietary_requirements)
            
            curr_row += 1

        # MALIAU bed bookings
        maliau_query = ((db.bed_reservations_maliau.research_visit_id == v.id) &
                        (db.bed_reservations_maliau.research_visit_member_id == db.research_visit_member.id) &
                        (db.research_visit_member.user_id == db.health_and_safety.user_id))
        
        maliau_data = db(maliau_query).select(orderby=db.bed_reservations_maliau.arrival_date)

        for r in maliau_data:

            name = uname(r.research_visit_member.user_id, r.research_visit_member.id)

            # calculate cost - entry, bed and food costs
            days = (r.bed_reservations_maliau.departure_date -
                    r.bed_reservations_maliau.arrival_date).days
            # admin and conservation on entry
            cost = costs_dict['maliau_entry']['admin']['standard'] + \
                   costs_dict['maliau_entry']['cons']['standard']
            # annex/hostel are only alternatives at the moment
            if r.bed_reservations_maliau.type == 'Annex':
                cost += days * costs_dict['maliau_accom']['annex']['standard']
            else:
                cost += days * costs_dict['maliau_accom']['hostel']['standard']
                
            # food
            if r.bed_reservations_maliau.breakfast is True:
                cost += costs_dict['maliau_food']['breakfast']['standard'] * days
            if r.bed_reservations_maliau.lunch is True:
                cost += costs_dict['maliau_food']['lunch']['standard'] * days
            if r.bed_reservations_maliau.dinner is True:
                cost += costs_dict['maliau_food']['dinner']['standard'] * days

            # content
            catering = ''.join(['B' if r.bed_reservations_maliau.breakfast else ''] + 
                               ['L' if r.bed_reservations_maliau.lunch else ''] + 
                               ['D' if r.bed_reservations_maliau.dinner else ''])
            if catering != "":
                catering = ' with catered ' + catering
            
            content = '{} in {} {} (Dietary restrictions: {})\n'.format(
                                name,
                                r.bed_reservations_maliau.type,
                                catering,
                                r.health_and_safety.dietary_requirements)
            
            dat = [curr_row, r.bed_reservations_maliau.arrival_date,
                   r.bed_reservations_maliau.departure_date,
                   r.bed_reservations_maliau.research_visit_id,
                   'Maliau booking', content, v.admin_status, cost]

            write_event(*dat)
            
            summary.update('maliau', 'beds',
                           r.bed_reservations_maliau.arrival_date,
                           r.bed_reservations_maliau.departure_date - datetime.timedelta(days=1))

            summary.update('maliau', 'diets',
                           r.bed_reservations_maliau.arrival_date,
                           r.bed_reservations_maliau.departure_date - datetime.timedelta(days=1),
                           r.health_and_safety.dietary_requirements)

            curr_row += 1

        # TRANSFERS
        transfer_query = ((db.transfers.research_visit_id == v.id) &
                          (db.transfers.research_visit_member_id == db.research_visit_member.id))
        transfer_data = db(transfer_query).select(orderby=db.transfers.transfer_date)

        for r in transfer_data:

            name = uname(r.research_visit_member.user_id, r.research_visit_member.id)

            # costs
            if r.transfers.transfer in ['SAFE to Tawau', 'Tawau to SAFE']:
                cost = costs_dict['transfers']['tawau_safe']['cost']
            elif r.transfers.transfer in ['SAFE to Maliau', 'Maliau to SAFE']:
                cost = costs_dict['transfers']['safe_maliau']['cost']
            elif r.transfers.transfer in ['Tawau to Maliau', 'Maliau to Tawau']:
                cost = costs_dict['transfers']['tawau_maliau']['cost']
            elif r.transfers.transfer in ['SAFE to Danum', 'Danum to SAFE']:
                cost = costs_dict['transfers']['safe_danum']['cost']

            dat = [curr_row, r.transfers.transfer_date, r.transfers.transfer_date,
                   r.transfers.research_visit_id, 'Transfer',
                   name + ': ' + r.transfers.transfer, v.admin_status, cost]

            write_event(*dat)
            # update the transfers summary, keying by the transfer type
            summary.update('transfers', r.transfers.transfer,
                           r.transfers.transfer_date,
                           r.transfers.transfer_date)
            curr_row += 1

        # RA requests
        rassist_query = (db.research_assistant_bookings.research_visit_id == v.id)
        rassist_data = db(rassist_query).select(orderby=db.research_assistant_bookings.start_date)

        for r in rassist_data:

            # costs
            if r.site_time in ['All day at SAFE', 'All day at Maliau']:
                cost = costs_dict['ra_costs']['full']
            else:
                cost = costs_dict['ra_costs']['half']

            cost = cost[r.work_type]

            dat = [curr_row, r.start_date, r.finish_date, r.research_visit_id, 'RA booking',
                   r.site_time, v.admin_status, cost]

            write_event(*dat)

            if r.site_time in ['All day at SAFE', 'Morning only at SAFE', 'Afternoon only at SAFE']:
                summary.update('safe', 'ras', r.start_date, r.finish_date)
            else:
                summary.update('maliau', 'ras', r.start_date, r.finish_date)

            curr_row += 1

    # Insert the summary information
    summary_row = 8

    # loop over the summary keys
    summary_keys = [('safe', 'beds', 'Beds requested at SAFE'), 
                    ('safe', 'ras', 'RAs requested at SAFE'), 
                    ('maliau', 'beds', 'Beds requested at Maliau'), 
                    ('maliau', 'ras', 'RAs requested at Maliau')]
                    
    for t in current.transfer_set:
        summary_keys.append(('transfers', t, t))
    
    for blck, itm, lab in summary_keys:

        c = ws.cell(row=summary_row, column=2)
        c.value = lab

        for d, c in zip(dates, dates_column):
            c = ws.cell(row=summary_row, column=c)
            c.value = summary.summary[d][blck][itm]
            if c.value == 0:
                c.font = zero_summary
            else:
                c.font = non_zero_summary

        summary_row += 1

    # food summary
    food_keys = [('safe',  4), 
                 ('maliau',  6)]
    
    for blck, col in food_keys:

        for i, d in enumerate(dates):
            dat = summary.summary[d][blck]
            c = food.cell(row=food_start_row + i, column=col)
            c.value = dat['beds']
            
            if dat['diets']:
                diets = ['{} x {}'.format(v, k) 
                         for k, v in Counter(dat['diets']).iteritems()]
                
                c = food.cell(row=food_start_row + i, column=col + 1)
                c.value = ', '.join(diets)

    # freeze the rows
    c = ws.cell(row=start_row, column=data_start_col)
    ws.freeze_panes = c
        
    # return as a string
    try:
        return str(writer.excel.save_virtual_workbook(wb))
    finally:
        del wb


def all_rv_summary_text():

    """
    This creates a text file compiling ongoing and future research visit data.
    """

    db = current.db

    # GET THE ONGOING RVs
    today = datetime.date.today()
    rv_query = (db.research_visit.departure_date >= today)

    # no records?
    if db(rv_query).count() == 0:
        return
    else:
        # grab the data from those queries starting with the earliest arrivals
        rv_data = db(rv_query).select(orderby=db.research_visit.arrival_date)

    # get the date range for all the activities and create a summary tracker
    start_all = min([r.arrival_date for r in rv_data])
    end_all = max([r.departure_date for r in rv_data])
    summary = SummaryTracker(start_all, end_all)

    # SETUP THE TEXT FILE AND DETAILS FILE. Because the summary is populated
    # while the details are being written out, and because we want the summary
    # at the top, keep two streams and then merge

    output = cStringIO.StringIO()
    details = cStringIO.StringIO()
    output.write('Research visit plans for the SAFE Project as of {}\n\n'.format(today.isoformat()))

    # loop over the research visits.
    for v in rv_data:

        details.write("\nProject " + str(v.project_id) + ": " + v.title +
                      ' [Status: ' + v.admin_status + ']\n')

        # SAFE bed bookings
        safe_query = ((db.bed_reservations_safe.research_visit_id == v.id) &
                      (db.bed_reservations_safe.research_visit_member_id == db.research_visit_member.id) &
                      (db.research_visit_member.user_id == db.health_and_safety.user_id))
        safe_data = db(safe_query).select(orderby=db.bed_reservations_safe.arrival_date)

        for r in safe_data:

            # check for unknown users
            name = uname(r.research_visit_member.user_id, r.research_visit_member.id)
            
            details.write('     SAFE Booking: {} -- {} for {} (Dietary restrictions: {})\n'.format(
                                r.bed_reservations_safe.arrival_date.isoformat(),
                                r.bed_reservations_safe.departure_date.isoformat(),
                                name,
                                r.health_and_safety.dietary_requirements))
            
            summary.update('safe', 'beds',
                           r.bed_reservations_safe.arrival_date,
                           r.bed_reservations_safe.departure_date - datetime.timedelta(days=1))

            summary.update('safe', 'diets',
                           r.bed_reservations_safe.arrival_date,
                           r.bed_reservations_safe.departure_date - datetime.timedelta(days=1),
                           r.health_and_safety.dietary_requirements)


        # MALIAU bed bookings
        maliau_query = ((db.bed_reservations_maliau.research_visit_id == v.id) &
                        (db.bed_reservations_maliau.research_visit_member_id == db.research_visit_member.id) &
                        (db.research_visit_member.user_id == db.health_and_safety.user_id))
        
        maliau_data = db(maliau_query).select(orderby=db.bed_reservations_maliau.arrival_date)

        for r in maliau_data:

            name = uname(r.research_visit_member.user_id, r.research_visit_member.id)

            # content
            catering = ''.join(['B' if r.bed_reservations_maliau.breakfast else ''] + 
                               ['L' if r.bed_reservations_maliau.lunch else ''] + 
                               ['D' if r.bed_reservations_maliau.dinner else ''])
            if catering != "":
                catering = ' with catered ' + catering
            
            details.write('     Maliau Booking: {} -- {} for {} in {} {} (Diet: {})\n'.format(
                                r.bed_reservations_maliau.arrival_date.isoformat(),
                                r.bed_reservations_maliau.departure_date.isoformat(),
                                name,
                                r.bed_reservations_maliau.type,
                                catering,
                                r.health_and_safety.dietary_requirements))
                        
            summary.update('maliau', 'beds',
                           r.bed_reservations_maliau.arrival_date,
                           r.bed_reservations_maliau.departure_date - datetime.timedelta(days=1))

            summary.update('maliau','diets',
                           r.bed_reservations_maliau.arrival_date,
                           r.bed_reservations_maliau.departure_date - datetime.timedelta(days=1),
                           r.health_and_safety.dietary_requirements)

        # TRANSFERS
        transfer_query = ((db.transfers.research_visit_id == v.id) &
                          (db.transfers.research_visit_member_id == db.research_visit_member.id))
        transfer_data = db(transfer_query).select(orderby=db.transfers.transfer_date)

        for r in transfer_data:

            name = uname(r.research_visit_member.user_id, r.research_visit_member.id)
            details.write('     Transfer: ' + r.transfers.transfer_date.isoformat() + ' from ' +
                         r.transfers.transfer + ' for ' + name + '\n')
            # update the transfers summary, keying by the transfer type
            summary.update('transfers', r.transfers.transfer,
                           r.transfers.transfer_date,
                           r.transfers.transfer_date)

        # RA requests
        rassist_query = (db.research_assistant_bookings.research_visit_id == v.id)
        rassist_data = db(rassist_query).select(orderby=db.research_assistant_bookings.start_date)

        for r in rassist_data:
            details.write('     RA time: ' + r.start_date.isoformat() + ' -- ' +
                          r.finish_date.isoformat() +
                          ' (' + r.site_time + ', ' + r.work_type + ')\n')

            if r.site_time in ['All day at SAFE', 'Morning only at SAFE', 'Afternoon only at SAFE']:
                summary.update('safe', 'ras', r.start_date, r.finish_date)
            else:
                summary.update('maliau','ras', r.start_date, r.finish_date)

    # now add the summary to the output and then transfer the details at the bottom
    output.write('Daily resource request totals for current projects\n'
                 '==================================================\n\n')

    dates = date_range(today, end_all)
    for d in dates:
        
        # extract and format the contents of the SummaryTracker
        if summary.summary[d]['include']:
            output.write(d.strftime('%a %d %b %Y') + '\n')
            
        if summary.summary[d]['include_safe']:
            output.write('    SAFE\n')
            
            if summary.summary[d]['safe']['beds'] > 0:
                output.write('      - Beds requested: ' +  str(summary.summary[d]['safe']['beds'])  + '\n')
                diets = ['{} x {}'.format(v, k) 
                         for k, v in Counter(summary.summary[d]['safe']['diets']).iteritems()]
                output.write('      - Dietary requirements: ' + ', '.join(diets) + '\n')
                
            if summary.summary[d]['safe']['ras'] > 0:
                output.write('      - RAs requested: ' +  str(summary.summary[d]['safe']['ras'])  + '\n')
            
        if summary.summary[d]['include_maliau']:
            output.write('    Maliau\n')
            
            if summary.summary[d]['maliau']['beds'] > 0:
                output.write('      - Beds requested: ' +  str(summary.summary[d]['maliau']['beds'])  + '\n')
                diets = ['{} x {}'.format(v, k) 
                         for k, v in Counter(summary.summary[d]['maliau']['diets']).iteritems()]
                output.write('      - Dietary restrictions: ' + ', '.join(diets) + '\n')
                
            if summary.summary[d]['maliau']['ras'] > 0:
                output.write('      - RAs requested: ' +  str(summary.summary[d]['maliau']['ras'])  + '\n')
        
        if summary.summary[d]['include_transfers']:
            output.write('    Transfers\n')
            for k, v in summary.summary[d]['transfers'].iteritems():
                if v > 0:
                    output.write('      - ' + k + ': ' + str(v) + '\n')
            
    # now add the details at the end
    output.write('\n\nProject by project details\n'
                 '==========================\n\n')
    output.write(details.getvalue())
    details.close()

    try:
        return output.getvalue().replace('\n', '\r\n')
    finally:
        output.close()


def health_and_safety_report():
    
    """
    Function to generate a report containing H&S information for
    upcoming visitors.
    """

    # get some time stamps
    now = datetime.datetime.now()
    old_hs = (now - datetime.timedelta(days=180)).date()
    
    # get a list of upcoming visitors to SAFE and Maliau and their H&S, if any, using a left
    # join to ensure that all visitors are included, with None in H&S if missing
    db = current.db
    
    # Find people with bed reservations at SAFE or Maliau. There may be a way of combining 
    # these two but it seems likely to be slower
    safe_vis = db((db.bed_reservations_safe.arrival_date - now <=  14) &
                  (db.bed_reservations_safe.departure_date >= now) &
                  (db.bed_reservations_safe.research_visit_member_id == db.research_visit_member.id) &
                  (db.research_visit_member.user_id == db.auth_user.id)
                  ).select(db.auth_user.ALL, 
                           db.health_and_safety.ALL, 
                           left=db.health_and_safety.on(db.auth_user.id == db.health_and_safety.user_id),
                           distinct=True,
                           orderby=(db.auth_user.last_name, db.auth_user.first_name))

    mali_vis = db((db.bed_reservations_maliau.arrival_date - now <=  14) &
                  (db.bed_reservations_maliau.departure_date >= now) &
                  (db.bed_reservations_maliau.research_visit_member_id == db.research_visit_member.id) &
                  (db.research_visit_member.user_id == db.auth_user.id)
                  ).select(db.auth_user.ALL, 
                           db.health_and_safety.ALL, 
                           left=db.health_and_safety.on(db.auth_user.id == db.health_and_safety.user_id),
                           distinct=True,
                           orderby=(db.auth_user.last_name, db.auth_user.first_name))
    
    rv = safe_vis + mali_vis
    
    # set of fields to report
    hs_fields = ['date_last_edited', 'passport_number', 'emergency_contact_name',
    			 'emergency_contact_address', 'emergency_contact_phone',
    			 'emergency_contact_email', 'insurance_company',
    			 'insurance_emergency_phone', 'insurance_policy_number',
    			 'medical_conditions', 'dietary_requirements']

    # Subclass FPDF to set the header
    class PDF(FPDF.FPDF):
        def header(self):
            self.set_font('Arial', 'B', 15)
            self.set_text_color(r=255, g=12, b=12)
            self.multi_cell(w=0, h=10, txt='SAFE Project Health and Safety information \n' +
    			 	       'Confidential. Report created: ' + now.date().isoformat(), border=1, align='C')
            self.ln(5)

    # Create the pdf with links from the cover page of names to individual pages
    pdf = PDF()
    # pdf.set_text_color(r=0, g=0, b=0)
    # Add the medical summary page
    pdf.add_page(orientation='P', format='A4')
    pdf.set_font('Arial', 'B', 15)
    pdf.cell(w=0, h=10, txt='Medical summary', border=1, align='C', ln=1)
    pdf.set_font('Arial', '', 12)
    
    for visitor in rv:
        
        if visitor.health_and_safety.id is not None:
            medic = visitor.health_and_safety.medical_conditions
            if (medic is not None) and (medic.strip() != '') and (medic.lower() not in ['n/a', 'na', 'none']):
                pdf.cell(h=10, w=70, txt='{last_name}, {first_name}'.format(**visitor.auth_user))
                pdf.cell(h=10, w=0, txt=visitor.health_and_safety.medical_conditions, ln=1)


    # Add the dietary requirements summary page
    pdf.add_page(orientation='P', format='A4')
    pdf.set_font('Arial', 'B', 15)
    pdf.cell(w=0, h=10, txt='Dietary requirements summary', border=1, align='C', ln=1)
    pdf.set_font('Arial', '', 12)
    
    for visitor in rv:
        
        if visitor.health_and_safety.id is not None:
            diet = visitor.health_and_safety.dietary_requirements
            if (diet is not None) and (diet.strip() != '') and (diet.lower() not in ['n/a', 'na', 'none']):
                pdf.cell(h=10, w=70, txt='{last_name}, {first_name}'.format(**visitor.auth_user))
                pdf.cell(h=10, w=0, txt=visitor.health_and_safety.medical_conditions, ln=1)

    links={}
    
    # write the index of the full pages and status
    pdf.add_page(orientation='P', format='A4')    
    pdf.set_font('Arial', 'B', 15)
    pdf.cell(w=0, h=10, txt='Person index and info status', border=1, align='C', ln=1)
    pdf.set_font('Arial', 'B', 12)

    pdf.cell(h=10, w=70, txt='Name (click to go to full details)')
    pdf.cell(h=10, w=0, txt='H&S information status', ln=1)
    pdf.set_font('Arial', '', 12)
    
    for visitor in rv:
    
        links[visitor.auth_user.id] = pdf.add_link()

        if visitor.health_and_safety.id is None:
            visitor.auth_user.info = "None found"
        elif visitor.health_and_safety.date_last_edited < old_hs:
            visitor.auth_user.info = "Outdated (> 6 months)"
        else:
            visitor.auth_user.info = "Up to date (< 6 months)"
    
        pdf.cell(h=10, w=70, 
    			 txt='{last_name}, {first_name}'.format(**visitor.auth_user),
                 link = links[visitor.auth_user.id])
			 
        pdf.cell(h=10, w=0, txt=visitor.auth_user.info, ln=1)
    

    
	# Add visitor pages
    for visitor in rv:
        pdf.add_page(orientation='P', format='A4')
        pdf.set_font('Arial', 'B', 15)
        pdf.write(h=10, txt='{last_name}, {first_name}: {info}\n'.format(**visitor.auth_user))
   
        pdf.set_link(links[visitor.auth_user.id])

        for fld in hs_fields:
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(h=10, w=70, txt = fld)
            pdf.set_font('Arial', '', 12)
            pdf.multi_cell(h=10, w=0, txt = str(visitor.health_and_safety[fld]))
        
    # this property is needed to place links but isn't defined in the code
    pdf.orientation_changes = []
    pdf.close()
    
    # output to an IO
    return pdf.output(dest='S')
